#!/usr/bin/env python3
"""
Disclosure Report K.303 Validator (Migdal Variant)

Validates K.303 disclosure reports and produces an Excel report with:
- Summary sheet
- Check specifications details
- Check status summary
- Per-check exception sheets

Inputs:
  - Mutual Funds List CSV: filtered to Mizrahi trustee funds
  - Current month disclosure report CSV
  - Previous month disclosure report CSV
  - K.303 checklist CSV (for פירוט בדיקות sheet)

Outputs:
  - XLSX workbook with validation results

Example:
  python disclosure_k303_validator.py \
    --mutual-funds-list "Mutual_Funds_List.xlsx - Worksheet.csv" \
    --current-report "disclosure_migdal_current_month.csv" \
    --previous-report "disclosure_migdal_previous_month.xlsx - Sheet1.csv" \
    --spec-file "בדיקת דוח גילוי נאות ק.303.xlsx - גיליון1.csv" \
    --output-xlsx "disclosure_validation_output.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import logging
import sys
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Global log directory for current run
LOG_RUN_DIR: Optional[Path] = None

# Loggers
logger = logging.getLogger(__name__)
logger_chk1a = logging.getLogger('CHK_1A')
logger_chk1b = logging.getLogger('CHK_1B')
logger_chk2a = logging.getLogger('CHK_2A')
logger_chk2b = logging.getLogger('CHK_2B')
logger_chk3 = logging.getLogger('CHK_3')


def setup_logging(log_base_dir: Path = Path("log")) -> Path:
    """Set up logging with separate files for each check."""
    global LOG_RUN_DIR

    from datetime import datetime as datetime_module
    timestamp = datetime_module.now().strftime("%Y%m%d_%H%M%S")
    short_uuid = str(uuid.uuid4())[:8]
    run_id = f"{timestamp}_{short_uuid}"

    run_dir = log_base_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    LOG_RUN_DIR = run_dir

    log_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    detailed_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(name)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(log_format)

    logger.setLevel(logging.DEBUG)
    logger.addHandler(console_handler)
    main_file_handler = logging.FileHandler(run_dir / "main.log", encoding='utf-8')
    main_file_handler.setLevel(logging.DEBUG)
    main_file_handler.setFormatter(log_format)
    logger.addHandler(main_file_handler)

    spec_loggers = [
        (logger_chk1a, "chk1a_fund_completeness.log", "CHK_1A - Fund Completeness"),
        (logger_chk1b, "chk1b_date_validity.log", "CHK_1B - Date Validity"),
        (logger_chk2a, "chk2a_prev_month.log", "CHK_2A - Previous Month Comparison"),
        (logger_chk2b, "chk2b_exposure_profile.log", "CHK_2B - Exposure Profile Check"),
        (logger_chk3, "chk3_combinations.log", "CHK_3 - Within-Month Combinations"),
    ]

    for spec_logger, filename, description in spec_loggers:
        spec_logger.setLevel(logging.DEBUG)
        file_handler = logging.FileHandler(run_dir / filename, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(detailed_format)
        spec_logger.addHandler(file_handler)
        spec_logger.addHandler(main_file_handler)
        spec_logger.info("=" * 70)
        spec_logger.info(description)
        spec_logger.info("=" * 70)

    logger.info("Log directory created: %s", run_dir)
    return run_dir


# -----------------------------
# Configuration
# -----------------------------

MIZRAHI_TRUSTEE_NAME = 'מזרחי טפחות חברה לנאמנות בע"מ'

# Disclosure report columns
D_COL_FUND_NO = "מספר קרן"
D_COL_FUND_NAME = "שם קרן"
D_COL_LEVEL_1 = "רמה 1"
D_COL_LEVEL_2 = "רמה 2"
D_COL_LEVEL_3 = "רמה 3"
D_COL_LEVEL_4 = "רמה 4"
D_COL_PERCENT = "%מקרן"
D_COL_EXTRA_DATA = "נתונים נוספים"
D_COL_REPORT_DATE = "תאריך דוח"
D_COL_RECORD_NO = "מס.רשומה"
D_COL_TOTAL_RECORDS = "סהכ רשומות"
D_COL_MANAGER_NO = "מס.מנהל ברשם"

# Mutual Funds List columns
MF_COL_FUND_ID = "מספר בורסה"
MF_COL_TRUSTEE = "שם נאמן"
MF_COL_MANAGER = "שם מנהל"
MF_COL_EXPOSURE_PROFILE = "פרופיל החשיפה"

# Excel styling
HEADER_FONT = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
DEFAULT_FONT = Font(name='Calibri', size=11)

WRAP_HEADERS = {
    "שם קרן",
    "סיבה",
    "הערות",
    "שדה",
    "ערך",
    "פירוט הבדיקה",
    "אופן ביצוע",
}

HEBREW_MONTHS = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר"
}

# Exposure profile mapping - determines equity/FX exposure from profile code
# Profile format: <digit><letter> e.g., "0A", "3B", "4D"
# First digit: max equity exposure (code 01)
# Letter: max FX exposure (code 06)
EQUITY_EXPOSURE_PROFILES = {
    '0': 0,    # 0% equity
    '1': 10,   # up to 10% equity
    '2': 30,   # up to 30% equity
    '3': 50,   # up to 50% equity
    '4': 120,  # up to 120% equity
    '5': 200,  # up to 200% equity
    '6': None, # above 200% (unlimited)
}

FX_EXPOSURE_PROFILES = {
    '0': 0,    # no FX exposure
    'A': 10,   # up to 10% FX
    'B': 30,   # up to 30% FX
    'C': 50,   # up to 50% FX
    'D': 120,  # up to 120% FX
    'E': 200,  # up to 200% FX
    'F': None, # above 200% (unlimited)
}


# -----------------------------
# Data structures
# -----------------------------

@dataclass
class DisclosureRow:
    """A single row from the disclosure report."""
    row_num: int
    fund_no: Optional[int]
    fund_name: Optional[str]
    level_1: Optional[str]
    level_2: Optional[str]
    level_3: Optional[str]
    level_4: Optional[str]
    percent_from_fund: Optional[float]
    extra_data: Optional[str]
    report_date: Optional[dt.date]
    record_no: Optional[int]
    total_records: Optional[int]
    manager_no: Optional[str]

    @property
    def effective_code(self) -> Optional[str]:
        """Get the most granular non-empty level code."""
        for level in [self.level_4, self.level_3, self.level_2, self.level_1]:
            if level and level.strip():
                return level.strip()
        return None


@dataclass
class MutualFund:
    """A fund from the mutual funds list."""
    fund_id: int
    fund_name: str
    trustee_name: str
    manager_name: str
    exposure_profile: Optional[str]


@dataclass
class ExceptionRow:
    """An exception found during validation."""
    check_id: str
    reason: str
    fund_no: Optional[int] = None
    fund_name: Optional[str] = None
    effective_code: Optional[str] = None
    percent_from_fund: Optional[float] = None
    report_date: Optional[dt.date] = None
    row_num: Optional[int] = None
    extra_info: dict = field(default_factory=dict)


# -----------------------------
# Helpers
# -----------------------------

# Regex pattern for illegal Excel XML characters
import re
_ILLEGAL_XML_CHARS_RE = re.compile(
    '[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


def _clean_excel_string(s: Optional[str]) -> Optional[str]:
    """Remove illegal XML characters that cause Excel file corruption."""
    if s is None:
        return None
    # Remove illegal XML characters
    return _ILLEGAL_XML_CHARS_RE.sub('', s)


def _norm_spaces(s: str) -> str:
    return " ".join((s or "").strip().split())


def _to_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    s = _clean_excel_string(s)  # Remove illegal XML characters
    return s if s else None


def _to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except Exception:
        return None


def _parse_ddmmyyyy(v: Any) -> Optional[dt.date]:
    """Parse date in DDMMYYYY format (e.g., 30112025 for Nov 30, 2025)."""
    if v is None or v == "":
        return None
    s = str(int(float(str(v).strip()))) if v else ""
    s = s.strip().zfill(8)
    try:
        day = int(s[0:2])
        month = int(s[2:4])
        year = int(s[4:8])
        return dt.date(year, month, day)
    except Exception:
        return None


def _format_report_month_hebrew(report_month: str) -> str:
    """Convert YYYY-MM to Hebrew format like 'נובמבר 2025'."""
    try:
        year, month = report_month.split("-")
        return f"{HEBREW_MONTHS[int(month)]} {year}"
    except (ValueError, KeyError):
        return report_month


def _fmt_date(d: Optional[dt.date]) -> str:
    return d.strftime("%d-%m-%Y") if d else ""


# -----------------------------
# Data loaders
# -----------------------------

def load_mutual_funds_csv(path: Path) -> dict[int, MutualFund]:
    """Load mutual funds list from CSV, return dict keyed by fund ID."""
    funds: dict[int, MutualFund] = {}

    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [h.strip() for h in reader.fieldnames]

        for row in reader:
            fund_id = _to_int(row.get(MF_COL_FUND_ID))
            if fund_id is None:
                continue

            fund = MutualFund(
                fund_id=fund_id,
                fund_name=_to_str(row.get("שם קרן בעברית")) or "",
                trustee_name=_to_str(row.get(MF_COL_TRUSTEE)) or "",
                manager_name=_to_str(row.get(MF_COL_MANAGER)) or "",
                exposure_profile=_to_str(row.get(MF_COL_EXPOSURE_PROFILE)),
            )
            funds[fund_id] = fund

    logger.info("Loaded %d funds from mutual funds list", len(funds))
    return funds


def load_disclosure_report_csv(path: Path) -> list[DisclosureRow]:
    """Load disclosure report from CSV."""
    rows: list[DisclosureRow] = []

    # Try different encodings
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1255', 'iso-8859-8', 'windows-1252']

    for encoding in encodings_to_try:
        try:
            with open(path, 'r', encoding=encoding) as f:
                reader = csv.DictReader(f)
                if reader.fieldnames:
                    # Strip whitespace and \r from header names
                    reader.fieldnames = [h.strip().replace('\r', '') for h in reader.fieldnames]

                for row_num, csv_row in enumerate(reader, start=2):
                    row = DisclosureRow(
                        row_num=row_num,
                        fund_no=_to_int(csv_row.get(D_COL_FUND_NO)),
                        fund_name=_to_str(csv_row.get(D_COL_FUND_NAME)),
                        level_1=_to_str(csv_row.get(D_COL_LEVEL_1)),
                        level_2=_to_str(csv_row.get(D_COL_LEVEL_2)),
                        level_3=_to_str(csv_row.get(D_COL_LEVEL_3)),
                        level_4=_to_str(csv_row.get(D_COL_LEVEL_4)),
                        percent_from_fund=_to_float(csv_row.get(D_COL_PERCENT)),
                        extra_data=_to_str(csv_row.get(D_COL_EXTRA_DATA)),
                        report_date=_parse_ddmmyyyy(csv_row.get(D_COL_REPORT_DATE)),
                        record_no=_to_int(csv_row.get(D_COL_RECORD_NO)),
                        total_records=_to_int(csv_row.get(D_COL_TOTAL_RECORDS)),
                        manager_no=_to_str(csv_row.get(D_COL_MANAGER_NO)),
                    )
                    rows.append(row)
            logger.info("Loaded %d rows from disclosure report (encoding: %s): %s", len(rows), encoding, path.name)
            return rows
        except UnicodeDecodeError:
            rows = []
            continue

    raise ValueError(f"Could not decode file {path} with any supported encoding")


def load_disclosure_report_xlsx(path: Path) -> list[DisclosureRow]:
    """Load disclosure report from XLSX."""
    import shutil
    import tempfile

    rows: list[DisclosureRow] = []

    # If file has wrong extension, copy to temp file with .xlsx
    if path.suffix.lower() != '.xlsx':
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            shutil.copy2(path, tmp.name)
            actual_path = Path(tmp.name)
    else:
        actual_path = path

    try:
        wb = openpyxl.load_workbook(actual_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Build header map from row 1
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val:
                headers[str(val).strip().replace('\r', '')] = col

        for row_num in range(2, ws.max_row + 1):
            def get_val(col_name: str, col_idx_cache=headers):
                col_idx = col_idx_cache.get(col_name)
                if col_idx:
                    return ws.cell(row_num, col_idx).value
                return None

            row = DisclosureRow(
                row_num=row_num,
                fund_no=_to_int(get_val(D_COL_FUND_NO)),
                fund_name=_to_str(get_val(D_COL_FUND_NAME)),
                level_1=_to_str(get_val(D_COL_LEVEL_1)),
                level_2=_to_str(get_val(D_COL_LEVEL_2)),
                level_3=_to_str(get_val(D_COL_LEVEL_3)),
                level_4=_to_str(get_val(D_COL_LEVEL_4)),
                percent_from_fund=_to_float(get_val(D_COL_PERCENT)),
                extra_data=_to_str(get_val(D_COL_EXTRA_DATA)),
                report_date=_parse_ddmmyyyy(get_val(D_COL_REPORT_DATE)),
                record_no=_to_int(get_val(D_COL_RECORD_NO)),
                total_records=_to_int(get_val(D_COL_TOTAL_RECORDS)),
                manager_no=_to_str(get_val(D_COL_MANAGER_NO)),
            )
            # Skip empty rows
            if row.fund_no is not None:
                rows.append(row)

        wb.close()
    finally:
        # Clean up temp file if we created one
        if actual_path != path:
            import os
            try:
                os.unlink(actual_path)
            except:
                pass

    logger.info("Loaded %d rows from disclosure report (XLSX): %s", len(rows), path.name)
    return rows


def load_disclosure_report(path: Path) -> list[DisclosureRow]:
    """Load disclosure report - auto-detect format (CSV or XLSX)."""
    # Check file magic bytes to detect actual format
    with open(path, 'rb') as f:
        magic = f.read(4)

    # PK is ZIP (XLSX) magic
    if magic[:2] == b'PK':
        logger.info("Detected XLSX format for: %s", path.name)
        return load_disclosure_report_xlsx(path)
    else:
        logger.info("Detected CSV format for: %s", path.name)
        return load_disclosure_report_csv(path)


def get_trustee_fund_ids(funds: dict[int, MutualFund], trustee_name: str) -> set[int]:
    """Get fund IDs that are under specified trusteeship."""
    trustee_ids = set()
    for fund_id, fund in funds.items():
        if _norm_spaces(fund.trustee_name) == _norm_spaces(trustee_name):
            trustee_ids.add(fund_id)
    logger.info("Found %d funds for trustee: %s", len(trustee_ids), trustee_name)
    return trustee_ids


# -----------------------------
# Checks
# -----------------------------

def check_1a_fund_completeness(
    disclosure_rows: list[DisclosureRow],
    mizrahi_fund_ids: set[int],
    all_funds: dict[int, MutualFund]
) -> list[ExceptionRow]:
    """Check 1א: Validate fund completeness between mutual funds list and disclosure report."""
    logger_chk1a.info("Starting fund completeness check")

    # Get unique fund IDs from disclosure report
    funds_in_report = set(r.fund_no for r in disclosure_rows if r.fund_no is not None)

    exceptions: list[ExceptionRow] = []

    # Funds in mutual list but missing from report
    missing_from_report = mizrahi_fund_ids - funds_in_report
    for fund_id in sorted(missing_from_report):
        fund = all_funds.get(fund_id)
        fund_name = fund.fund_name if fund else ""
        logger_chk1a.warning("Fund %d (%s) is in mutual funds list but missing from report", fund_id, fund_name)
        exceptions.append(ExceptionRow(
            check_id="1א",
            reason="קרן חסרה בדוח",
            fund_no=fund_id,
            fund_name=fund_name,
        ))

    # Funds in report but missing from mutual list (unexpected)
    extra_in_report = funds_in_report - mizrahi_fund_ids
    # Only flag funds that are completely missing from the mutual funds list
    # (funds under different trustees are expected)
    for fund_id in sorted(extra_in_report):
        if fund_id not in all_funds:
            # Get name from disclosure report
            fund_name = ""
            for r in disclosure_rows:
                if r.fund_no == fund_id:
                    fund_name = r.fund_name or ""
                    break
            logger_chk1a.warning("Fund %d (%s) is in report but not in mutual funds list", fund_id, fund_name)
            exceptions.append(ExceptionRow(
                check_id="1א",
                reason="קרן לא קיימת ברשימת קרנות",
                fund_no=fund_id,
                fund_name=fund_name,
            ))

    logger_chk1a.info("Fund completeness check completed: %d exceptions", len(exceptions))
    return exceptions


def check_1b_report_month_validity(
    disclosure_rows: list[DisclosureRow],
    expected_month: str,
    mizrahi_fund_ids: set[int]
) -> list[ExceptionRow]:
    """Check 1ב: Validate that report_date matches expected report month."""
    logger_chk1b.info("Starting date validity check for month: %s", expected_month)

    year, month = expected_month.split("-")
    expected_year = int(year)
    expected_month_num = int(month)

    exceptions: list[ExceptionRow] = []

    for row in disclosure_rows:
        # Only check in-scope funds
        if row.fund_no not in mizrahi_fund_ids:
            continue

        if row.report_date is None:
            logger_chk1b.warning("Row %d: Missing report date for fund %d", row.row_num, row.fund_no)
            exceptions.append(ExceptionRow(
                check_id="1ב",
                reason="תאריך דוח חסר",
                fund_no=row.fund_no,
                fund_name=row.fund_name,
                effective_code=row.effective_code,
                row_num=row.row_num,
            ))
            continue

        if row.report_date.year != expected_year or row.report_date.month != expected_month_num:
            logger_chk1b.warning(
                "Row %d: Date mismatch for fund %d - expected %s, got %s",
                row.row_num, row.fund_no, expected_month, row.report_date
            )
            exceptions.append(ExceptionRow(
                check_id="1ב",
                reason=f"תאריך לא תואם (צפוי: {expected_month})",
                fund_no=row.fund_no,
                fund_name=row.fund_name,
                effective_code=row.effective_code,
                report_date=row.report_date,
                row_num=row.row_num,
            ))

    logger_chk1b.info("Date validity check completed: %d exceptions", len(exceptions))
    return exceptions


def check_2a_prev_month_comparison(
    current_rows: list[DisclosureRow],
    prev_rows: list[DisclosureRow],
    mizrahi_fund_ids: set[int]
) -> list[ExceptionRow]:
    """Check 2א: Compare disclosure codes between current and previous month."""
    logger_chk2a.info("Starting previous month comparison check")

    # Build lookup: (fund_no, effective_code) -> percent
    def build_lookup(rows: list[DisclosureRow], fund_ids: set[int]) -> dict[tuple[int, str], float]:
        lookup: dict[tuple[int, str], float] = {}
        for row in rows:
            if row.fund_no not in fund_ids:
                continue
            code = row.effective_code
            if code and row.percent_from_fund is not None:
                key = (row.fund_no, code)
                # If multiple rows with same code, sum them (shouldn't happen but be safe)
                lookup[key] = lookup.get(key, 0) + row.percent_from_fund
        return lookup

    current_lookup = build_lookup(current_rows, mizrahi_fund_ids)
    prev_lookup = build_lookup(prev_rows, mizrahi_fund_ids)

    # Get fund names from current rows
    fund_names: dict[int, str] = {}
    for row in current_rows:
        if row.fund_no and row.fund_name:
            fund_names[row.fund_no] = row.fund_name
    for row in prev_rows:
        if row.fund_no and row.fund_name and row.fund_no not in fund_names:
            fund_names[row.fund_no] = row.fund_name

    exceptions: list[ExceptionRow] = []

    all_keys = set(current_lookup.keys()) | set(prev_lookup.keys())

    for fund_no, code in sorted(all_keys):
        current_pct = current_lookup.get((fund_no, code))
        prev_pct = prev_lookup.get((fund_no, code))

        if current_pct is None and prev_pct is not None:
            # Code removed
            logger_chk2a.warning("Fund %d: Code %s removed (was %.2f%%)", fund_no, code, prev_pct)
            exceptions.append(ExceptionRow(
                check_id="2א",
                reason=f"קוד נעלם (היה: {prev_pct:.2f}%)",
                fund_no=fund_no,
                fund_name=fund_names.get(fund_no),
                effective_code=code,
                percent_from_fund=None,
                extra_info={"prev_pct": prev_pct},
            ))
        elif current_pct is not None and prev_pct is None:
            # Code added
            logger_chk2a.warning("Fund %d: Code %s added (now %.2f%%)", fund_no, code, current_pct)
            exceptions.append(ExceptionRow(
                check_id="2א",
                reason=f"קוד חדש (כעת: {current_pct:.2f}%)",
                fund_no=fund_no,
                fund_name=fund_names.get(fund_no),
                effective_code=code,
                percent_from_fund=current_pct,
                extra_info={"prev_pct": None},
            ))
        elif current_pct is not None and prev_pct is not None:
            # Check delta
            delta = abs(current_pct - prev_pct)
            if delta > 10.0:
                logger_chk2a.warning(
                    "Fund %d: Code %s changed by %.2f%% (%.2f%% -> %.2f%%)",
                    fund_no, code, delta, prev_pct, current_pct
                )
                exceptions.append(ExceptionRow(
                    check_id="2א",
                    reason=f"סטייה > 10% (שינוי: {delta:.2f}%)",
                    fund_no=fund_no,
                    fund_name=fund_names.get(fund_no),
                    effective_code=code,
                    percent_from_fund=current_pct,
                    extra_info={"prev_pct": prev_pct, "delta": delta},
                ))

    logger_chk2a.info("Previous month comparison completed: %d exceptions", len(exceptions))
    return exceptions


def check_2b_exposure_profile(
    disclosure_rows: list[DisclosureRow],
    all_funds: dict[int, MutualFund],
    mizrahi_fund_ids: set[int]
) -> list[ExceptionRow]:
    """Check 2ב: Cross-check disclosure exposure codes against fund's exposure profile."""
    logger_chk2b.info("Starting exposure profile check")

    exceptions: list[ExceptionRow] = []

    # Group rows by fund
    fund_rows: dict[int, list[DisclosureRow]] = defaultdict(list)
    for row in disclosure_rows:
        if row.fund_no in mizrahi_fund_ids:
            fund_rows[row.fund_no].append(row)

    for fund_no, rows in fund_rows.items():
        fund = all_funds.get(fund_no)
        if not fund or not fund.exposure_profile:
            continue

        profile = fund.exposure_profile.strip()
        if len(profile) < 2:
            continue

        equity_code = profile[0]
        fx_code = profile[1].upper()

        max_equity = EQUITY_EXPOSURE_PROFILES.get(equity_code, None)
        max_fx = FX_EXPOSURE_PROFILES.get(fx_code, None)

        # Aggregate exposure by code prefix for this fund
        equity_total = 0.0  # code 01 - מניות
        fx_total = 0.0      # code 06 - מט"ח

        for row in rows:
            code = row.effective_code
            if not code:
                continue
            pct = row.percent_from_fund or 0
            if code.startswith("01"):
                equity_total += pct
            elif code.startswith("06"):
                fx_total += pct

        # Check equity exposure (code 01 - מניות) against profile limit
        if max_equity is not None and equity_total > max_equity:
            logger_chk2b.warning(
                "Fund %d: Profile %s allows max %d%% equity but total is %.2f%%",
                fund_no, profile, max_equity, equity_total
            )
            exceptions.append(ExceptionRow(
                check_id="2ב",
                reason=f"פרופיל {profile} מתיר עד {max_equity}% מניות אך סה\"כ חשיפה = {equity_total:.2f}%",
                fund_no=fund_no,
                fund_name=rows[0].fund_name if rows else "",
                effective_code="01",
                percent_from_fund=equity_total,
                row_num=None,
            ))

        # Check FX exposure (code 06 - מט"ח) against profile limit
        if max_fx is not None and fx_total > max_fx:
            logger_chk2b.warning(
                "Fund %d: Profile %s allows max %d%% FX but total is %.2f%%",
                fund_no, profile, max_fx, fx_total
            )
            exceptions.append(ExceptionRow(
                check_id="2ב",
                reason=f'פרופיל {profile} מתיר עד {max_fx}% מט"ח אך סה\"כ חשיפה = {fx_total:.2f}%',
                fund_no=fund_no,
                fund_name=rows[0].fund_name if rows else "",
                effective_code="06",
                percent_from_fund=fx_total,
                row_num=None,
            ))

    logger_chk2b.info("Exposure profile check completed: %d exceptions", len(exceptions))
    return exceptions


def check_3_combinations(
    disclosure_rows: list[DisclosureRow],
    mizrahi_fund_ids: set[int]
) -> dict[str, list[ExceptionRow]]:
    """Check 3א-3ח: Within-month code combinations and cross-checks."""
    logger_chk3.info("Starting within-month combinations check")

    # Group rows by fund
    fund_rows: dict[int, list[DisclosureRow]] = defaultdict(list)
    fund_names: dict[int, str] = {}
    for row in disclosure_rows:
        if row.fund_no in mizrahi_fund_ids:
            fund_rows[row.fund_no].append(row)
            if row.fund_name:
                fund_names[row.fund_no] = row.fund_name

    def has_code_prefix(codes: set[str], prefix: str) -> bool:
        """Check if any code starts with the given prefix."""
        return any(c.startswith(prefix) for c in codes)

    def get_codes_with_prefix(codes: set[str], prefix: str) -> set[str]:
        """Get all codes that start with the given prefix."""
        return {c for c in codes if c.startswith(prefix)}

    results: dict[str, list[ExceptionRow]] = {
        "3א": [],  # FX exposure
        "3ב": [],  # Bond exposure
        "3ג": [],  # Government bonds (shekel)
        "3ד": [],  # Government bonds (linked)
        "3ה": [],  # Government bonds (linked FX)
        "3ו": [],  # Corporate bonds (shekel)
        "3ז": [],  # Corporate bonds (linked)
        "3ח": [],  # Corporate bonds (linked FX)
    }

    for fund_no, rows in fund_rows.items():
        # Get all effective codes for this fund (most granular level)
        codes = set()
        # Get TIER 1 codes specifically (column C / level_1) for check 3ב
        tier1_codes = set()
        # Get TIER 2 codes specifically (column D / level_2) for check 3א
        tier2_codes = set()
        for row in rows:
            code = row.effective_code
            if code:
                codes.add(code)
            # Collect TIER 1 codes for check 3ב
            if row.level_1 and str(row.level_1).strip():
                tier1_codes.add(str(row.level_1).strip())
            # Collect TIER 2 codes for check 3א
            if row.level_2 and row.level_2.strip():
                tier2_codes.add(row.level_2.strip())

        fund_name = fund_names.get(fund_no, "")

        # Check 3א - FX Exposure
        # If 0102, 0302, or 0502 exists in TIER 2 -> 06 must exist in TIER 2
        # If 06 exists in TIER 2 -> at least one of 0102, 0302, 0502 must exist in TIER 2
        # NOTE: We check TIER 2 specifically because these codes have different meanings at deeper levels
        fx_related = "0102" in tier2_codes or "0302" in tier2_codes or "0502" in tier2_codes
        has_06 = any(c.startswith("06") for c in tier2_codes)

        if fx_related and not has_06:
            logger_chk3.warning("Fund %d: Has FX-related codes but missing code 06", fund_no)
            results["3א"].append(ExceptionRow(
                check_id="3א",
                reason='יש חשיפה למט"ח (0102/0302/0502) אך חסר קוד 06',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_06 and not fx_related:
            logger_chk3.warning("Fund %d: Has code 06 but missing FX-related codes", fund_no)
            results["3א"].append(ExceptionRow(
                check_id="3א",
                reason='יש קוד 06 אך חסרים קודי חשיפה למט"ח (0102/0302/0502)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ב - Bond Exposure
        # If 03 exists -> 07 and 08 must exist
        # If 07 or 08 exists -> 03 must exist
        # NOTE: We check TIER 1 (level_1) because 03, 07, 08 are TIER 1 codes per ISA spec
        has_03 = "3" in tier1_codes or "03" in tier1_codes
        has_07 = "7" in tier1_codes or "07" in tier1_codes
        has_08 = "8" in tier1_codes or "08" in tier1_codes

        if has_03 and not (has_07 and has_08):
            missing = []
            if not has_07:
                missing.append("07 (דירוגים)")
            if not has_08:
                missing.append('08 (מח"מ)')
            logger_chk3.warning("Fund %d: Has bonds (03) but missing %s", fund_no, ", ".join(missing))
            results["3ב"].append(ExceptionRow(
                check_id="3ב",
                reason=f'יש אג"ח (03) אך חסר: {", ".join(missing)}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif (has_07 or has_08) and not has_03:
            logger_chk3.warning("Fund %d: Has ratings/duration but missing bonds (03)", fund_no)
            results["3ב"].append(ExceptionRow(
                check_id="3ב",
                reason='יש קוד 07/08 אך חסר קוד אג"ח (03)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ג - Government Bonds (Shekel): 03010101 <-> 080201
        has_03010101 = "03010101" in codes
        has_080201 = "080201" in codes
        if has_03010101 and not has_080201:
            results["3ג"].append(ExceptionRow(
                check_id="3ג",
                reason='יש אג"ח ממשלתי שקלי (03010101) אך חסר מח"מ (080201)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080201 and not has_03010101:
            results["3ג"].append(ExceptionRow(
                check_id="3ג",
                reason='יש מח"מ ממשלתי שקלי (080201) אך חסר אג"ח (03010101)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ד - Government Bonds (Linked): 03010102 <-> 080202
        has_03010102 = "03010102" in codes
        has_080202 = "080202" in codes
        if has_03010102 and not has_080202:
            results["3ד"].append(ExceptionRow(
                check_id="3ד",
                reason='יש אג"ח ממשלתי צמוד (03010102) אך חסר מח"מ (080202)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080202 and not has_03010102:
            results["3ד"].append(ExceptionRow(
                check_id="3ד",
                reason='יש מח"מ ממשלתי צמוד (080202) אך חסר אג"ח (03010102)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ה - Government Bonds (Linked FX): 03010103 <-> 080203
        has_03010103 = "03010103" in codes
        has_080203 = "080203" in codes
        if has_03010103 and not has_080203:
            results["3ה"].append(ExceptionRow(
                check_id="3ה",
                reason='יש אג"ח ממשלתי צמוד מט"ח (03010103) אך חסר מח"מ (080203)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080203 and not has_03010103:
            results["3ה"].append(ExceptionRow(
                check_id="3ה",
                reason='יש מח"מ ממשלתי צמוד מט"ח (080203) אך חסר אג"ח (03010103)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ו - Corporate Bonds (Shekel): 03010202 or 03010203 <-> 080204
        has_corp_shekel = "03010202" in codes or "03010203" in codes
        has_080204 = "080204" in codes
        if has_corp_shekel and not has_080204:
            results["3ו"].append(ExceptionRow(
                check_id="3ו",
                reason='יש אג"ח קונצרני שקלי (03010202/03010203) אך חסר מח"מ (080204)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080204 and not has_corp_shekel:
            results["3ו"].append(ExceptionRow(
                check_id="3ו",
                reason='יש מח"מ קונצרני שקלי (080204) אך חסר אג"ח (03010202/03010203)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ז - Corporate Bonds (Linked): 03010201 <-> 080205
        has_03010201 = "03010201" in codes
        has_080205 = "080205" in codes
        if has_03010201 and not has_080205:
            results["3ז"].append(ExceptionRow(
                check_id="3ז",
                reason='יש אג"ח קונצרני צמוד (03010201) אך חסר מח"מ (080205)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080205 and not has_03010201:
            results["3ז"].append(ExceptionRow(
                check_id="3ז",
                reason='יש מח"מ קונצרני צמוד (080205) אך חסר אג"ח (03010201)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ח - Corporate Bonds (Linked FX): 03010204 <-> 080206
        has_03010204 = "03010204" in codes
        has_080206 = "080206" in codes
        if has_03010204 and not has_080206:
            results["3ח"].append(ExceptionRow(
                check_id="3ח",
                reason='יש אג"ח קונצרני צמוד מט"ח (03010204) אך חסר מח"מ (080206)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080206 and not has_03010204:
            results["3ח"].append(ExceptionRow(
                check_id="3ח",
                reason='יש מח"מ קונצרני צמוד מט"ח (080206) אך חסר אג"ח (03010204)',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

    total_exceptions = sum(len(v) for v in results.values())
    logger_chk3.info("Combinations check completed: %d total exceptions", total_exceptions)
    return results


# -----------------------------
# Excel output helpers
# -----------------------------

def _rtl(ws) -> None:
    ws.sheet_view.rightToLeft = True


def _style_header(ws, row: int = 1) -> None:
    """Apply header styling to specified row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_cells(ws, start_row: int = 2) -> None:
    """Apply styling to all data cells."""
    wrap_cols: set[int] = set()
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(1, col_idx)
        header_text = str(header_cell.value).strip() if header_cell.value is not None else ""
        if header_text in WRAP_HEADERS:
            wrap_cols.add(col_idx)

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            wrap = cell.col_idx in wrap_cols
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=wrap)
            cell.border = THIN_BORDER


def _calculate_text_width(text: str, is_bold: bool = False) -> float:
    """Calculate approximate display width of text in Excel units."""
    if not text:
        return 0.0

    width = 0.0
    for char in text:
        if char == " ":
            width += 0.5
        elif '\u0590' <= char <= '\u05FF':
            width += 1.05
        elif char in '０１２３４５６７８９':
            width += 2.0
        elif ord(char) > 0x4E00:
            width += 2.0
        elif char in 'WMwm':
            width += 1.2
        elif char in 'il|!.,;:\'"':
            width += 0.6
        else:
            width += 1.0

    if is_bold:
        width *= 1.05

    return width


def _auto_fit_columns(ws, min_width: float = 8.0, max_width: float = 50.0, padding: float = 1.2) -> dict[str, float]:
    """Auto-fit column widths based on content."""
    column_widths: dict[str, float] = {}

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_width_found = 0.0

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)

            if cell.value is None:
                continue

            cell_text = str(cell.value)
            is_bold = cell.font.bold if cell.font else False
            is_header = (row_idx == 1)

            lines = cell_text.split('\n')
            for line in lines:
                line_width = _calculate_text_width(line, is_bold or is_header)
                if line_width > max_width_found:
                    max_width_found = line_width

        final_width = max_width_found + padding
        final_width = max(min_width, min(final_width, max_width))

        ws.column_dimensions[column_letter].width = final_width
        column_widths[column_letter] = final_width

    return column_widths


def _auto_fit_rows(ws, column_widths: dict[str, float] = None, line_height: float = 15.0, header_line_height: float = 18.0, min_height: float = 15.0, max_height: float = 120.0) -> None:
    """Auto-fit row heights based on content."""
    if column_widths is None:
        column_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            column_widths[col_letter] = width if width else 10.0

    for row_idx in range(1, ws.max_row + 1):
        max_lines_needed = 1
        is_header = (row_idx == 1)

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)

            if cell.value is None:
                continue

            cell_text = str(cell.value)
            col_letter = get_column_letter(col_idx)
            col_width = column_widths.get(col_letter, 10.0)
            has_wrap = cell.alignment.wrap_text if cell.alignment else False

            lines_in_cell = 0

            for line in cell_text.split('\n'):
                if not line:
                    lines_in_cell += 1
                    continue

                is_bold = cell.font.bold if cell.font else False
                text_width = _calculate_text_width(line, is_bold or is_header)
                available_width = max(col_width - 1.0, 6.0)

                if has_wrap and text_width > available_width:
                    wrapped_lines = int((text_width / available_width) + 0.99)
                    lines_in_cell += max(1, wrapped_lines)
                else:
                    lines_in_cell += 1

            if lines_in_cell > max_lines_needed:
                max_lines_needed = lines_in_cell

        if is_header:
            row_height = header_line_height * max_lines_needed
        else:
            row_height = line_height * max_lines_needed

        row_height = max(min_height, min(row_height, max_height))
        ws.row_dimensions[row_idx].height = row_height


def _set_font_calibri(ws) -> None:
    """Set Calibri font for all cells in worksheet."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name='Calibri',
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                    size=cell.font.size or 11
                )
            else:
                cell.font = Font(name='Calibri', size=11)


def _header(ws, headers: list[str], row: int = 1) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i)
        cell.value = h
    _style_header(ws, row)
    if row == 1:
        ws.freeze_panes = "A2"


# -----------------------------
# Excel output
# -----------------------------

def write_output_xlsx(
    output_path: Path,
    *,
    report_month: str,
    manager_name: str,
    trustee_name: str,
    summary: dict[str, Any],
    exceptions_1a: list[ExceptionRow],
    exceptions_1b: list[ExceptionRow],
    exceptions_2a: list[ExceptionRow],
    exceptions_2b: list[ExceptionRow],
    exceptions_3: dict[str, list[ExceptionRow]],
    spec_file_path: Optional[Path] = None,
) -> None:
    """Write validation results to Excel workbook."""
    wb = openpyxl.Workbook()

    VALIDATION_COLS = ["טופל?", "שם הבודק"]

    # Sheet 1: Summary (סיכום)
    ws_sum = wb.active
    ws_sum.title = "סיכום"
    _rtl(ws_sum)
    _header(ws_sum, ["שדה", "ערך"])

    hebrew_month = _format_report_month_hebrew(report_month)
    summary_rows = [
        ("מנהל קרן", manager_name),
        ("נאמן", trustee_name),
        ("חודש נבדק", f"דוח חודשי-{hebrew_month}"),
        ("מספר קרנות בדוח", summary.get("total_funds_in_report", "")),
        ("מספר קרנות בתחום (מזרחי)", summary.get("in_scope_funds", "")),
        ("מספר קרנות מחוץ לתחום", summary.get("out_of_scope_funds", "")),
        ("סה\"כ שורות בדוח", summary.get("total_rows", "")),
    ]

    rr = 2
    for k, v in summary_rows:
        ws_sum.cell(rr, 1).value = k
        ws_sum.cell(rr, 2).value = v
        rr += 1

    _style_cells(ws_sum)

    # Sheet 2: סטטוס בדיקות (Check Status)
    ws_checks = wb.create_sheet("סטטוס בדיקות")
    _rtl(ws_checks)
    ws_checks.append(["בדיקה", "תיאור", "סטטוס", "חריגות", "טופל?", "שם הבודק"])
    _style_header(ws_checks, 1)

    # Calculate counts for check 3
    count_3a = len(exceptions_3.get("3א", []))
    count_3b = len(exceptions_3.get("3ב", []))
    count_3c = len(exceptions_3.get("3ג", []))
    count_3d = len(exceptions_3.get("3ד", []))
    count_3e = len(exceptions_3.get("3ה", []))
    count_3f = len(exceptions_3.get("3ו", []))
    count_3g = len(exceptions_3.get("3ז", []))
    count_3h = len(exceptions_3.get("3ח", []))

    check_statuses = [
        ("בדיקה #1א - שלמות קרנות", "הצלבה בין רשימת קרנות לדוח", len(exceptions_1a) == 0, len(exceptions_1a)),
        ("בדיקה #1ב - תקינות תאריכים", "התאמת תאריך לחודש הדיווח", len(exceptions_1b) == 0, len(exceptions_1b)),
        ("בדיקה #2א - סבירות מול דוח קודם", "השוואה לחודש קודם", len(exceptions_2a) == 0, len(exceptions_2a)),
        ("בדיקה #2ב - סבירות מול מאפייני הקרן", "הצלבה מול פרופיל חשיפה", len(exceptions_2b) == 0, len(exceptions_2b)),
        ('בדיקה #3א - חשיפה למט"ח', 'הצלבת קודי חשיפה למט"ח', count_3a == 0, count_3a),
        ('בדיקה #3ב - חשיפה לאג"ח', 'הצלבת קודי אג"ח/דירוגים/מח"מ', count_3b == 0, count_3b),
        ('בדיקה #3ג - אג"ח ממשלתי שקלי', 'הצלבת 03010101 מול 080201', count_3c == 0, count_3c),
        ('בדיקה #3ד - אג"ח ממשלתי צמוד', 'הצלבת 03010102 מול 080202', count_3d == 0, count_3d),
        ('בדיקה #3ה - אג"ח ממשלתי צמוד מט"ח', 'הצלבת 03010103 מול 080203', count_3e == 0, count_3e),
        ('בדיקה #3ו - אג"ח קונצרני שקלי', 'הצלבת 03010202/03010203 מול 080204', count_3f == 0, count_3f),
        ('בדיקה #3ז - אג"ח קונצרני צמוד', 'הצלבת 03010201 מול 080205', count_3g == 0, count_3g),
        ('בדיקה #3ח - אג"ח קונצרני צמוד מט"ח', 'הצלבת 03010204 מול 080206', count_3h == 0, count_3h),
    ]

    for row_idx, (name, description, passed, count) in enumerate(check_statuses, start=2):
        ws_checks.append([name, description, "✓ תקין" if passed else "✗ חריגה", count, "", ""])
        fill = PASS_FILL if passed else FAIL_FILL
        for col in range(1, 5):
            ws_checks.cell(row=row_idx, column=col).fill = fill

    _style_cells(ws_checks)

    # Track optional sheets
    optional_sheets = []

    # Exception sheet helper
    def create_exception_sheet(sheet_name: str, exceptions: list[ExceptionRow], extra_columns: list[str] = None):
        if not exceptions:
            return None

        ws = wb.create_sheet(sheet_name)
        _rtl(ws)

        headers = ["בדיקה", "סיבה", "מספר קרן", "שם קרן"]
        if extra_columns:
            headers.extend(extra_columns)
        headers.extend(VALIDATION_COLS)

        _header(ws, headers)

        for ex in exceptions:
            # Clean string values to remove illegal XML characters
            row_data = [
                _clean_excel_string(ex.check_id),
                _clean_excel_string(ex.reason),
                ex.fund_no,
                _clean_excel_string(ex.fund_name)
            ]
            if extra_columns:
                for col in extra_columns:
                    if col == "קוד חשיפה":
                        row_data.append(_clean_excel_string(ex.effective_code))
                    elif col == "%מקרן":
                        row_data.append(ex.percent_from_fund)
                    elif col == "תאריך דוח":
                        row_data.append(_fmt_date(ex.report_date))
                    elif col == "שורה בקובץ":
                        row_data.append(ex.row_num)
                    elif col == "% קודם":
                        row_data.append(ex.extra_info.get("prev_pct"))
                    elif col == "הפרש":
                        row_data.append(ex.extra_info.get("delta"))
            row_data.extend(["", ""])  # Validation columns
            ws.append(row_data)

        optional_sheets.append(ws)
        return ws

    # Create exception sheets
    # Note: Excel sheet names must be <= 31 characters
    create_exception_sheet("בדיקה א1 - שלמות קרנות", exceptions_1a)
    create_exception_sheet("בדיקה ב1 - תקינות תאריכים", exceptions_1b,
                          ["קוד חשיפה", "תאריך דוח", "שורה בקובץ"])
    create_exception_sheet("בדיקה א2 - סבירות דוח קודם", exceptions_2a,
                          ["קוד חשיפה", "%מקרן", "% קודם", "הפרש"])
    create_exception_sheet("בדיקה ב2 - מאפייני קרן", exceptions_2b,
                          ["קוד חשיפה", "%מקרן", "שורה בקובץ"])

    # Create check 3 sheets
    # Note: Sheet names must be <= 31 chars and no invalid chars: " / \ [ ] : * ?
    check_3_names = [
        ("3א", 'בדיקה א3 - חשיפה למט"ח'),
        ("3ב", 'בדיקה ב3 - חשיפה לאג"ח'),
        ("3ג", 'בדיקה ג3 - אג"ח ממשלתי שקלי'),
        ("3ד", 'בדיקה ד3 - אג"ח ממשלתי צמוד'),
        ("3ה", 'בדיקה ה3 - אג"ח ממשלתי מט"ח'),
        ("3ו", 'בדיקה ו3 - אג"ח קונצרני שקלי'),
        ("3ז", 'בדיקה ז3 - אג"ח קונצרני צמוד'),
        ("3ח", 'בדיקה ח3 - אג"ח קונצרני מט"ח'),
    ]

    for check_key, sheet_name in check_3_names:
        create_exception_sheet(sheet_name, exceptions_3.get(check_key, []))

    # Apply styling to all optional sheets
    for ws in optional_sheets:
        _style_cells(ws)

    # Sheet: פירוט בדיקות (copy from spec file if available)
    ws_spec = None
    if spec_file_path and spec_file_path.exists():
        ws_spec = wb.create_sheet("פירוט בדיקות")
        _rtl(ws_spec)

        with open(spec_file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                # Clean each cell value to remove illegal XML characters
                cleaned_row = [_clean_excel_string(cell) if cell else cell for cell in row]
                ws_spec.append(cleaned_row)

        _style_header(ws_spec, 1)
        _style_cells(ws_spec)

    # Auto-fit all sheets
    all_sheets = [ws_sum, ws_checks] + optional_sheets
    if ws_spec:
        all_sheets.append(ws_spec)

    for ws in all_sheets:
        _set_font_calibri(ws)
        column_widths = _auto_fit_columns(ws)
        _auto_fit_rows(ws, column_widths=column_widths)

    wb.save(output_path)
    wb.close()
    logger.info("Output saved to: %s", output_path)


# -----------------------------
# CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Validate K.303 Disclosure Report (Migdal Variant)"
    )
    parser.add_argument(
        "--mutual-funds-list",
        type=Path,
        required=True,
        help="Path to Mutual Funds List CSV"
    )
    parser.add_argument(
        "--current-report",
        type=Path,
        required=True,
        help="Path to current month disclosure report CSV"
    )
    parser.add_argument(
        "--previous-report",
        type=Path,
        required=True,
        help="Path to previous month disclosure report CSV"
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        required=True,
        help="Path for output Excel file"
    )
    parser.add_argument(
        "--email-json",
        type=Path,
        default=None,
        help="Path to write email JSON payload for n8n workflow (not yet implemented)"
    )
    parser.add_argument(
        "--report-month",
        type=str,
        required=True,
        help="Report month in YYYY-MM format (e.g., 2025-12)"
    )
    parser.add_argument(
        "--manager-name",
        type=str,
        default=None,
        help="Fund manager name for report header (optional)"
    )
    parser.add_argument(
        "--trustee-name",
        type=str,
        default=MIZRAHI_TRUSTEE_NAME,
        help="Trustee name filter (default: Mizrahi)"
    )
    parser.add_argument(
        "--spec-file",
        type=Path,
        default=None,
        help="Path to K.303 checklist CSV (for פירוט בדיקות sheet)"
    )

    args = parser.parse_args()

    # Setup logging
    setup_logging()

    logger.info("=" * 70)
    logger.info("Disclosure Report K.303 Validator")
    logger.info("=" * 70)

    # Load data
    logger.info("Loading mutual funds list...")
    all_funds = load_mutual_funds_csv(args.mutual_funds_list)
    in_scope_fund_ids = get_trustee_fund_ids(all_funds, args.trustee_name)

    logger.info("Loading current month disclosure report...")
    current_rows = load_disclosure_report(args.current_report)

    logger.info("Loading previous month disclosure report...")
    prev_rows = load_disclosure_report(args.previous_report)

    # Report month from argument
    report_month = args.report_month
    logger.info("Report month: %s", report_month)

    # Calculate summary stats
    funds_in_report = set(r.fund_no for r in current_rows if r.fund_no is not None)
    in_scope_funds = funds_in_report & in_scope_fund_ids
    out_of_scope_funds = funds_in_report - in_scope_fund_ids

    summary = {
        "total_funds_in_report": len(funds_in_report),
        "in_scope_funds": len(in_scope_funds),
        "out_of_scope_funds": len(out_of_scope_funds),
        "total_rows": len(current_rows),
    }

    # Run checks
    logger.info("Running validation checks...")

    exceptions_1a = check_1a_fund_completeness(current_rows, in_scope_fund_ids, all_funds)
    exceptions_1b = check_1b_report_month_validity(current_rows, report_month, in_scope_fund_ids)
    exceptions_2a = check_2a_prev_month_comparison(current_rows, prev_rows, in_scope_fund_ids)
    exceptions_2b = check_2b_exposure_profile(current_rows, all_funds, in_scope_fund_ids)
    exceptions_3 = check_3_combinations(current_rows, in_scope_fund_ids)

    # Write output
    logger.info("Writing output Excel file...")
    write_output_xlsx(
        args.output_xlsx,
        report_month=report_month,
        manager_name=args.manager_name or "",
        trustee_name=args.trustee_name,
        summary=summary,
        exceptions_1a=exceptions_1a,
        exceptions_1b=exceptions_1b,
        exceptions_2a=exceptions_2a,
        exceptions_2b=exceptions_2b,
        exceptions_3=exceptions_3,
        spec_file_path=args.spec_file,
    )

    # Print summary
    logger.info("=" * 70)
    logger.info("Validation Summary")
    logger.info("=" * 70)
    logger.info("Report Month: %s", report_month)
    logger.info("Total Funds in Report: %d", summary["total_funds_in_report"])
    logger.info("In-Scope Funds (Mizrahi): %d", summary["in_scope_funds"])
    logger.info("Out-of-Scope Funds: %d", summary["out_of_scope_funds"])
    logger.info("")
    logger.info("Check Results:")
    logger.info("  1א Fund Completeness: %d exceptions", len(exceptions_1a))
    logger.info("  1ב Date Validity: %d exceptions", len(exceptions_1b))
    logger.info("  2א Previous Month Comparison: %d exceptions", len(exceptions_2a))
    logger.info("  2ב Exposure Profile: %d exceptions", len(exceptions_2b))
    for check_key in ["3א", "3ב", "3ג", "3ד", "3ה", "3ו", "3ז", "3ח"]:
        count = len(exceptions_3.get(check_key, []))
        logger.info("  %s: %d exceptions", check_key, count)
    logger.info("")
    logger.info("Output saved to: %s", args.output_xlsx)
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
