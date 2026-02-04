#!/usr/bin/env python3
"""
Fund Automation - Complete Standalone Script
Fetches files from Apify actors and processes them into final Excel report.

Usage:
    python fund_automation_complete.py --fund-name "סיגמא"
    python fund_automation_complete.py --fund-name "סיגמא" --output-dir ./reports
"""

import os
import sys
import base64
import argparse
from datetime import datetime, timedelta
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional
import warnings

# Try to import pandas/openpyxl - will fail gracefully with instructions if missing
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Install with: pip install pandas openpyxl requests")
    sys.exit(1)

# Add parent directory to path for shared imports (if run from subdirectory)
sys.path.insert(0, str(Path(__file__).parent))

from shared.apify_client import apify_request, run_actor_and_wait, log
from shared.constants import (
    FUND_MANAGER_CODES, APIFY_ACTORS, HEBREW_MONTHS,
    UNUSUAL_ASSET_TYPES, REQUIRED_COMBINATIONS, COMBINATION_111_THRESHOLD
)
from shared.data_utils import fix_shifted_encoding
from shared.excel_styles import (
    HEADER_FONT, HEADER_FILL, PASS_FILL, FAIL_FILL, THIN_BORDER,
    set_rtl, style_header_row, style_data_cells
)

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

APIFY_TOKEN = os.environ.get("APIFY_TOKEN", "")

# Hardcoded trustee
TRUSTEE_NAME = "מזרחי טפחות"


# ============================================================================
# APIFY FUNCTIONS
# ============================================================================

def build_maya_url(fund_code):
    """Build Maya URL for fund reports"""
    today = datetime.now()
    one_year_ago = today - timedelta(days=365)

    return (
        f"https://maya.tase.co.il/he/reports/funds?"
        f"fromDate={one_year_ago.strftime('%Y-%m-%d')}&toDate={today.strftime('%Y-%m-%d')}"
        f"&noMeetings=false&isSingle=false&isIntendToTaseMember=false"
        f"&by=group&groupId=7&itemId={fund_code}&eventsIds%5B%5D=5618"
    )


def fetch_funds_list():
    """Fetch master funds list from Apify"""
    log("Fetching funds list (master Excel)...")

    run_data = run_actor_and_wait(APIFY_TOKEN, APIFY_ACTORS["FUNDS_LIST"], {})
    dataset_id = run_data["defaultDatasetId"]

    resp = apify_request(APIFY_TOKEN, "GET", f"/datasets/{dataset_id}/items")
    items = resp.json()

    if not items:
        raise Exception("No items in dataset")

    file_base64 = items[0].get("fileBase64")
    if not file_base64:
        raise Exception(f"No fileBase64 in response. Keys: {items[0].keys()}")

    return base64.b64decode(file_base64)


def fetch_fund_reports(fund_code):
    """Fetch fund manager reports from Apify"""
    log(f"Fetching fund reports for code: {fund_code}")

    maya_url = build_maya_url(fund_code)
    log(f"Maya URL: {maya_url[:80]}...")

    run_data = run_actor_and_wait(APIFY_TOKEN, APIFY_ACTORS["FUND_REPORTS"], {"url": maya_url}, timeout=180)

    # Get report name from dataset
    dataset_id = run_data["defaultDatasetId"]
    resp = apify_request(APIFY_TOKEN, "GET", f"/datasets/{dataset_id}/items")
    items = resp.json()

    report_name = ""
    if items and items[0].get("downloadedFiles"):
        report_name = items[0]["downloadedFiles"][0].get("reportName", "")

    if not report_name:
        # Fallback: calculate from current date
        today = datetime.now()
        prev_month = today.replace(day=1) - timedelta(days=1)
        report_name = f"{HEBREW_MONTHS[prev_month.month]} {prev_month.year}"

    log(f"Report name: {report_name}")

    # Get CSVs from key-value store
    kv_store_id = run_data["defaultKeyValueStoreId"]

    current_resp = apify_request(APIFY_TOKEN, "GET", f"/key-value-stores/{kv_store_id}/records/report_latest_month.csv")
    previous_resp = apify_request(APIFY_TOKEN, "GET", f"/key-value-stores/{kv_store_id}/records/report_previous_month.csv")

    log(f"Fetched CSVs - current: {len(current_resp.content)} bytes, previous: {len(previous_resp.content)} bytes")

    return current_resp.content, previous_resp.content, report_name


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class FundDiscrepancy:
    fund_number: str
    fund_name: str
    status: str = ""
    trustee: str = ""

@dataclass
class AssetAlert:
    fund_number: str
    fund_name: str
    asset_name: str
    asset_id: str
    asset_type: int
    alert_type: str
    details: str = ""

@dataclass
class CheckStatus:
    name: str
    description: str
    passed: bool
    issue_count: int

@dataclass
class ProcessingResult:
    manager_name: str
    trustee_filter: str
    magna_funds_count: int
    manager_funds_count: int
    matching_funds_count: int
    other_trustee_funds_count: int = 0

    only_in_magna: list = field(default_factory=list)
    only_in_manager: list = field(default_factory=list)

    unusual_assets_current: list = field(default_factory=list)
    unusual_assets_previous: list = field(default_factory=list)
    new_assets: list = field(default_factory=list)
    changed_assets: list = field(default_factory=list)
    clause_214_issues: list = field(default_factory=list)
    clause_328_issues: list = field(default_factory=list)
    combination_issues: list = field(default_factory=list)
    price_issues: list = field(default_factory=list)

    has_discrepancies: bool = False
    email_address: str = ""
    report_date: str = ""

    def get_check_statuses(self) -> list:
        combined_count = len(self.clause_214_issues) + len(self.combination_issues)
        return [
            CheckStatus("בדיקת שלמות", "הצלבת קרנות מגנא מול דוח מנהל",
                       len(self.only_in_magna) == 0 and len(self.only_in_manager) == 0,
                       len(self.only_in_magna) + len(self.only_in_manager)),
            CheckStatus("סוגי נכסים חריגים", "זיהוי נכסים מסוגים חריגים עם שווי > 0",
                       len(self.unusual_assets_current) == 0,
                       len(self.unusual_assets_current)),
            CheckStatus("נכסים חדשים", "נכסים חריגים שנוספו מהחודש הקודם",
                       len(self.new_assets) == 0,
                       len(self.new_assets)),
            CheckStatus("שינויים בכמות", "שינויים בכמות נכסים חריגים",
                       len(self.changed_assets) == 0,
                       len(self.changed_assets)),
            CheckStatus("סעיף 328", "כמות שהושאלה - בדיקת עקביות",
                       len(self.clause_328_issues) == 0,
                       len(self.clause_328_issues)),
            CheckStatus("שילובים נדרשים", "בדיקת צמדי סוגי נכסים נדרשים (כולל סעיף 214)",
                       combined_count == 0,
                       combined_count),
            CheckStatus("סבירות מחירים", "יחס בין מחירים - סף 7.5%",
                       len(self.price_issues) == 0,
                       len(self.price_issues)),
        ]


# ============================================================================
# PROCESSING FUNCTIONS
# ============================================================================

def load_data(funds_list_path: str, current_report_path: str, previous_report_path: Optional[str] = None) -> tuple:
    """Load and return all data files"""
    def read_file(path):
        if path.lower().endswith('.csv'):
            return pd.read_csv(path, encoding='utf-8-sig')
        return pd.read_excel(path)

    funds_list = read_file(funds_list_path)
    current_report = read_file(current_report_path)
    previous_report = None
    if previous_report_path:
        try:
            previous_report = read_file(previous_report_path)
        except Exception as e:
            log(f"Warning: Could not load previous report: {e}")
    return funds_list, current_report, previous_report


def filter_funds_by_trustee_and_manager(funds_list: pd.DataFrame, manager_name: str, trustee_name: str) -> pd.DataFrame:
    """Filter funds list by manager and trustee names."""
    filtered = funds_list.copy()

    if trustee_name:
        filtered = filtered[filtered['שם נאמן'].str.contains(trustee_name, na=False)]

    if manager_name:
        filtered = filtered[filtered['שם מנהל'].str.contains(manager_name, na=False)]

    if 'מצב הקרן' in filtered.columns:
        filtered = filtered[filtered['מצב הקרן'] == 'פעיל']

    return filtered


def check_completeness(filtered_funds: pd.DataFrame, manager_report: pd.DataFrame, full_funds_list: Optional[pd.DataFrame] = None) -> tuple:
    """Check 1: Cross-reference funds between Magna list and Manager report."""
    magna_funds = set(filtered_funds['מספר בורסה'].astype(str).unique())
    manager_funds = set(manager_report['מספר קרן'].astype(str).unique())

    only_in_magna = magna_funds - manager_funds
    matching = magna_funds & manager_funds
    raw_only_in_manager = manager_funds - magna_funds

    other_trustee_funds = set()
    true_only_in_manager = set()

    if full_funds_list is not None:
        all_fund_trustees = full_funds_list.set_index(
            full_funds_list['מספר בורסה'].astype(str)
        )['שם נאמן'].to_dict()

        our_trustee = filtered_funds['שם נאמן'].iloc[0] if len(filtered_funds) > 0 else ""

        for fund_num in raw_only_in_manager:
            fund_trustee = all_fund_trustees.get(fund_num, "")
            if fund_trustee and our_trustee and our_trustee in fund_trustee:
                true_only_in_manager.add(fund_num)
            elif fund_trustee:
                other_trustee_funds.add(fund_num)
            else:
                true_only_in_manager.add(fund_num)
    else:
        true_only_in_manager = raw_only_in_manager

    magna_details = []
    for fund_num in only_in_magna:
        fund = filtered_funds[filtered_funds['מספר בורסה'].astype(str) == fund_num]
        if len(fund) > 0:
            fund = fund.iloc[0]
            magna_details.append(FundDiscrepancy(
                fund_number=fund_num,
                fund_name=fund.get('שם קרן בעברית', ''),
                status=fund.get('מצב הקרן', ''),
                trustee=fund.get('שם נאמן', '')
            ))

    manager_details = []
    seen = set()
    for _, row in manager_report.iterrows():
        fund_num = str(row['מספר קרן'])
        if fund_num in true_only_in_manager and fund_num not in seen:
            manager_details.append(FundDiscrepancy(
                fund_number=fund_num,
                fund_name=row.get('שם קרן', '')
            ))
            seen.add(fund_num)

    return magna_details, manager_details, len(matching), len(other_trustee_funds)


def check_unusual_asset_types(manager_report: pd.DataFrame, filtered_fund_numbers: set, previous_report: Optional[pd.DataFrame] = None, label: str = "current") -> list:
    """Check 2: Flag holdings with unusual asset types."""
    alerts = []

    our_holdings = manager_report[manager_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)]

    unusual = our_holdings[
        (our_holdings['סוג נכס'].isin(UNUSUAL_ASSET_TYPES)) &
        (our_holdings['שווי בשקלים'].fillna(0) != 0)
    ].copy()

    if previous_report is not None:
        previous_filtered = previous_report[
            (previous_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)) &
            (previous_report['סוג נכס'].isin(UNUSUAL_ASSET_TYPES))
        ]

        def make_key(row):
            return (str(row['מספר קרן']), str(row.get('מספר מזהה', '')))

        unusual['_key'] = unusual.apply(make_key, axis=1)
        previous_keys = set(previous_filtered.apply(make_key, axis=1))
        unusual = unusual[unusual['_key'].isin(previous_keys)]

    for _, row in unusual.iterrows():
        alerts.append(AssetAlert(
            fund_number=str(row['מספר קרן']),
            fund_name=row.get('שם קרן', ''),
            asset_name=row.get('שם נכס', ''),
            asset_id=str(row.get('מספר מזהה', '')),
            asset_type=int(row['סוג נכס']),
            alert_type='unusual',
            details=f"שווי: {row.get('שווי בשקלים', 0):,.2f}"
        ))

    return alerts


def check_new_and_changed_assets(current_report: pd.DataFrame, previous_report: Optional[pd.DataFrame], filtered_fund_numbers: set) -> tuple:
    """Check 3: Compare current vs previous month to find new/changed assets."""
    if previous_report is None:
        return [], []

    new_assets = []
    changed_assets = []

    current = current_report[
        (current_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)) &
        (current_report['סוג נכס'].isin(UNUSUAL_ASSET_TYPES))
    ].copy()
    previous = previous_report[
        (previous_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)) &
        (previous_report['סוג נכס'].isin(UNUSUAL_ASSET_TYPES))
    ].copy()

    def make_key(row):
        return (str(row['מספר קרן']), str(row.get('מספר מזהה', '')))

    current['_key'] = current.apply(make_key, axis=1)
    previous['_key'] = previous.apply(make_key, axis=1)

    current_positions = set(current['_key'])
    previous_positions = set(previous['_key'])

    new_position_keys = current_positions - previous_positions
    for _, row in current[current['_key'].isin(new_position_keys)].iterrows():
        if (row.get('שווי בשקלים', 0) or 0) != 0:
            new_assets.append(AssetAlert(
                fund_number=str(row['מספר קרן']),
                fund_name=row.get('שם קרן', ''),
                asset_name=row.get('שם נכס', ''),
                asset_id=str(row.get('מספר מזהה', '')),
                asset_type=int(row.get('סוג נכס', 0)),
                alert_type='new',
                details=f"כמות: {row.get('כמות', 0):,.2f}, שווי: {row.get('שווי בשקלים', 0):,.2f}"
            ))

    common_keys = current_positions & previous_positions
    current_qty = current.set_index('_key')['כמות'].to_dict()
    previous_qty = previous.set_index('_key')['כמות'].to_dict()

    for key in common_keys:
        curr_q = current_qty.get(key, 0) or 0
        prev_q = previous_qty.get(key, 0) or 0

        if abs(curr_q - prev_q) > 0.001:
            row = current[current['_key'] == key].iloc[0]
            if (row.get('שווי בשקלים', 0) or 0) != 0:
                changed_assets.append(AssetAlert(
                    fund_number=str(row['מספר קרן']),
                    fund_name=row.get('שם קרן', ''),
                    asset_name=row.get('שם נכס', ''),
                    asset_id=str(row.get('מספר מזהה', '')),
                    asset_type=int(row.get('סוג נכס', 0)),
                    alert_type='changed',
                    details=f"כמות: {prev_q:,.2f} → {curr_q:,.2f} (Δ: {curr_q - prev_q:+,.2f})"
                ))

    return new_assets, changed_assets


def check_clause_214(manager_report: pd.DataFrame, funds_list: pd.DataFrame, filtered_fund_numbers: set) -> list:
    """Check 4: Clause 214 - If asset type 214 exists, expect variable management fees."""
    alerts = []

    holdings_214 = manager_report[
        (manager_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)) &
        (manager_report['סוג נכס'] == 214)
    ]

    fee_lookup = funds_list.set_index('מספר בורסה')['דמי ניהול משתנים'].to_dict()

    our_holdings = manager_report[manager_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)]
    fund_names = our_holdings.groupby('מספר קרן')['שם קרן'].first().to_dict()

    for fund_num in holdings_214['מספר קרן'].unique():
        fund_214 = holdings_214[holdings_214['מספר קרן'] == fund_num]
        asset_214_value = fund_214['שווי בשקלים'].sum() if len(fund_214) > 0 else 0

        if asset_214_value == 0 or pd.isna(asset_214_value):
            continue

        fee = fee_lookup.get(fund_num, 0)
        if pd.isna(fee):
            fee = 0

        if fee == 0:
            fund_name = fund_names.get(fund_num, '')
            alerts.append(AssetAlert(
                fund_number=str(fund_num),
                fund_name=fund_name,
                asset_name="N/A",
                asset_id="N/A",
                asset_type=214,
                alert_type='clause_214',
                details=f"קיים סוג נכס 214 (שווי: {asset_214_value:,.2f}) אך דמי ניהול משתנים = 0"
            ))

    return alerts


def check_clause_328(manager_report: pd.DataFrame, filtered_fund_numbers: set) -> list:
    """Check 5: Clause 328 - If asset type 328 has non-zero value, sum of borrowed quantity must be non-zero."""
    alerts = []

    our_holdings = manager_report[manager_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)]
    holdings_328 = our_holdings[our_holdings['סוג נכס'] == 328]
    borrowed_by_fund = our_holdings.groupby('מספר קרן')['כמות שהושאלה'].sum()

    for _, row in holdings_328.iterrows():
        fund_num = row['מספר קרן']
        value_328 = row.get('שווי בשקלים', 0) or 0

        if value_328 != 0:
            total_borrowed = borrowed_by_fund.get(fund_num, 0) or 0

            if total_borrowed == 0:
                alerts.append(AssetAlert(
                    fund_number=str(fund_num),
                    fund_name=row.get('שם קרן', ''),
                    asset_name=row.get('שם נכס', ''),
                    asset_id=str(row.get('מספר מזהה', '')),
                    asset_type=328,
                    alert_type='clause_328',
                    details=f"Asset type 328 has value {value_328:,.2f} but total borrowed quantity is 0"
                ))

    return alerts


def check_required_combinations(manager_report: pd.DataFrame, filtered_fund_numbers: set) -> list:
    """Check 6: Verify required asset type combinations."""
    alerts = []

    our_holdings = manager_report[manager_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)]
    fund_asset_types = our_holdings.groupby('מספר קרן')['סוג נכס'].apply(set).to_dict()
    fund_names = our_holdings.groupby('מספר קרן')['שם קרן'].first().to_dict()

    for fund_num, asset_types in fund_asset_types.items():
        fund_data = our_holdings[our_holdings['מספר קרן'] == fund_num]

        for required_type, trigger_types in REQUIRED_COMBINATIONS.items():
            if required_type == 111:
                trigger_assets = fund_data[
                    (fund_data['סוג נכס'].isin(trigger_types)) &
                    (fund_data['שווי בשקלים'].fillna(0) >= COMBINATION_111_THRESHOLD)
                ]
                if len(trigger_assets) > 0 and 111 not in asset_types:
                    present_triggers = trigger_assets['סוג נכס'].unique().tolist()
                    max_value = trigger_assets['שווי בשקלים'].max()
                    alerts.append(AssetAlert(
                        fund_number=str(fund_num),
                        fund_name=fund_names.get(fund_num, ''),
                        asset_name="N/A",
                        asset_id="N/A",
                        asset_type=required_type,
                        alert_type='combination',
                        details=f"סוג נכס {present_triggers} עם שווי {max_value:,.0f} >= 100,000 - חסר סוג 111"
                    ))
            else:
                has_trigger = any(t in asset_types for t in trigger_types)
                if has_trigger and required_type not in asset_types:
                    present_triggers = [t for t in trigger_types if t in asset_types]
                    alerts.append(AssetAlert(
                        fund_number=str(fund_num),
                        fund_name=fund_names.get(fund_num, ''),
                        asset_name="N/A",
                        asset_id="N/A",
                        asset_type=required_type,
                        alert_type='combination',
                        details=f"קיים סוג נכס {present_triggers} - חסר סוג {required_type}"
                    ))

    return alerts


def check_price_reasonableness(manager_report: pd.DataFrame, filtered_fund_numbers: set) -> list:
    """Check 7: Price reasonableness"""
    alerts = []

    our_holdings = manager_report[manager_report['מספר קרן'].astype(str).isin(filtered_fund_numbers)]
    fund_names = our_holdings.groupby('מספר קרן')['שם קרן'].first().to_dict()

    for fund_num in our_holdings['מספר קרן'].unique():
        fund_data = our_holdings[our_holdings['מספר קרן'] == fund_num]
        fund_name = fund_names.get(fund_num, '')

        values_by_type = {}
        for _, row in fund_data.iterrows():
            asset_type = row['סוג נכס']
            if asset_type in [300, 314, 316]:
                values_by_type[asset_type] = row.get('שווי בשקלים', 0) or 0
            elif asset_type in [301, 313, 315]:
                values_by_type[asset_type] = row.get('כמות', 0) or 0

        ratios = []
        ratio_details = []

        pairs = [(300, 301), (314, 313), (316, 315)]
        for t_h, t_g in pairs:
            v_h = values_by_type.get(t_h, 0)
            v_g = values_by_type.get(t_g, 0)
            if v_h > 0 and v_g > 0:
                ratio = v_h / v_g
                ratios.append(ratio)
                ratio_details.append(f"{t_h}/{t_g}={ratio:.4f}")

        if len(ratios) >= 2:
            min_ratio = min(ratios)
            max_ratio = max(ratios)

            if max_ratio > 0:
                diff_pct = (max_ratio - min_ratio) / max_ratio * 100

                if diff_pct > 7.5:
                    alerts.append(AssetAlert(
                        fund_number=str(fund_num),
                        fund_name=fund_name,
                        asset_name="יחסי מחירים",
                        asset_id="N/A",
                        asset_type=0,
                        alert_type='price_ratio',
                        details=f"פער {diff_pct:.2f}% בין יחסים ({', '.join(ratio_details)}) - חריגה מ-7.5%"
                    ))

    return alerts


# ============================================================================
# EXCEL REPORT GENERATION
# ============================================================================

def generate_excel_report(result: ProcessingResult, output_path: str):
    """Generate Excel report with all results in separate sheets."""
    wb = Workbook()

    def style_sheet(ws, start_row=2):
        set_rtl(ws)
        style_header_row(ws, row=1)
        style_data_cells(ws, start_row=start_row)

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "סיכום"

    summary_data = [
        ["שדה", "ערך"],
        ["מנהל קרן", result.manager_name],
        ["נאמן", result.trustee_filter],
        ["חודש נבדק", result.report_date],
        ["קרנות במגנא", result.magna_funds_count],
        ["קרנות בדוח מנהל", result.manager_funds_count],
        ["קרנות תואמות", result.matching_funds_count],
        ["קרנות רק במגנא", len(result.only_in_magna)],
        ["קרנות רק בדוח מנהל", len(result.only_in_manager)],
        ["קרנות מנאמנים אחרים", result.other_trustee_funds_count],
    ]
    for row in summary_data:
        ws_summary.append(row)
    style_sheet(ws_summary)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # Sheet 2: Check Statuses
    ws_checks = wb.create_sheet("סטטוס בדיקות")
    ws_checks.append(["בדיקה", "תיאור", "סטטוס", "חריגות", "טופל?", "שם הבודק"])
    for check in result.get_check_statuses():
        ws_checks.append([check.name, check.description, "✓ תקין" if check.passed else "✗ חריגה", check.issue_count, "", ""])
    style_sheet(ws_checks)
    for row_idx, check in enumerate(result.get_check_statuses(), start=2):
        fill = PASS_FILL if check.passed else FAIL_FILL
        for col in range(1, 5):
            ws_checks.cell(row=row_idx, column=col).fill = fill
    ws_checks.column_dimensions['A'].width = 20
    ws_checks.column_dimensions['B'].width = 40
    ws_checks.column_dimensions['C'].width = 15
    ws_checks.column_dimensions['D'].width = 12

    # Sheet 3: Missing funds
    if result.only_in_magna or result.only_in_manager:
        ws_missing = wb.create_sheet("קרנות חסרות")
        ws_missing.append(["מספר קרן", "שם קרן", "סטטוס", "מקור", "האם תקין?", "שם הבודק"])
        for fund in result.only_in_magna:
            f = fund if isinstance(fund, dict) else asdict(fund)
            ws_missing.append([f['fund_number'], f['fund_name'], f.get('status', ''), "חסר בדוח מנהל", "", ""])
        for fund in result.only_in_manager:
            f = fund if isinstance(fund, dict) else asdict(fund)
            ws_missing.append([f['fund_number'], f['fund_name'], '', "חסר במגנא", "", ""])
        style_sheet(ws_missing)

    # Sheet 4: Unusual assets
    if result.unusual_assets_current:
        ws_unusual = wb.create_sheet("נכסים חריגים")
        ws_unusual.append(["מספר קרן", "שם קרן", "סוג נכס", "מספר נייר", "שם נייר", "שווי", "האם תקין?", "שם הבודק"])
        for alert in result.unusual_assets_current:
            a = alert if isinstance(alert, dict) else asdict(alert)
            ws_unusual.append([a['fund_number'], a['fund_name'], a['asset_type'], a['asset_id'], a['asset_name'], a['details'], "", ""])
        style_sheet(ws_unusual)

    # Sheet 5: New assets
    if result.new_assets:
        ws_new = wb.create_sheet("נכסים חדשים")
        ws_new.append(["מספר קרן", "שם קרן", "סוג נכס", "מספר נייר", "שם נייר", "פרטים", "האם תקין?", "שם הבודק"])
        for alert in result.new_assets:
            a = alert if isinstance(alert, dict) else asdict(alert)
            ws_new.append([a['fund_number'], a['fund_name'], a['asset_type'], a['asset_id'], a['asset_name'], a['details'], "", ""])
        style_sheet(ws_new)

    # Sheet 6: Changed assets
    if result.changed_assets:
        ws_changed = wb.create_sheet("שינויים בכמות")
        ws_changed.append(["מספר קרן", "שם קרן", "סוג נכס", "מספר נייר", "שם נייר", "פרטים", "האם תקין?", "שם הבודק"])
        for alert in result.changed_assets:
            a = alert if isinstance(alert, dict) else asdict(alert)
            ws_changed.append([a['fund_number'], a['fund_name'], a['asset_type'], a['asset_id'], a['asset_name'], a['details'], "", ""])
        style_sheet(ws_changed)

    # Sheet 7: Clause 328
    if result.clause_328_issues:
        ws_328 = wb.create_sheet("סעיף 328")
        ws_328.append(["מספר קרן", "שם קרן", "סוג נכס", "מספר נייר", "שם נייר", "פרטים", "האם תקין?", "שם הבודק"])
        for alert in result.clause_328_issues:
            a = alert if isinstance(alert, dict) else asdict(alert)
            ws_328.append([a['fund_number'], a['fund_name'], a['asset_type'], a['asset_id'], a['asset_name'], a['details'], "", ""])
        style_sheet(ws_328)

    # Sheet 8: Combinations (includes 214)
    combined_issues = result.clause_214_issues + result.combination_issues
    if combined_issues:
        ws_combo = wb.create_sheet("שילובים נדרשים")
        ws_combo.append(["מספר קרן", "שם קרן", "סוג נכס", "מספר נייר", "שם נייר", "פרטים", "האם תקין?", "שם הבודק"])
        for alert in combined_issues:
            a = alert if isinstance(alert, dict) else asdict(alert)
            ws_combo.append([a['fund_number'], a['fund_name'], a['asset_type'], a['asset_id'], a['asset_name'], a['details'], "", ""])
        style_sheet(ws_combo)

    # Sheet 9: Price issues
    if result.price_issues:
        ws_price = wb.create_sheet("סבירות מחירים")
        ws_price.append(["מספר קרן", "שם קרן", "פרטים", "טופל?", "שם הבודק"])
        for alert in result.price_issues:
            a = alert if isinstance(alert, dict) else asdict(alert)
            ws_price.append([a['fund_number'], a['fund_name'], a['details'], "", ""])
        style_sheet(ws_price)

    wb.save(output_path)
    log(f"Excel report saved to: {output_path}")


# ============================================================================
# MAIN PROCESSING
# ============================================================================

def process_fund_reports(funds_list_path: str, current_report_path: str, previous_report_path: str,
                         manager_name: str, trustee_name: str, report_month: str) -> ProcessingResult:
    """Main processing function - runs all checks and returns structured results."""

    log("Loading data files...")
    funds_list, current_report, previous_report = load_data(
        funds_list_path, current_report_path, previous_report_path
    )

    log(f"Filtering funds by trustee='{trustee_name}' and manager='{manager_name}'...")
    filtered_funds = filter_funds_by_trustee_and_manager(funds_list, manager_name, trustee_name)
    filtered_fund_numbers = set(filtered_funds['מספר בורסה'].astype(str))

    log(f"Found {len(filtered_funds)} matching funds in Magna list")

    result = ProcessingResult(
        manager_name=manager_name,
        trustee_filter=trustee_name,
        magna_funds_count=len(filtered_funds),
        manager_funds_count=0,
        matching_funds_count=0,
        report_date=report_month
    )

    # Check 1: Completeness
    log("Running Check 1: Fund completeness cross-reference...")
    only_magna, only_manager, matching_count, other_trustee_count = check_completeness(
        filtered_funds, current_report, funds_list
    )
    result.only_in_magna = [asdict(f) for f in only_magna]
    result.only_in_manager = [asdict(f) for f in only_manager]
    result.matching_funds_count = matching_count
    result.other_trustee_funds_count = other_trustee_count
    result.manager_funds_count = len(current_report['מספר קרן'].unique())
    result.has_discrepancies = len(only_magna) > 0 or len(only_manager) > 0

    # Check 2: Unusual asset types
    log("Running Check 2: Unusual asset types...")
    result.unusual_assets_current = [
        asdict(a) for a in check_unusual_asset_types(current_report, filtered_fund_numbers, previous_report)
    ]

    # Check 3: New and changed assets
    log("Running Check 3: New and changed assets...")
    new_assets, changed_assets = check_new_and_changed_assets(current_report, previous_report, filtered_fund_numbers)
    result.new_assets = [asdict(a) for a in new_assets]
    result.changed_assets = [asdict(a) for a in changed_assets]

    # Check 4: Clause 214
    log("Running Check 4: Clause 214...")
    result.clause_214_issues = [asdict(a) for a in check_clause_214(current_report, funds_list, filtered_fund_numbers)]

    # Check 5: Clause 328
    log("Running Check 5: Clause 328...")
    result.clause_328_issues = [asdict(a) for a in check_clause_328(current_report, filtered_fund_numbers)]

    # Check 6: Required combinations
    log("Running Check 6: Required combinations...")
    result.combination_issues = [asdict(a) for a in check_required_combinations(current_report, filtered_fund_numbers)]

    # Check 7: Price reasonableness
    log("Running Check 7: Price reasonableness...")
    result.price_issues = [asdict(a) for a in check_price_reasonableness(current_report, filtered_fund_numbers)]

    log("Processing complete!")
    return result


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Fund Automation - Complete Pipeline")
    parser.add_argument('--fund-name', required=True, help=f'Fund manager name. Options: {list(FUND_MANAGER_CODES.keys())}')
    parser.add_argument('--output-dir', default='.', help='Output directory for files')
    parser.add_argument('--keep-temp', action='store_true', help='Keep temporary files')

    args = parser.parse_args()

    fund_name = args.fund_name
    fund_code = FUND_MANAGER_CODES.get(fund_name)

    if not fund_code:
        print(f"ERROR: Unknown fund manager: {fund_name}")
        print(f"Available: {list(FUND_MANAGER_CODES.keys())}")
        sys.exit(1)

    if not APIFY_TOKEN:
        print("ERROR: APIFY_TOKEN environment variable not set")
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)
    temp_dir = os.path.join(args.output_dir, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    log("=" * 60)
    log("FUND AUTOMATION - COMPLETE PIPELINE")
    log(f"Fund: {fund_name} (code: {fund_code})")
    log("=" * 60)

    try:
        # Step 1: Fetch funds list
        log("\n--- STEP 1: Fetching master funds list ---")
        funds_list_bytes = fetch_funds_list()
        funds_list_path = os.path.join(temp_dir, "funds_list.xlsx")
        with open(funds_list_path, "wb") as f:
            f.write(funds_list_bytes)
        log(f"Saved: {funds_list_path} ({len(funds_list_bytes):,} bytes)")

        # Step 2: Fetch fund reports
        log("\n--- STEP 2: Fetching fund reports ---")
        current_csv, previous_csv, report_name = fetch_fund_reports(fund_code)

        current_path = os.path.join(temp_dir, "current_month.csv")
        with open(current_path, "wb") as f:
            f.write(fix_shifted_encoding(current_csv))

        previous_path = os.path.join(temp_dir, "previous_month.csv")
        with open(previous_path, "wb") as f:
            f.write(fix_shifted_encoding(previous_csv))

        log(f"Saved: {current_path} ({len(current_csv):,} bytes)")
        log(f"Saved: {previous_path} ({len(previous_csv):,} bytes)")
        log(f"Report month: {report_name}")

        # Step 3: Process
        log("\n--- STEP 3: Processing data ---")
        result = process_fund_reports(
            funds_list_path=funds_list_path,
            current_report_path=current_path,
            previous_report_path=previous_path,
            manager_name=fund_name,
            trustee_name=TRUSTEE_NAME,
            report_month=report_name
        )

        # Step 4: Generate Excel
        log("\n--- STEP 4: Generating Excel report ---")
        output_filename = f"דוח_{fund_name}_{report_name.replace(' ', '_')}.xlsx"
        output_path = os.path.join(args.output_dir, output_filename)
        generate_excel_report(result, output_path)

        # Cleanup temp files
        if not args.keep_temp:
            import shutil
            shutil.rmtree(temp_dir)
            log("Cleaned up temporary files")

        # Summary
        log("\n" + "=" * 60)
        log("SUCCESS!")
        log("=" * 60)
        log(f"Output file: {output_path}")
        log(f"Manager: {result.manager_name}")
        log(f"Trustee: {result.trustee_filter}")
        log(f"Report month: {result.report_date}")
        log(f"Magna funds: {result.magna_funds_count}")
        log(f"Manager funds: {result.manager_funds_count}")
        log(f"Matching: {result.matching_funds_count}")
        log(f"Issues found: {result.has_discrepancies}")

        return output_path

    except Exception as e:
        log(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
