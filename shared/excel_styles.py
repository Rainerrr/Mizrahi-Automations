#!/usr/bin/env python3
"""Shared Excel styling utilities for Mizrahi automations."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# COLOR CONSTANTS
# =============================================================================

HEADER_BG_COLOR = "4472C4"
PASS_BG_COLOR = "C6EFCE"
FAIL_BG_COLOR = "FFC7CE"
ERROR_BG_COLOR = "C00000"
STRIPE_BG_COLOR = "D6DCE4"
BORDER_COLOR = "B4B4B4"

# =============================================================================
# FONT STYLES
# =============================================================================

HEADER_FONT = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
DEFAULT_FONT = Font(name='Calibri', size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=18, color="FFFFFF")
ERROR_FONT = Font(color="FF0000")
BOLD_FONT = Font(bold=True)

# =============================================================================
# FILL STYLES
# =============================================================================

HEADER_FILL = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
PASS_FILL = PatternFill(start_color=PASS_BG_COLOR, end_color=PASS_BG_COLOR, fill_type="solid")
FAIL_FILL = PatternFill(start_color=FAIL_BG_COLOR, end_color=FAIL_BG_COLOR, fill_type="solid")
ERROR_FILL = PatternFill(start_color=ERROR_BG_COLOR, end_color=ERROR_BG_COLOR, fill_type="solid")
STRIPE_FILL = PatternFill(start_color=STRIPE_BG_COLOR, end_color=STRIPE_BG_COLOR, fill_type="solid")

# =============================================================================
# BORDER STYLES
# =============================================================================

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

LIGHT_BORDER = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

# =============================================================================
# ALIGNMENT STYLES
# =============================================================================

RTL_ALIGNMENT = Alignment(horizontal='right', vertical='center')
WRAP_ALIGNMENT = Alignment(horizontal='right', vertical='top', wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center')

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================


def set_rtl(ws) -> None:
    """
    Set worksheet to right-to-left mode for Hebrew content.

    Args:
        ws: openpyxl Worksheet object
    """
    ws.sheet_view.rightToLeft = True


def style_header_row(ws, row: int = 1, font=None, fill=None, alignment=None, border=None) -> None:
    """
    Apply standard header styling to a row.

    Args:
        ws: openpyxl Worksheet object
        row: Row number to style (default: 1)
        font: Font to use (default: HEADER_FONT)
        fill: Fill to use (default: HEADER_FILL)
        alignment: Alignment to use (default: RTL_ALIGNMENT)
        border: Border to use (default: THIN_BORDER)
    """
    font = font or HEADER_FONT
    fill = fill or HEADER_FILL
    alignment = alignment or RTL_ALIGNMENT
    border = border or THIN_BORDER

    for cell in ws[row]:
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border


def style_data_cells(ws, start_row: int = 2, wrap_columns: set = None, alignment=None, border=None) -> None:
    """
    Apply standard styling to data cells.

    Args:
        ws: openpyxl Worksheet object
        start_row: First data row (default: 2)
        wrap_columns: Set of column indices (1-based) that should wrap text
        alignment: Default alignment (default: RTL_ALIGNMENT)
        border: Border to apply (default: THIN_BORDER)
    """
    wrap_columns = wrap_columns or set()
    alignment = alignment or RTL_ALIGNMENT
    border = border or THIN_BORDER

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.col_idx in wrap_columns:
                cell.alignment = WRAP_ALIGNMENT
            else:
                cell.alignment = alignment
            cell.border = border


def apply_alternating_stripes(ws, start_row: int, end_row: int, num_columns: int, stripe_fill=None) -> None:
    """
    Apply alternating row stripes to data rows.

    Args:
        ws: openpyxl Worksheet object
        start_row: First row to stripe
        end_row: Last row to stripe
        num_columns: Number of columns to style
        stripe_fill: Fill pattern for striped rows (default: STRIPE_FILL)
    """
    stripe_fill = stripe_fill or STRIPE_FILL

    for row_idx in range(start_row, end_row + 1):
        row_offset = row_idx - start_row
        if row_offset % 2 == 1:
            for col_idx in range(1, num_columns + 1):
                cell = ws.cell(row_idx, col_idx)
                # Only apply if cell doesn't have error font (red color)
                if cell.font.color is None or getattr(cell.font.color, 'rgb', None) != "00FF0000":
                    cell.fill = stripe_fill


def apply_pass_fail_fill(cell, passed: bool) -> None:
    """
    Apply pass/fail fill color to a cell.

    Args:
        cell: openpyxl Cell object
        passed: True for pass (green), False for fail (red)
    """
    cell.fill = PASS_FILL if passed else FAIL_FILL


def auto_fit_column_width(ws, column_letter: str, min_width: float = 8, max_width: float = 50) -> None:
    """
    Auto-fit a column width based on content.

    Args:
        ws: openpyxl Worksheet object
        column_letter: Column letter (e.g., 'A', 'B')
        min_width: Minimum column width
        max_width: Maximum column width
    """
    max_length = 0
    for cell in ws[column_letter]:
        try:
            cell_length = len(str(cell.value)) if cell.value else 0
            max_length = max(max_length, cell_length)
        except (TypeError, AttributeError):
            pass

    adjusted_width = min(max(max_length + 2, min_width), max_width)
    ws.column_dimensions[column_letter].width = adjusted_width


def set_column_widths(ws, widths: dict) -> None:
    """
    Set multiple column widths at once.

    Args:
        ws: openpyxl Worksheet object
        widths: Dict mapping column letters to widths, e.g., {'A': 15, 'B': 20}
    """
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def calculate_hebrew_text_width(text: str, base_width: float = 1.0) -> float:
    """
    Calculate approximate column width for Hebrew text.

    Hebrew characters are generally wider than ASCII characters in most fonts.

    Args:
        text: Text to measure
        base_width: Base width multiplier

    Returns:
        Approximate width value for openpyxl
    """
    if not text:
        return 0

    width = 0
    for char in str(text):
        if '\u0590' <= char <= '\u05FF':  # Hebrew
            width += 1.05
        elif ord(char) > 0x4E00:  # CJK
            width += 2.0
        else:
            width += 1.0

    return width * base_width
