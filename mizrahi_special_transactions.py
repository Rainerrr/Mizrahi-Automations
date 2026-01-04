#!/usr/bin/env python3
"""
Mizrahi Special Transactions - CLI Processor (single script)

Implements checks #1–#7 from the current specification.

Inputs:
  - Mutual Funds List (XLSX): filtered to Mizrahi trustee funds (by 'שם נאמן')
  - Manager special transactions report (CSV or XLSX): e.g., 1702431.csv or 1702431.xlsx
  - Specification table (XLSX, optional): for פירוט בדיקות sheet

Outputs:
  - Output XLSX: summary + check statuses + exceptions + samples (+ out-of-scope funds)
  - Email JSON: for n8n workflow - contains ONLY two JSON objects with transaction info (no full email body)

Dependencies:
  - Python 3.10+
  - openpyxl

Example:
  python mizrahi_special_transactions.py \
    --mutual-funds-list "Mutual Funds List.xlsx" \
    --input-report "1702431.csv" \
    --output-xlsx "output.xlsx" \
    --email-json "email.json" \
    --manager-name "איילון" \
    --spec-file "Special Transactions Report Testing Specifications.xlsx" \
    --price-threshold 5.0 \
    --seed 123

  # Without manager name (column will be empty):
  python mizrahi_special_transactions.py \
    --mutual-funds-list "Mutual Funds List.xlsx" \
    --input-report "1702431.csv" \
    --output-xlsx "output.xlsx" \
    --email-json "email.json" \
    --price-threshold 3.5 \
    --seed 123

Notes:
  - Hebrew column headers are expected (as in your provided files).
  - Dates in the manager report are often stored as numbers DDMMYYYY without leading zeros.
  - Times are often stored as numbers HHMMSS without leading zeros.
  - CSV files are expected to be UTF-8 encoded (BOM is handled automatically).
  - Unique ID for transactions is built from security number + date (as per specification step 1).
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import logging
import random
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl
import uuid
from datetime import datetime as datetime_module

# Global log directory for current run (will be set in main)
LOG_RUN_DIR: Optional[Path] = None

# Loggers for each specification
logger = logging.getLogger(__name__)  # Main logger
logger_chk1 = logging.getLogger('CHK_1')  # Inter-fund transactions
logger_chk3 = logging.getLogger('CHK_3')  # Date validation
logger_chk4 = logging.getLogger('CHK_4')  # Decision method rules
logger_chk4c = logging.getLogger('CHK_4C')  # Price/type consistency for same security+date+time
logger_chk6_price = logging.getLogger('CHK_6_PRICE')  # TASE price checks
logger_chk6_limit = logging.getLogger('CHK_6_LIMIT')  # Price > 100 checks
logger_chk6_url_fail = logging.getLogger('CHK_6_URL_FAIL')  # Failed URL fetches (for manual verification)
logger_chk7 = logging.getLogger('CHK_7')  # Problematic securities


def setup_logging(log_base_dir: Path = Path("log")) -> Path:
    """Set up logging with separate files for each specification.

    Creates a unique directory for each run with separate log files:
    - main.log: General processing log
    - chk1_inter_fund.log: Inter-fund transaction checks
    - chk3_date.log: Date validation checks
    - chk4_decision.log: Decision method rule checks
    - chk6_tase_price.log: TASE price comparison checks
    - chk6_price_limit.log: Price > 100 checks
    - chk7_problematic.log: Problematic securities checks

    Returns:
        Path to the run directory
    """
    global LOG_RUN_DIR

    # Generate unique run ID: timestamp + short UUID
    timestamp = datetime_module.now().strftime("%Y%m%d_%H%M%S")
    short_uuid = str(uuid.uuid4())[:8]
    run_id = f"{timestamp}_{short_uuid}"

    # Create log directory structure
    run_dir = log_base_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    LOG_RUN_DIR = run_dir

    # Common log format
    log_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # Detailed format for specification logs (includes more context)
    detailed_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(name)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # Console handler for main logger only
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(log_format)

    # Set up main logger
    logger.setLevel(logging.DEBUG)
    logger.addHandler(console_handler)
    main_file_handler = logging.FileHandler(run_dir / "main.log", encoding='utf-8')
    main_file_handler.setLevel(logging.DEBUG)
    main_file_handler.setFormatter(log_format)
    logger.addHandler(main_file_handler)

    # Set up specification-specific loggers
    spec_loggers = [
        (logger_chk1, "chk1_inter_fund.log", "CHK_1 - Inter-fund Transactions"),
        (logger_chk3, "chk3_date.log", "CHK_3 - Date Validation"),
        (logger_chk4, "chk4_decision.log", "CHK_4 - Decision Method Rules"),
        (logger_chk4c, "chk4c_price_type_consistency.log", "CHK_4C - Price/Type Consistency"),
        (logger_chk6_price, "chk6_tase_price.log", "CHK_6 - TASE Price Checks"),
        (logger_chk6_limit, "chk6_price_limit.log", "CHK_6 - Price > 100 Checks"),
        (logger_chk6_url_fail, "chk6_failed_urls.log", "CHK_6 - Failed URL Fetches (Manual Verification Required)"),
        (logger_chk7, "chk7_problematic.log", "CHK_7 - Problematic Securities"),
    ]

    for spec_logger, filename, description in spec_loggers:
        spec_logger.setLevel(logging.DEBUG)

        # File handler for this specification
        file_handler = logging.FileHandler(run_dir / filename, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(detailed_format)
        spec_logger.addHandler(file_handler)

        # Also add to main log file
        spec_logger.addHandler(main_file_handler)

        # Write header to the log file
        spec_logger.info("=" * 70)
        spec_logger.info(description)
        spec_logger.info("=" * 70)

    logger.info("Log directory created: %s", run_dir)
    return run_dir
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Selenium imports (optional - for price checks)
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# -----------------------------
# Configuration (column headers)
# -----------------------------

MIZRAHI_TRUSTEE_NAME_DEFAULT = 'מזרחי טפחות חברה לנאמנות בע"מ'

# Mutual Funds List headers
MF_COL_FUND_ID = "מספר בורסה"
MF_COL_TRUSTEE = "שם נאמן"

# Manager report headers
R_COL_FUND_NO = "מספר קרן"
R_COL_FUND_NAME = "שם קרן"
R_COL_SECURITY_NAME = "שם נייר"
R_COL_SECURITY_NO = "מספר נייר"
R_COL_QUANTITY = "כמות"
R_COL_PRICE = "מחיר"
R_COL_DATE = "תאריך"
R_COL_TIME = "שעה"
R_COL_TYPE = "סוג"
R_COL_DECISION = "אופן החלטה"
R_COL_REPORT_DATE = "ת.דוח"  # used to infer month if --report-month omitted

# Decision-method rules (check #4)
TYPE_REQUIRES_DECISION_1 = {12, 22}
TYPE_REQUIRES_DECISION_1_OR_2 = {31, 32, 33, 34, 35, 36}

# Price checks (spec #6)
TASE_PRICE_CHECK_TYPES = {12, 21, 22}  # Types requiring TASE price comparison
TASE_SAMPLES_PER_TYPE = 2
TASE_VARIANCE_THRESHOLD_DEFAULT = 5.0  # 5% (in percent, will be converted to decimal)
PRICE_LIMIT_TYPES = {31, 32, 33, 34, 35, 36}  # Types with price > 100 check
PRICE_LIMIT = 100.0

# Maximum exceptions before smart sampling is applied
MAX_EXCEPTIONS_BEFORE_SAMPLING = 100

# Problematic securities lists (spec #7)
PROBLEMATIC_LISTS_CONFIG = {
    'low_liquidity': {
        'url': 'https://market.tase.co.il/he/market_data/securities/data/all?dType=1&cl1=0&cl2=2',
        'name_he': 'דלי סחירות',
    },
    'maintenance': {
        'url': 'https://market.tase.co.il/he/market_data/securities/data/all?dType=1&cl1=0&cl2=3',
        'name_he': 'רשימת שימור',
    },
    'suspended': {
        'url': 'https://market.tase.co.il/he/market_data/securities/data/all?dType=1&cl1=0&cl2=4',
        'name_he': 'מושעים',
    },
}

# Excel styling constants (matching fund_automation_complete.py)
HEADER_FONT = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
DEFAULT_FONT = Font(name='Calibri', size=11)

# Headers that should have wrap_text=True (long text fields)
# Other columns will have wrap_text=False to keep rows compact
WRAP_HEADERS = {
    "שם קרן",
    "שם נייר",
    "סיבה",
    "רשימות בעייתיות",
    "סיבה סבירות",
    "הערות",
    "שדה",  # Summary sheet field column
    "ערך",  # Summary sheet value column
}

# Hebrew month names
HEBREW_MONTHS = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר"
}


def _format_report_month_hebrew(report_month: str) -> str:
    """Convert YYYY-MM to Hebrew format like 'ספטמבר 2025'."""
    try:
        year, month = report_month.split("-")
        return f"{HEBREW_MONTHS[int(month)]} {year}"
    except (ValueError, KeyError):
        return report_month


# -----------------------------
# Data structures
# -----------------------------

@dataclass(frozen=True)
class TxnRow:
    row_num: int  # row number in the original sheet (1-based)
    fund_no: Optional[int]
    fund_name: Optional[str]
    security_name: Optional[str]
    security_no: Optional[str]
    quantity: Optional[float]
    price: Optional[float]
    tx_date: Optional[dt.date]
    tx_time: Optional[dt.time]
    tx_type: Optional[int]
    decision_method: Optional[int]
    report_date: Optional[dt.date]

    @property
    def unique_id(self) -> str:
        """Spec: unique id = security number + date."""
        d = self.tx_date.strftime("%d%m%Y") if self.tx_date else ""
        s = self.security_no or ""
        return f"{s}|{d}"


@dataclass(frozen=True)
class ExceptionRow:
    check_id: str
    reason: str
    row: TxnRow
    group_key: str = ""


@dataclass(frozen=True)
class Samples:
    decision_1: Optional[TxnRow]
    decision_2: Optional[TxnRow]


@dataclass
class PriceCheckResult:
    """Result of TASE price comparison check."""
    row: TxnRow
    tase_closing_price: Optional[float] = None
    variance_pct: Optional[float] = None
    is_exception: bool = False
    error_message: Optional[str] = None


@dataclass
class PriceLimitResult:
    """Result of price > 100 check for types 31-36."""
    row: TxnRow
    is_exception: bool = False


@dataclass
class ProblematicSecurityResult:
    """Result of checking a transaction against problematic lists."""
    row: TxnRow
    matched_lists: list = field(default_factory=list)
    is_exception: bool = False


# -----------------------------
# Helpers: parsing / normalization
# -----------------------------

def _norm_spaces(s: str) -> str:
    return " ".join((s or "").strip().split())


def _to_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None


def _parse_ddmmyyyy(v: Any) -> Optional[dt.date]:
    """Parse manager report numeric date DDMMYYYY with zero-padding."""
    if v is None or v == "":
        return None
    s = str(int(v)) if isinstance(v, (int, float)) else str(v)
    s = s.strip().zfill(8)
    try:
        day = int(s[0:2])
        month = int(s[2:4])
        year = int(s[4:8])
        return dt.date(year, month, day)
    except Exception:
        return None


def _parse_hhmmss(v: Any) -> Optional[dt.time]:
    """Parse manager report numeric time HHMMSS with zero-padding."""
    if v is None or v == "":
        return None
    s = str(int(v)) if isinstance(v, (int, float)) else str(v)
    s = s.strip().zfill(6)
    try:
        hh = int(s[0:2])
        mm = int(s[2:4])
        ss = int(s[4:6])
        return dt.time(hh, mm, ss)
    except Exception:
        return None


def _headers_row1(ws) -> dict[str, int]:
    """Return mapping from header text to column index, assuming headers are in row 1."""
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c
    return headers


# -----------------------------
# Loaders
# -----------------------------

def load_mizrahi_fund_ids(mutual_funds_list_path: Path, trustee_name: str) -> set[int]:
    """Load fund IDs under Mizrahi trusteeship from the Mutual Funds List.

    Uses:
      - fund id column: מספר בורסה
      - trustee name column: שם נאמן
    """
    wb = openpyxl.load_workbook(mutual_funds_list_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = _headers_row1(ws)
    fund_id_col = headers.get(MF_COL_FUND_ID)
    trustee_col = headers.get(MF_COL_TRUSTEE)

    if fund_id_col is None or trustee_col is None:
        wb.close()
        raise ValueError(f"Mutual Funds List must contain columns: '{MF_COL_FUND_ID}', '{MF_COL_TRUSTEE}'")

    target = _norm_spaces(trustee_name)
    fund_ids: set[int] = set()

    for r in range(2, ws.max_row + 1):
        trustee = ws.cell(r, trustee_col).value
        if trustee is None:
            continue
        if _norm_spaces(str(trustee)) != target:
            continue

        fid_val = ws.cell(r, fund_id_col).value
        try:
            fid = int(float(fid_val))
        except Exception:
            continue

        fund_ids.add(fid)

    wb.close()
    return fund_ids


def load_manager_report_xlsx(input_report_path: Path) -> tuple[list[TxnRow], dict]:
    """Load manager special-transactions report from XLSX. Expects headers in row 1."""
    wb = openpyxl.load_workbook(input_report_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = _headers_row1(ws)

    def col(name: str) -> Optional[int]:
        return headers.get(name)

    required = [R_COL_FUND_NO, R_COL_SECURITY_NO, R_COL_QUANTITY, R_COL_PRICE, R_COL_DATE, R_COL_TIME, R_COL_TYPE, R_COL_DECISION]
    missing = [r for r in required if col(r) is None]
    if missing:
        wb.close()
        raise ValueError(f"Manager report missing required columns: {missing}")

    rows: list[TxnRow] = []
    rec_col = headers.get("מס. רשומה")  # optional, helps skip trailing note rows

    for r in range(2, ws.max_row + 1):
        rec = ws.cell(r, rec_col).value if rec_col else None
        fund_no = _to_int(ws.cell(r, col(R_COL_FUND_NO)).value)
        sec_no = _to_str(ws.cell(r, col(R_COL_SECURITY_NO)).value)

        if fund_no is None and sec_no is None and rec is None:
            continue

        row = TxnRow(
            row_num=r,
            fund_no=fund_no,
            fund_name=_to_str(ws.cell(r, col(R_COL_FUND_NAME)).value) if col(R_COL_FUND_NAME) else None,
            security_name=_to_str(ws.cell(r, col(R_COL_SECURITY_NAME)).value) if col(R_COL_SECURITY_NAME) else None,
            security_no=sec_no,
            quantity=_to_float(ws.cell(r, col(R_COL_QUANTITY)).value),
            price=_to_float(ws.cell(r, col(R_COL_PRICE)).value),
            tx_date=_parse_ddmmyyyy(ws.cell(r, col(R_COL_DATE)).value),
            tx_time=_parse_hhmmss(ws.cell(r, col(R_COL_TIME)).value),
            tx_type=_to_int(ws.cell(r, col(R_COL_TYPE)).value),
            decision_method=_to_int(ws.cell(r, col(R_COL_DECISION)).value),
            report_date=_parse_ddmmyyyy(ws.cell(r, col(R_COL_REPORT_DATE)).value) if col(R_COL_REPORT_DATE) else None,
        )
        rows.append(row)

    inferred_month = None
    for row in rows:
        if row.report_date:
            inferred_month = f"{row.report_date.year:04d}-{row.report_date.month:02d}"
            break

    wb.close()
    meta = {"sheet": ws.title, "rows_parsed": len(rows), "report_month_inferred": inferred_month}
    return rows, meta


def load_manager_report_csv(input_report_path: Path) -> tuple[list[TxnRow], dict]:
    """Load manager special-transactions report from CSV. Expects headers in row 1.

    The CSV may have a BOM (byte order mark) which is handled by utf-8-sig encoding.
    """
    rows: list[TxnRow] = []

    with open(input_report_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)

        # Normalize header names (strip whitespace)
        if reader.fieldnames:
            reader.fieldnames = [h.strip() for h in reader.fieldnames]

        required = [R_COL_FUND_NO, R_COL_SECURITY_NO, R_COL_QUANTITY, R_COL_PRICE, R_COL_DATE, R_COL_TIME, R_COL_TYPE, R_COL_DECISION]
        missing = [r for r in required if r not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(f"Manager report CSV missing required columns: {missing}")

        for row_num, csv_row in enumerate(reader, start=2):  # start=2 because row 1 is header
            fund_no = _to_int(csv_row.get(R_COL_FUND_NO))
            sec_no = _to_str(csv_row.get(R_COL_SECURITY_NO))
            rec = csv_row.get("מס. רשומה")

            if fund_no is None and sec_no is None and not rec:
                continue

            row = TxnRow(
                row_num=row_num,
                fund_no=fund_no,
                fund_name=_to_str(csv_row.get(R_COL_FUND_NAME)),
                security_name=_to_str(csv_row.get(R_COL_SECURITY_NAME)),
                security_no=sec_no,
                quantity=_to_float(csv_row.get(R_COL_QUANTITY)),
                price=_to_float(csv_row.get(R_COL_PRICE)),
                tx_date=_parse_ddmmyyyy(csv_row.get(R_COL_DATE)),
                tx_time=_parse_hhmmss(csv_row.get(R_COL_TIME)),
                tx_type=_to_int(csv_row.get(R_COL_TYPE)),
                decision_method=_to_int(csv_row.get(R_COL_DECISION)),
                report_date=_parse_ddmmyyyy(csv_row.get(R_COL_REPORT_DATE)),
            )
            rows.append(row)

    inferred_month = None
    for row in rows:
        if row.report_date:
            inferred_month = f"{row.report_date.year:04d}-{row.report_date.month:02d}"
            break

    meta = {"source": str(input_report_path), "rows_parsed": len(rows), "report_month_inferred": inferred_month}
    return rows, meta


def load_manager_report(input_report_path: Path) -> tuple[list[TxnRow], dict]:
    """Load manager special-transactions report. Detects format by file extension."""
    suffix = input_report_path.suffix.lower()
    if suffix == '.csv':
        return load_manager_report_csv(input_report_path)
    elif suffix in ('.xlsx', '.xls'):
        return load_manager_report_xlsx(input_report_path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}. Use .csv or .xlsx")


# -----------------------------
# Checks (spec 1–5.1)
# -----------------------------

def check_1_duplicates_exact(rows: list[TxnRow]) -> list[ExceptionRow]:
    """Spec #1: identical date/time/security number/quantity/price -> flag all rows in the group."""
    logger.info("CHK_1 (Exact Duplicates): Starting check on %d rows", len(rows))
    buckets: dict[tuple, list[TxnRow]] = defaultdict(list)
    for r in rows:
        key = (
            r.tx_date.isoformat() if r.tx_date else "",
            r.tx_time.isoformat() if r.tx_time else "",
            r.security_no or "",
            r.quantity if r.quantity is not None else "",
            r.price if r.price is not None else "",
        )
        buckets[key].append(r)

    out: list[ExceptionRow] = []
    for key, group in buckets.items():
        if len(group) <= 1:
            continue
        group_key = "|".join(map(str, key))
        for row in group:
            logger.warning(
                "CHK_1 EXCEPTION: DUPLICATE_EXACT - Row %d: security_no=%s, date=%s, time=%s, qty=%s, price=%s | "
                "Group has %d identical rows | Reason: Multiple rows share identical date/time/security/quantity/price",
                row.row_num, row.security_no, row.tx_date, row.tx_time, row.quantity, row.price, len(group)
            )
            out.append(ExceptionRow(check_id="CHK_1", reason="DUPLICATE_EXACT", row=row, group_key=group_key))

    logger.info("CHK_1 (Exact Duplicates): Completed - found %d exceptions", len(out))
    return out


def check_1_abs_quantity_pairs(rows: list[TxnRow]) -> list[ExceptionRow]:
    """Spec #1: within unique_id(security+date), if there are two rows with same abs(quantity) but DIFFERENT SIGNS -> flag.

    Only flags when one quantity is positive and the other is negative (inter-fund transactions).
    Does NOT flag if both quantities have the same sign.
    """
    logger_chk1.info("Starting inter-fund transaction check on %d rows", len(rows))
    logger.info("CHK_1 (Inter-fund Transactions): Starting check on %d rows", len(rows))

    by_uid: dict[str, list[TxnRow]] = defaultdict(list)
    for r in rows:
        by_uid[r.unique_id].append(r)

    logger_chk1.info("Grouped into %d unique IDs (security+date)", len(by_uid))

    out: list[ExceptionRow] = []
    for uid, group in by_uid.items():
        abs_map: dict[float, list[TxnRow]] = defaultdict(list)
        for r in group:
            if r.quantity is None or r.quantity == 0:
                continue
            abs_map[abs(r.quantity)].append(r)

        for abs_qty, rs in abs_map.items():
            # Check if there are rows with OPPOSITE signs (one positive, one negative)
            has_positive = any(r.quantity > 0 for r in rs if r.quantity is not None)
            has_negative = any(r.quantity < 0 for r in rs if r.quantity is not None)

            # Only flag if we have BOTH positive and negative quantities with same abs value
            if has_positive and has_negative:
                positive_rows = [r for r in rs if r.quantity is not None and r.quantity > 0]
                negative_rows = [r for r in rs if r.quantity is not None and r.quantity < 0]
                for r in rs:
                    logger_chk1.warning(
                        "EXCEPTION FOUND:\n"
                        "  Row Number: %d\n"
                        "  Security Number: %s\n"
                        "  Security Name: %s\n"
                        "  Fund Number: %s\n"
                        "  Fund Name: %s\n"
                        "  Transaction Date: %s\n"
                        "  Quantity: %s\n"
                        "  Price: %s\n"
                        "  Unique ID: %s\n"
                        "  Absolute Quantity: %s\n"
                        "  Positive quantity rows: %s\n"
                        "  Negative quantity rows: %s\n"
                        "  Reason: Matching abs(quantity) with opposite signs indicates inter-fund transaction",
                        r.row_num, r.security_no, r.security_name, r.fund_no, r.fund_name,
                        r.tx_date, r.quantity, r.price, uid, abs_qty,
                        [pr.row_num for pr in positive_rows],
                        [nr.row_num for nr in negative_rows]
                    )
                    out.append(ExceptionRow(check_id="CHK_1", reason="עסקה בין קרנות", row=r, group_key=f"{uid}|abs={abs_qty}"))

    logger_chk1.info("Check completed - found %d exceptions", len(out))
    logger.info("CHK_1 (Inter-fund Transactions): Completed - found %d exceptions", len(out))
    return out


def check_3_dates_in_report_month(rows: list[TxnRow], report_month: str) -> list[ExceptionRow]:
    """Spec #3: tx_date must be within same month (YYYY-MM)."""
    logger_chk3.info("Starting date validation check on %d rows", len(rows))
    logger_chk3.info("Expected report month: %s", report_month)
    logger.info("CHK_3 (Date in Report Month): Starting check on %d rows for month %s", len(rows), report_month)

    y, m = report_month.split("-")
    year = int(y)
    month = int(m)

    out: list[ExceptionRow] = []
    for r in rows:
        if r.tx_date is None:
            logger_chk3.warning(
                "EXCEPTION FOUND - MISSING DATE:\n"
                "  Row Number: %d\n"
                "  Security Number: %s\n"
                "  Security Name: %s\n"
                "  Fund Number: %s\n"
                "  Fund Name: %s\n"
                "  Expected Month: %s\n"
                "  Actual Date: None\n"
                "  Reason: Transaction date is missing/null",
                r.row_num, r.security_no, r.security_name, r.fund_no, r.fund_name, report_month
            )
            out.append(ExceptionRow(check_id="CHK_3", reason="MISSING_TX_DATE", row=r, group_key=report_month))
            continue
        if r.tx_date.year != year or r.tx_date.month != month:
            logger_chk3.warning(
                "EXCEPTION FOUND - DATE OUT OF RANGE:\n"
                "  Row Number: %d\n"
                "  Security Number: %s\n"
                "  Security Name: %s\n"
                "  Fund Number: %s\n"
                "  Fund Name: %s\n"
                "  Expected Month: %s (year=%d, month=%d)\n"
                "  Actual Date: %s (year=%d, month=%d)\n"
                "  Reason: Transaction date does not match report month",
                r.row_num, r.security_no, r.security_name, r.fund_no, r.fund_name,
                report_month, year, month, r.tx_date, r.tx_date.year, r.tx_date.month
            )
            out.append(ExceptionRow(check_id="CHK_3", reason="DATE_OUT_OF_REPORT_MONTH", row=r, group_key=report_month))

    logger_chk3.info("Check completed - found %d exceptions", len(out))
    logger.info("CHK_3 (Date in Report Month): Completed - found %d exceptions", len(out))
    return out


def check_4_decision_method_rules(rows: list[TxnRow]) -> list[ExceptionRow]:
    """Spec #4: decision method allowed values depend on type."""
    logger_chk4.info("Starting decision method rules check on %d rows", len(rows))
    logger_chk4.info("Types requiring decision_method=1: %s", TYPE_REQUIRES_DECISION_1)
    logger_chk4.info("Types requiring decision_method=1 or 2: %s", TYPE_REQUIRES_DECISION_1_OR_2)
    logger.info("CHK_4 (Decision Method Rules): Starting check on %d rows", len(rows))

    out: list[ExceptionRow] = []
    for r in rows:
        if r.tx_type is None or r.decision_method is None:
            logger_chk4.warning(
                "EXCEPTION FOUND - MISSING DATA:\n"
                "  Row Number: %d\n"
                "  Security Number: %s\n"
                "  Security Name: %s\n"
                "  Fund Number: %s\n"
                "  Fund Name: %s\n"
                "  Transaction Type: %s\n"
                "  Decision Method: %s\n"
                "  Reason: Transaction type or decision method is missing/null",
                r.row_num, r.security_no, r.security_name, r.fund_no, r.fund_name,
                r.tx_type, r.decision_method
            )
            out.append(ExceptionRow(check_id="CHK_4", reason="MISSING_TYPE_OR_DECISION_METHOD", row=r))
            continue

        if r.tx_type in TYPE_REQUIRES_DECISION_1 and r.decision_method != 1:
            logger_chk4.warning(
                "EXCEPTION FOUND - WRONG DECISION METHOD:\n"
                "  Row Number: %d\n"
                "  Security Number: %s\n"
                "  Security Name: %s\n"
                "  Fund Number: %s\n"
                "  Fund Name: %s\n"
                "  Transaction Type: %d\n"
                "  Decision Method: %d\n"
                "  Required Decision Method: 1\n"
                "  Reason: Type %d requires decision_method=1, but got %d",
                r.row_num, r.security_no, r.security_name, r.fund_no, r.fund_name,
                r.tx_type, r.decision_method, r.tx_type, r.decision_method
            )
            out.append(ExceptionRow(check_id="CHK_4", reason=f"TYPE_{r.tx_type}_REQUIRES_DECISION_1", row=r, group_key=f"type={r.tx_type}"))

        if r.tx_type in TYPE_REQUIRES_DECISION_1_OR_2 and r.decision_method not in (1, 2):
            logger_chk4.warning(
                "EXCEPTION FOUND - WRONG DECISION METHOD:\n"
                "  Row Number: %d\n"
                "  Security Number: %s\n"
                "  Security Name: %s\n"
                "  Fund Number: %s\n"
                "  Fund Name: %s\n"
                "  Transaction Type: %d\n"
                "  Decision Method: %d\n"
                "  Required Decision Method: 1 or 2\n"
                "  Reason: Type %d requires decision_method=1 or 2, but got %d",
                r.row_num, r.security_no, r.security_name, r.fund_no, r.fund_name,
                r.tx_type, r.decision_method, r.tx_type, r.decision_method
            )
            out.append(ExceptionRow(check_id="CHK_4", reason=f"TYPE_{r.tx_type}_REQUIRES_DECISION_1_OR_2", row=r, group_key=f"type={r.tx_type}"))

    logger_chk4.info("Check completed - found %d exceptions", len(out))
    logger.info("CHK_4 (Decision Method Rules): Completed - found %d exceptions", len(out))
    return out


def check_4c_price_type_consistency(rows: list[TxnRow]) -> list[ExceptionRow]:
    """Spec #4c: For transactions with same security_no, date, and time, verify price and type are identical.

    Groups transactions by (security_no, tx_date, tx_time) and flags groups where:
    - Price is not identical across all rows in the group, OR
    - Type is not identical across all rows in the group
    """
    logger_chk4c.info("Starting price/type consistency check on %d rows", len(rows))
    logger.info("CHK_4C (Price/Type Consistency): Starting check on %d rows", len(rows))

    # Group by (security_no, date, time)
    groups: dict[tuple, list[TxnRow]] = defaultdict(list)
    for r in rows:
        if r.security_no is None or r.tx_date is None or r.tx_time is None:
            continue
        key = (r.security_no, r.tx_date.isoformat() if r.tx_date else "", r.tx_time.isoformat() if r.tx_time else "")
        groups[key].append(r)

    logger_chk4c.info("Grouped into %d unique (security+date+time) combinations", len(groups))

    out: list[ExceptionRow] = []
    for key, group in groups.items():
        if len(group) <= 1:
            continue

        # Check price consistency
        prices = set(r.price for r in group if r.price is not None)
        types = set(r.tx_type for r in group if r.tx_type is not None)

        price_mismatch = len(prices) > 1
        type_mismatch = len(types) > 1

        if price_mismatch or type_mismatch:
            group_key = f"{key[0]}|{key[1]}|{key[2]}"

            # Determine reason
            reasons = []
            if price_mismatch:
                reasons.append(f"מחירים שונים: {sorted(prices)}")
            if type_mismatch:
                reasons.append(f"סוגים שונים: {sorted(types)}")
            reason = "; ".join(reasons)

            for row in group:
                logger_chk4c.warning(
                    "EXCEPTION FOUND - INCONSISTENT PRICE/TYPE:\n"
                    "  Row Number: %d\n"
                    "  Security Number: %s\n"
                    "  Security Name: %s\n"
                    "  Fund Number: %s\n"
                    "  Fund Name: %s\n"
                    "  Transaction Date: %s\n"
                    "  Transaction Time: %s\n"
                    "  Price: %s\n"
                    "  Type: %s\n"
                    "  Group Size: %d\n"
                    "  Prices in Group: %s\n"
                    "  Types in Group: %s\n"
                    "  Reason: %s",
                    row.row_num, row.security_no, row.security_name, row.fund_no, row.fund_name,
                    row.tx_date, row.tx_time, row.price, row.tx_type, len(group),
                    sorted(prices), sorted(types), reason
                )
                out.append(ExceptionRow(check_id="CHK_4C", reason=reason, row=row, group_key=group_key))

    logger_chk4c.info("Check completed - found %d exceptions", len(out))
    logger.info("CHK_4C (Price/Type Consistency): Completed - found %d exceptions", len(out))
    return out


def pick_samples(valid_rows: list[TxnRow], seed: Optional[int]) -> Samples:
    """Spec #5: random transaction with decision method 1 and 2 from valid lines.

    Samples one transaction with decision_method=1 and one with decision_method=2
    from the valid (non-exception) rows.
    """
    logger.info("CHK_5 (Sampling): Starting sample selection from %d valid rows", len(valid_rows))

    rng = random.Random(seed)
    dm1 = [r for r in valid_rows if r.decision_method == 1]
    dm2 = [r for r in valid_rows if r.decision_method == 2]

    logger.info("CHK_5 (Sampling): Found %d rows with decision_method=1", len(dm1))
    logger.info("CHK_5 (Sampling): Found %d rows with decision_method=2", len(dm2))

    s1 = rng.choice(dm1) if dm1 else None
    s2 = rng.choice(dm2) if dm2 else None

    if s1:
        logger.info("CHK_5 (Sampling): Selected decision_method=1 sample: row %d, security=%s, fund=%s",
                   s1.row_num, s1.security_no, s1.fund_no)
    else:
        logger.warning("CHK_5 (Sampling): No valid rows with decision_method=1 available for sampling")

    if s2:
        logger.info("CHK_5 (Sampling): Selected decision_method=2 sample: row %d, security=%s, fund=%s",
                   s2.row_num, s2.security_no, s2.fund_no)
    else:
        logger.warning("CHK_5 (Sampling): No valid rows with decision_method=2 available for sampling")

    return Samples(decision_1=s1, decision_2=s2)


def smart_sample_exceptions(exceptions: list[ExceptionRow], max_count: int, seed: Optional[int] = None) -> tuple[list[ExceptionRow], int]:
    """Smart sample exceptions when count exceeds max_count.

    Uses stratified sampling to ensure representation across different:
    - Reasons (stratify by reason type)
    - Securities (try to include diverse securities)
    - Funds (try to include diverse funds)

    Args:
        exceptions: List of exception rows
        max_count: Maximum number of exceptions to return
        seed: Random seed for reproducibility

    Returns:
        Tuple of (sampled exceptions, original count)
    """
    original_count = len(exceptions)

    if original_count <= max_count:
        return exceptions, original_count

    logger.info("Smart sampling %d exceptions down to max %d", original_count, max_count)

    rng = random.Random(seed)

    # Group by reason for stratified sampling
    by_reason: dict[str, list[ExceptionRow]] = defaultdict(list)
    for ex in exceptions:
        by_reason[ex.reason].append(ex)

    sampled: list[ExceptionRow] = []
    reasons = list(by_reason.keys())

    # First pass: take proportional samples from each reason
    for reason in reasons:
        reason_exceptions = by_reason[reason]
        # Proportional allocation with minimum of 1
        proportion = len(reason_exceptions) / original_count
        target_count = max(1, int(proportion * max_count))

        if len(reason_exceptions) <= target_count:
            sampled.extend(reason_exceptions)
        else:
            sampled.extend(rng.sample(reason_exceptions, target_count))

    # If we have room for more, add randomly from remaining
    remaining_budget = max_count - len(sampled)
    if remaining_budget > 0:
        sampled_set = {(e.row.row_num, e.reason) for e in sampled}
        unsampled = [e for e in exceptions if (e.row.row_num, e.reason) not in sampled_set]
        if unsampled:
            additional = rng.sample(unsampled, min(remaining_budget, len(unsampled)))
            sampled.extend(additional)

    # If we're over budget, trim down
    if len(sampled) > max_count:
        sampled = rng.sample(sampled, max_count)

    # Sort by row number for consistent output
    sampled.sort(key=lambda e: e.row.row_num)

    logger.info("Smart sampling complete: %d -> %d exceptions", original_count, len(sampled))
    return sampled, original_count


# -----------------------------
# Checks (spec #6 - Price)
# -----------------------------

def _init_selenium_driver():
    """Initialize Selenium Chrome driver."""
    if not SELENIUM_AVAILABLE:
        return None
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def _fetch_tase_closing_price(driver, security_no: str, tx_date: dt.date) -> tuple[Optional[float], Optional[str]]:
    """Fetch closing price from TASE website."""
    date_str = tx_date.strftime('%Y-%m-%d')
    url = f'https://market.tase.co.il/he/market_data/security/{security_no}/historical_data/eod?pType=8&oId=0{security_no}&dFrom={date_str}&dTo={date_str}'

    logger_chk6_price.info("=" * 50)
    logger_chk6_price.info("Fetching TASE closing price")
    logger_chk6_price.info("  Security Number: %s", security_no)
    logger_chk6_price.info("  Transaction Date: %s", tx_date)
    logger_chk6_price.info("  URL: %s", url)

    try:
        driver.get(url)
        time.sleep(3)
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'table')))
        except:
            pass
        time.sleep(2)

        body = driver.find_element(By.TAG_NAME, 'body')
        text = body.text

        logger_chk6_price.debug("Page body text (first 500 chars): %s", text[:500] if text else "EMPTY")

        if "לא נמצאו תוצאות" in text or "אין נתונים" in text:
            logger_chk6_price.warning("No data found on TASE website for this security/date")
            return None, "לא נמצאו נתונים לתאריך זה"

        tables = driver.find_elements(By.TAG_NAME, 'table')
        if not tables:
            logger_chk6_price.warning("No table element found on page")
            return None, "לא נמצאה טבלה"

        logger_chk6_price.info("Found %d table(s) on page", len(tables))

        for table_idx, table in enumerate(tables):
            rows = table.find_elements(By.TAG_NAME, 'tr')
            logger_chk6_price.debug("Table %d has %d rows", table_idx, len(rows))
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if cells:
                    row_text = [cell.text.strip() for cell in cells]
                    logger_chk6_price.debug("Scraped row: %s", row_text)
                    date_formatted = tx_date.strftime('%d/%m/%Y')
                    if row_text and date_formatted in row_text[0]:
                        logger_chk6_price.info("Found matching date row!")
                        logger_chk6_price.info("  Scraped data: %s", row_text)
                        if len(row_text) >= 2:
                            try:
                                price_str = row_text[1].replace(',', '')
                                price = float(price_str)
                                logger_chk6_price.info("  Parsed closing price: %s", price)
                                logger_chk6_price.info("  SUCCESS: Price retrieved")
                                return price, None
                            except ValueError:
                                logger_chk6_price.error("Failed to parse price value: '%s'", row_text[1])
                                return None, f"לא ניתן לפרסר מחיר: {row_text[1]}"

        logger_chk6_price.warning("Date %s not found in any table", date_formatted)
        return None, "התאריך לא נמצא בטבלה"
    except Exception as e:
        logger_chk6_price.error("Exception while fetching: %s", str(e))
        return None, f"שגיאה: {str(e)}"


def check_6_tase_prices(rows: list[TxnRow], seed: Optional[int] = None, variance_threshold_pct: float = TASE_VARIANCE_THRESHOLD_DEFAULT) -> list[PriceCheckResult]:
    """Spec #6 Part 1: Sample transactions of types 12,21,22 and compare with TASE closing price.

    Exception is raised ONLY when transaction price is HIGHER than TASE closing price by more than threshold.
    Transaction price being LOWER than closing price is acceptable.

    Failed URL fetches are logged to a separate file for manual verification and NOT included in results.

    Args:
        rows: List of transaction rows to check
        seed: Random seed for reproducible sampling
        variance_threshold_pct: Maximum allowed price above TASE closing price in percent (e.g., 5.0 for 5%)
    """
    logger_chk6_price.info("Starting TASE price comparison check")
    logger_chk6_price.info("Threshold (max price above closing): %.2f%%", variance_threshold_pct)
    logger_chk6_price.info("Target transaction types: %s", TASE_PRICE_CHECK_TYPES)
    logger_chk6_price.info("Samples per type: %d", TASE_SAMPLES_PER_TYPE)
    logger.info("CHK_6 (TASE Prices): Starting TASE price comparison check")

    if not SELENIUM_AVAILABLE:
        logger_chk6_price.warning("Selenium not available - skipping TASE price checks")
        logger.warning("CHK_6 (TASE Prices): Selenium not available, skipping TASE price checks")
        return []

    results: list[PriceCheckResult] = []
    rng = random.Random(seed)
    variance_threshold_decimal = variance_threshold_pct / 100.0

    # Select samples for each type
    samples_by_type: dict[int, list[TxnRow]] = {}
    for tx_type in TASE_PRICE_CHECK_TYPES:
        type_rows = [r for r in rows if r.tx_type == tx_type and r.security_no and r.tx_date and r.price]
        if type_rows:
            selected = rng.sample(type_rows, min(TASE_SAMPLES_PER_TYPE, len(type_rows)))
            samples_by_type[tx_type] = selected
            logger_chk6_price.info("Type %d: %d eligible rows, selected %d samples", tx_type, len(type_rows), len(selected))
            for s in selected:
                logger_chk6_price.info("  - Row %d: security=%s, date=%s, price=%.4f", s.row_num, s.security_no, s.tx_date, s.price)

    if not any(samples_by_type.values()):
        logger_chk6_price.warning("No eligible transactions found for TASE price checks")
        return results

    driver = _init_selenium_driver()
    if not driver:
        logger_chk6_price.error("Failed to initialize Selenium driver")
        return results

    failed_url_count = 0
    try:
        for tx_type, txns in samples_by_type.items():
            for txn in txns:
                closing_price, error = _fetch_tase_closing_price(driver, txn.security_no, txn.tx_date)
                if error:
                    # Log failed URLs to separate file for manual verification - NOT added to results
                    failed_url_count += 1
                    date_str = txn.tx_date.strftime('%Y-%m-%d') if txn.tx_date else 'N/A'
                    url = f'https://market.tase.co.il/he/market_data/security/{txn.security_no}/historical_data/eod?pType=8&oId=0{txn.security_no}&dFrom={date_str}&dTo={date_str}'
                    logger_chk6_url_fail.warning(
                        "FAILED URL - MANUAL VERIFICATION REQUIRED:\n"
                        "  URL: %s\n"
                        "  Row Number: %d\n"
                        "  Security Number: %s\n"
                        "  Security Name: %s\n"
                        "  Fund Number: %s\n"
                        "  Transaction Date: %s\n"
                        "  Transaction Price: %s\n"
                        "  Error: %s",
                        url, txn.row_num, txn.security_no, txn.security_name, txn.fund_no,
                        txn.tx_date, txn.price, error
                    )
                    logger_chk6_price.info(
                        "URL FETCH FAILED (logged to chk6_failed_urls.log for manual verification):\n"
                        "  Row Number: %d\n"
                        "  Security Number: %s\n"
                        "  Error: %s",
                        txn.row_num, txn.security_no, error
                    )
                else:
                    # Calculate how much HIGHER the transaction price is compared to closing price
                    # Only flag as exception if price is HIGHER by more than threshold
                    # Price being LOWER is acceptable
                    if closing_price and closing_price > 0:
                        price_above_pct = ((txn.price - closing_price) / closing_price) * 100
                        is_exception = price_above_pct > variance_threshold_pct
                    else:
                        price_above_pct = 0.0
                        is_exception = False

                    if is_exception:
                        logger_chk6_price.warning(
                            "EXCEPTION - PRICE TOO HIGH:\n"
                            "  Row Number: %d\n"
                            "  Security Number: %s\n"
                            "  Security Name: %s\n"
                            "  Fund Number: %s\n"
                            "  Transaction Date: %s\n"
                            "  Transaction Price: %.4f\n"
                            "  TASE Closing Price: %.4f\n"
                            "  Price Above Closing: %.2f%%\n"
                            "  Threshold: %.2f%%\n"
                            "  Reason: Transaction price exceeds TASE closing price by more than threshold",
                            txn.row_num, txn.security_no, txn.security_name, txn.fund_no,
                            txn.tx_date, txn.price, closing_price, price_above_pct, variance_threshold_pct
                        )
                    else:
                        status = "BELOW CLOSING" if price_above_pct < 0 else "WITHIN THRESHOLD"
                        logger_chk6_price.info(
                            "PASSED (%s):\n"
                            "  Row Number: %d\n"
                            "  Security Number: %s\n"
                            "  Transaction Price: %.4f\n"
                            "  TASE Closing Price: %.4f\n"
                            "  Price vs Closing: %.2f%% (threshold: +%.2f%%)",
                            status, txn.row_num, txn.security_no, txn.price, closing_price,
                            price_above_pct, variance_threshold_pct
                        )
                    results.append(PriceCheckResult(
                        row=txn,
                        tase_closing_price=closing_price,
                        variance_pct=price_above_pct,
                        is_exception=is_exception
                    ))
    finally:
        driver.quit()

    exception_count = sum(1 for r in results if r.is_exception)
    logger_chk6_price.info("Check completed - checked %d transactions, %d exceptions, %d failed URLs", len(results), exception_count, failed_url_count)
    logger.info("CHK_6 (TASE Prices): Completed - checked %d transactions, %d exceptions, %d failed URLs (see chk6_failed_urls.log)", len(results), exception_count, failed_url_count)
    return results


def check_6_price_limits(rows: list[TxnRow]) -> list[PriceLimitResult]:
    """Spec #6 Part 2: Check types 31-36 for price > 100."""
    logger_chk6_limit.info("Starting price > 100 check")
    logger_chk6_limit.info("Target transaction types: %s", PRICE_LIMIT_TYPES)
    logger_chk6_limit.info("Price limit threshold: %.2f", PRICE_LIMIT)
    logger.info("CHK_6 (Price Limits): Starting price > 100 check for types %s", PRICE_LIMIT_TYPES)

    results: list[PriceLimitResult] = []
    checked_count = 0
    for row in rows:
        if row.tx_type in PRICE_LIMIT_TYPES and row.price is not None:
            checked_count += 1
            is_exception = row.price > PRICE_LIMIT
            if is_exception:
                logger_chk6_limit.warning(
                    "EXCEPTION FOUND - PRICE > 100:\n"
                    "  Row Number: %d\n"
                    "  Security Number: %s\n"
                    "  Security Name: %s\n"
                    "  Fund Number: %s\n"
                    "  Fund Name: %s\n"
                    "  Transaction Type: %d\n"
                    "  Price: %.4f\n"
                    "  Price Limit: %.2f\n"
                    "  Reason: Transaction type %d should not have price exceeding %.2f",
                    row.row_num, row.security_no, row.security_name, row.fund_no, row.fund_name,
                    row.tx_type, row.price, PRICE_LIMIT, row.tx_type, PRICE_LIMIT
                )
                results.append(PriceLimitResult(row=row, is_exception=True))

    logger_chk6_limit.info("Check completed - checked %d transactions, found %d exceptions", checked_count, len(results))
    logger.info("CHK_6 (Price Limits): Completed - checked %d transactions, found %d exceptions", checked_count, len(results))
    return results


# -----------------------------
# Checks (spec #7 - Problematic Securities)
# -----------------------------

def _fetch_problematic_list(driver, list_type: str, url: str) -> set[str]:
    """Fetch a problematic securities list from TASE website."""
    logger_chk7.info("=" * 50)
    logger_chk7.info("Fetching problematic list: %s", list_type)
    logger_chk7.info("  URL: %s", url)

    securities: set[str] = set()
    try:
        driver.get(url)
        time.sleep(3)
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'table')))
        except:
            pass
        time.sleep(2)

        tables = driver.find_elements(By.TAG_NAME, 'table')
        logger_chk7.info("Found %d table(s) on page", len(tables))

        for table_idx, table in enumerate(tables):
            rows = table.find_elements(By.TAG_NAME, 'tr')
            logger_chk7.debug("Table %d has %d rows", table_idx, len(rows))
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) >= 4:
                    cell_texts = [cell.text.strip() for cell in cells]
                    logger_chk7.debug("Scraped row: %s", cell_texts)
                    for text in cell_texts:
                        if re.match(r'^\d{7}$', text):
                            securities.add(text)
                            logger_chk7.debug("Found security number: %s", text)
                            break

        logger_chk7.info("Successfully fetched list '%s'", list_type)
        logger_chk7.info("  Securities found: %d", len(securities))
        if securities:
            logger_chk7.info("  Sample (first 10): %s", list(securities)[:10])
    except Exception as e:
        logger_chk7.error("Error fetching '%s': %s", list_type, str(e))
    return securities


def fetch_problematic_lists(cache_path: Optional[Path] = None) -> dict[str, set[str]]:
    """Fetch all problematic securities lists from TASE."""
    logger_chk7.info("Starting fetch of all problematic securities lists")
    logger.info("CHK_7 (Problematic Lists): Starting fetch of all problematic lists")

    # Try to load from cache
    if cache_path and cache_path.exists():
        try:
            logger_chk7.info("Loading from cache: %s", cache_path)
            cached = json.loads(cache_path.read_text(encoding='utf-8'))
            result = {k: set(v) for k, v in cached.items()}
            for list_type, securities in result.items():
                logger_chk7.info("  Loaded '%s': %d securities", list_type, len(securities))
            return result
        except Exception as e:
            logger_chk7.warning("Failed to load from cache: %s", str(e))

    if not SELENIUM_AVAILABLE:
        logger_chk7.warning("Selenium not available - skipping problematic securities fetch")
        return {}

    driver = _init_selenium_driver()
    if not driver:
        logger_chk7.error("Failed to initialize Selenium driver")
        return {}

    all_lists: dict[str, set[str]] = {}
    try:
        for list_type, config in PROBLEMATIC_LISTS_CONFIG.items():
            logger_chk7.info("Fetching '%s' (%s)...", config['name_he'], list_type)
            securities = _fetch_problematic_list(driver, list_type, config['url'])
            all_lists[list_type] = securities
    finally:
        driver.quit()

    # Save to cache
    if cache_path:
        logger_chk7.info("Saving to cache: %s", cache_path)
        cache_data = {k: list(v) for k, v in all_lists.items()}
        cache_path.write_text(json.dumps(cache_data, ensure_ascii=False, indent=2), encoding='utf-8')

    total_securities = sum(len(s) for s in all_lists.values())
    logger_chk7.info("Fetch complete - %d lists, %d total securities", len(all_lists), total_securities)
    logger.info("CHK_7 (Problematic Lists): Fetch complete - %d lists, %d total securities", len(all_lists), total_securities)
    return all_lists


def check_7_problematic_securities(rows: list[TxnRow], problematic_lists: dict[str, set[str]]) -> list[ProblematicSecurityResult]:
    """Spec #7: Check all transactions against problematic securities lists."""
    logger_chk7.info("Starting problematic securities check on %d rows", len(rows))
    logger.info("CHK_7 (Problematic Securities): Starting check on %d rows", len(rows))

    for list_type, securities in problematic_lists.items():
        logger_chk7.info("Checking against '%s' list (%d securities)",
                        PROBLEMATIC_LISTS_CONFIG[list_type]['name_he'], len(securities))

    results: list[ProblematicSecurityResult] = []
    checked_count = 0

    for row in rows:
        if not row.security_no:
            continue

        checked_count += 1
        matched_lists = []
        for list_type, security_nos in problematic_lists.items():
            if row.security_no in security_nos:
                matched_lists.append(PROBLEMATIC_LISTS_CONFIG[list_type]['name_he'])

        if matched_lists:
            logger_chk7.warning(
                "EXCEPTION FOUND - PROBLEMATIC SECURITY:\n"
                "  Row Number: %d\n"
                "  Security Number: %s\n"
                "  Security Name: %s\n"
                "  Fund Number: %s\n"
                "  Fund Name: %s\n"
                "  Transaction Date: %s\n"
                "  Transaction Type: %s\n"
                "  Quantity: %s\n"
                "  Price: %s\n"
                "  Matched Lists: %s\n"
                "  Reason: Security appears in problematic lists",
                row.row_num, row.security_no, row.security_name, row.fund_no, row.fund_name,
                row.tx_date, row.tx_type, row.quantity, row.price, ", ".join(matched_lists)
            )
            results.append(ProblematicSecurityResult(
                row=row,
                matched_lists=matched_lists,
                is_exception=True
            ))

    logger_chk7.info("Check completed - checked %d transactions, found %d exceptions", checked_count, len(results))
    logger.info("CHK_7 (Problematic Securities): Completed - checked %d transactions, found %d exceptions", checked_count, len(results))
    return results


def build_email_json(samples: Samples) -> list[dict[str, Any]]:
    """Spec #5.1: JSON file should contain only valid JSON objects with transaction info.

    Output is a list containing only valid samples (no empty objects):
      - If decision method 1 sample exists, include it
      - If decision method 2 sample exists, include it
      - If neither exists, return empty list

    Each object contains:
      Fund number, Fund name, Security name, Security number, Quantity, Price, Date, Type, Decision method
    """
    def txn_obj(row: TxnRow) -> dict[str, Any]:
        return {
            "fund_number": row.fund_no,
            "fund_name": row.fund_name,
            "security_name": row.security_name,
            "security_number": row.security_no,
            "quantity": row.quantity,
            "price": row.price,
            "date": row.tx_date.strftime("%d-%m-%Y") if row.tx_date else None,
            "type": row.tx_type,
            "decision_method": row.decision_method,
        }

    result = []
    if samples.decision_1 is not None:
        result.append(txn_obj(samples.decision_1))
    if samples.decision_2 is not None:
        result.append(txn_obj(samples.decision_2))
    return result


# -----------------------------
# Output writing (XLSX)
# -----------------------------

def _rtl(ws) -> None:
    ws.sheet_view.rightToLeft = True


def _style_header(ws, row: int = 1) -> None:
    """Apply header styling to specified row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_cells(ws, start_row: int = 2) -> None:
    """Apply styling to all data cells.

    Only columns with headers in WRAP_HEADERS will have wrap_text=True.
    This prevents numeric/date columns from causing tall rows.
    """
    # Determine which columns should wrap based on header text in row 1
    wrap_cols: set[int] = set()
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(1, col_idx)
        header_text = str(header_cell.value).strip() if header_cell.value is not None else ""
        if header_text in WRAP_HEADERS:
            wrap_cols.add(col_idx)

    # Apply styling to data cells
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            wrap = cell.col_idx in wrap_cols
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=wrap)
            cell.border = THIN_BORDER


def _calculate_text_width(text: str, is_bold: bool = False) -> float:
    """Calculate approximate display width of text in Excel units.

    Excel column width is based on the number of characters that fit using the default font.
    For Calibri 11pt (default), each character is approximately 1 unit wide.
    Hebrew characters are slightly wider, spaces are narrower.
    """
    if not text:
        return 0.0

    width = 0.0
    for char in text:
        if char == " ":
            # Spaces are narrower
            width += 0.5
        elif '\u0590' <= char <= '\u05FF':
            # Hebrew characters - reduced from 1.3 to 1.05 to prevent inflation
            width += 1.05
        elif char in '０１２３４５６７８９':
            # Full-width digits
            width += 2.0
        elif ord(char) > 0x4E00:
            # CJK characters
            width += 2.0
        elif char in 'WMwm':
            # Wide Latin characters
            width += 1.2
        elif char in 'il|!.,;:\'"':
            # Narrow characters
            width += 0.6
        else:
            width += 1.0

    # Bold text is slightly wider
    if is_bold:
        width *= 1.05

    return width


def _auto_fit_columns(ws, min_width: float = 8.0, max_width: float = 50.0, padding: float = 1.2) -> dict[str, float]:
    """Auto-fit column widths based on content - simulates Excel double-click behavior.

    Returns a dict mapping column letters to their calculated widths for use by row height calculation.

    Each column gets its own optimal width based on the widest content in that column.

    Args:
        ws: Worksheet to process
        min_width: Minimum column width
        max_width: Maximum column width
        padding: Padding to add to calculated width (reduced from 2.5 to 1.2)
    """
    column_widths: dict[str, float] = {}

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_width_found = 0.0
        widest_cell_text = ""

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)

            if cell.value is None:
                continue

            cell_text = str(cell.value)
            is_bold = cell.font.bold if cell.font else False
            is_header = (row_idx == 1)

            # For multi-line content, find the longest line
            lines = cell_text.split('\n')
            for line in lines:
                line_width = _calculate_text_width(line, is_bold or is_header)
                if line_width > max_width_found:
                    max_width_found = line_width
                    widest_cell_text = line[:50]  # Truncate for logging

        # Add padding (reduced from 2.5 to 1.2)
        final_width = max_width_found + padding

        # Apply min/max bounds
        was_capped = final_width > max_width
        final_width = max(min_width, min(final_width, max_width))

        # Debug log when column hits max_width
        if was_capped:
            logger.debug(
                "Column %s in sheet '%s' hit max_width (%.1f): calculated=%.1f, text='%s'",
                column_letter, ws.title, max_width, max_width_found + padding, widest_cell_text
            )

        ws.column_dimensions[column_letter].width = final_width
        column_widths[column_letter] = final_width

    return column_widths


def _auto_fit_rows(ws, column_widths: dict[str, float] = None, line_height: float = 15.0, header_line_height: float = 18.0, min_height: float = 15.0, max_height: float = 120.0) -> None:
    """Auto-fit row heights based on content - simulates Excel double-click behavior.

    Each row gets its own optimal height based on:
    - Number of explicit line breaks in cell content
    - Text wrapping based on actual column widths
    - Different heights for header vs data rows

    Args:
        ws: Worksheet to process
        column_widths: Dict of column letter -> width (from _auto_fit_columns)
        line_height: Height per line for data rows (pixels)
        header_line_height: Height per line for header row (pixels)
        min_height: Minimum row height
        max_height: Maximum row height to prevent extremely tall rows (reduced from 200 to 120)
    """
    # If column_widths not provided, read from worksheet
    if column_widths is None:
        column_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            column_widths[col_letter] = width if width else 10.0

    for row_idx in range(1, ws.max_row + 1):
        max_lines_needed = 1
        is_header = (row_idx == 1)
        tallest_cell_text = ""
        tallest_cell_col = ""

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)

            if cell.value is None:
                continue

            cell_text = str(cell.value)
            col_letter = get_column_letter(col_idx)
            col_width = column_widths.get(col_letter, 10.0)

            # Check if cell has wrap_text enabled
            has_wrap = cell.alignment.wrap_text if cell.alignment else False

            # Calculate lines for this cell
            lines_in_cell = 0

            for line in cell_text.split('\n'):
                if not line:
                    lines_in_cell += 1
                    continue

                # Calculate text width using the same logic as column width calculation
                is_bold = cell.font.bold if cell.font else False
                text_width = _calculate_text_width(line, is_bold or is_header)

                # Available width for text (column width minus padding)
                # Raised minimum from 4.0 to 6.0 to prevent extreme line counts in narrow columns
                available_width = max(col_width - 1.0, 6.0)

                if has_wrap and text_width > available_width:
                    # Text will wrap - calculate how many lines needed
                    wrapped_lines = int((text_width / available_width) + 0.99)  # Ceiling
                    lines_in_cell += max(1, wrapped_lines)
                else:
                    lines_in_cell += 1

            if lines_in_cell > max_lines_needed:
                max_lines_needed = lines_in_cell
                tallest_cell_text = cell_text[:50]  # Truncate for logging
                tallest_cell_col = col_letter

        # Calculate row height based on number of lines
        if is_header:
            row_height = header_line_height * max_lines_needed
        else:
            row_height = line_height * max_lines_needed

        # Apply min/max bounds
        was_capped = row_height > max_height
        row_height = max(min_height, min(row_height, max_height))

        # Debug log when row hits max_height
        if was_capped:
            logger.debug(
                "Row %d in sheet '%s' hit max_height (%.1f): calculated=%.1f, col=%s, text='%s'",
                row_idx, ws.title, max_height, line_height * max_lines_needed, tallest_cell_col, tallest_cell_text
            )

        ws.row_dimensions[row_idx].height = row_height


def _set_font_calibri(ws) -> None:
    """Set Calibri font for all cells in worksheet."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                # Preserve other font properties (bold, color) but change name to Calibri
                cell.font = Font(
                    name='Calibri',
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                    size=cell.font.size or 11
                )
            else:
                cell.font = Font(name='Calibri', size=11)


def _header(ws, headers: list[str], row: int = 1) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i)
        cell.value = h
    _style_header(ws, row)
    if row == 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def _fmt_date(d: Optional[dt.date]) -> str:
    return d.strftime("%d-%m-%Y") if d else ""


def _fmt_time(t: Optional[dt.time]) -> str:
    return t.strftime("%H:%M:%S") if t else ""


def _txn_to_basic_list(r: TxnRow) -> list[Any]:
    return [
        r.fund_no,
        r.fund_name,
        r.security_name,
        r.security_no,
        r.quantity,
        r.price,
        _fmt_date(r.tx_date),
        _fmt_time(r.tx_time),
        r.tx_type,
        r.decision_method,
    ]


def _copy_spec_sheet_to_workbook(spec_file_path: Path, target_wb: openpyxl.Workbook, sheet_name: str = "פירוט בדיקות") -> None:
    """Copy specification sheet to target workbook preserving all original styling."""
    if not spec_file_path or not spec_file_path.exists():
        return

    from copy import copy

    src_wb = openpyxl.load_workbook(spec_file_path)
    # Use Hebrew sheet if available, otherwise first sheet
    src_sheet_name = 'Hebrew' if 'Hebrew' in src_wb.sheetnames else src_wb.sheetnames[0]
    src_ws = src_wb[src_sheet_name]

    # Create new sheet in target workbook
    target_ws = target_wb.create_sheet(sheet_name)
    target_ws.sheet_view.rightToLeft = True

    # Copy column dimensions
    for col_letter, col_dim in src_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
        target_ws.column_dimensions[col_letter].hidden = col_dim.hidden

    # Copy row dimensions
    for row_num, row_dim in src_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height
        target_ws.row_dimensions[row_num].hidden = row_dim.hidden

    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Copy all cells with their values and styles
    for row in src_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy cell styles
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.alignment = copy(cell.alignment)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)

    src_wb.close()


def write_output_xlsx(
    output_path: Path,
    *,
    report_month: str,
    manager_name: str,
    trustee_name: str,
    summary: dict[str, Any],
    exceptions_duplicates: list[ExceptionRow],
    exceptions_date: list[ExceptionRow],
    exceptions_decision: list[ExceptionRow],
    exceptions_price_type: list[ExceptionRow] = None,
    samples: Samples,
    in_scope_funds: set[int],
    price_check_results: list[PriceCheckResult] = None,
    price_limit_results: list[PriceLimitResult] = None,
    problematic_security_results: list[ProblematicSecurityResult] = None,
    spec_file_path: Path = None,
) -> None:
    wb = openpyxl.Workbook()

    # Validation columns to add to each sheet
    VALIDATION_COLS = ["טופל?", "שם הבודק"]

    # Extract counts for סטטוס בדיקות table
    count_inter_fund = summary.get("חריגות עסקאות בין קרנות", 0)
    count_date = summary.get("חריגות תאריך", 0)
    count_decision = summary.get("חריגות אופן החלטה", 0)
    count_price_type = summary.get("חריגות מחיר/סוג לא עקביים", 0)
    count_price_limit = summary.get("חריגות מחיר מעל 100", 0)
    count_problematic = summary.get("חריגות ניירות בעייתיים", 0)
    # Count TASE price exceptions
    count_tase_price = len([r for r in (price_check_results or []) if r.is_exception])

    # Sheet 1: Summary (סיכום) - new format
    ws_sum = wb.active
    ws_sum.title = "סיכום"
    _rtl(ws_sum)
    _header(ws_sum, ["שדה", "ערך"])

    # Build new summary rows in the required order
    hebrew_month = _format_report_month_hebrew(report_month)
    summary_rows = [
        ("מנהל קרן", manager_name),
        ("נאמן", trustee_name),
        ("חודש נבדק", f"דוח חודשי-{hebrew_month}"),
        ("מספר קרנות של מנהל הקרן", summary.get("מספר קרנות של מנהל הקרן", "")),
        ("מספר קרנות של מנהל הקרן – בנאמנות מזרחי", summary.get("מספר קרנות של מנהל הקרן – בנאמנות מזרחי", "")),
        ("מספר קרנות של מנהל הקרן – בניהול חיצוני", summary.get("קרנות מחוץ לתחום", "")),
    ]

    rr = 2
    for k, v in summary_rows:
        ws_sum.cell(rr, 1).value = k
        ws_sum.cell(rr, 2).value = v
        rr += 1

    _style_cells(ws_sum)

    # Sheet 2: סטטוס בדיקות (Check Status)
    ws_checks = wb.create_sheet("סטטוס בדיקות")
    _rtl(ws_checks)
    ws_checks.append(["בדיקה", "תיאור", "סטטוס", "חריגות", "טופל?", "שם הבודק"])
    _style_header(ws_checks, 1)

    # Define all checks with their descriptions (with specification numbers)
    # Note: Spec #2 (out-of-scope funds) is shown in summary table, not here
    check_statuses = [
        ("בדיקה #1 - חריגות עסקאות בין קרנות", "עסקאות עם כמות מנוגדת באותו יום", count_inter_fund == 0, count_inter_fund),
        ("בדיקה #3 - חריגות תאריך", "עסקאות מחוץ לחודש הדוח", count_date == 0, count_date),
        ("בדיקה #4 - חריגות אופן החלטה", "אי-התאמה בין סוג לאופן החלטה", count_decision == 0, count_decision),
        ("בדיקה #4c - חריגות מחיר/סוג", "אותו נייר+תאריך+שעה עם מחיר/סוג שונה", count_price_type == 0, count_price_type),
        ("בדיקה #5 - דגימה לבדיקה – אופן החלטה 1", "קיימת דגימה תקינה עם אופן החלטה 1" if samples.decision_1 is not None else "לא קיימת דגימה תקינה עם אופן החלטה 1", samples.decision_1 is not None, 0 if samples.decision_1 is not None else 1),
        ("בדיקה #5 - דגימה לבדיקה – אופן החלטה 2", "קיימת דגימה תקינה עם אופן החלטה 2" if samples.decision_2 is not None else "לא קיימת דגימה תקינה עם אופן החלטה 2", samples.decision_2 is not None, 0 if samples.decision_2 is not None else 1),
        ("בדיקה #6 - חריגות סטיית מחיר מבורסה", "סטייה מעל סף אחוז ממחיר סגירה בבורסה", count_tase_price == 0, count_tase_price),
        ("בדיקה #6 - חריגות מחיר מעל 100", "עסקאות מסוג 31-36 עם מחיר > 100", count_price_limit == 0, count_price_limit),
        ("בדיקה #7 - חריגות ניירות בעייתיים", "ניירות ברשימות דלי סחירות/שימור/מושעים", count_problematic == 0, count_problematic),
    ]

    for row_idx, (name, description, passed, count) in enumerate(check_statuses, start=2):
        ws_checks.append([name, description, "✓ תקין" if passed else "✗ חריגה", count, "", ""])
        fill = PASS_FILL if passed else FAIL_FILL
        for col in range(1, 5):
            ws_checks.cell(row=row_idx, column=col).fill = fill

    _style_cells(ws_checks)

    # Track which optional sheets are created
    optional_sheets = []

    # Note: Spec #2 (out-of-scope funds) count is shown in summary table, no separate sheet

    # Exceptions - duplicates (inter-fund transactions) - only create if there are exceptions
    ws_dup = None
    if exceptions_duplicates:
        ws_dup = wb.create_sheet("בדיקה #1 - עסקאות בין קרנות")
        _rtl(ws_dup)
        _header(
            ws_dup,
            [
                "בדיקה",
                "סיבה",
                "מפתח קבוצה",
                "מספר קרן",
                "שם קרן",
                "קרן של מזרחי?",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "שורה בקובץ",
            ] + VALIDATION_COLS,
        )
        for ex in exceptions_duplicates:
            is_mizrahi = "כן" if ex.row.fund_no in in_scope_funds else "לא"
            basic_list = _txn_to_basic_list(ex.row)
            # Insert is_mizrahi after fund_name (index 1 in basic_list corresponds to שם קרן)
            row_data = [ex.check_id, ex.reason, ex.group_key, basic_list[0], basic_list[1], is_mizrahi] + basic_list[2:] + [ex.row.row_num, "", ""]
            ws_dup.append(row_data)
        optional_sheets.append(ws_dup)

    # Exceptions - date - only create if there are exceptions
    ws_date = None
    if exceptions_date:
        ws_date = wb.create_sheet("בדיקה #3 - תאריך")
        _rtl(ws_date)
        _header(
            ws_date,
            [
                "בדיקה",
                "סיבה",
                "מספר קרן",
                "שם קרן",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "שורה בקובץ",
            ] + VALIDATION_COLS,
        )
        for ex in exceptions_date:
            ws_date.append([ex.check_id, ex.reason, *_txn_to_basic_list(ex.row), ex.row.row_num, "", ""])
        optional_sheets.append(ws_date)

    # Exceptions - decision method - only create if there are exceptions
    ws_dm = None
    if exceptions_decision:
        ws_dm = wb.create_sheet("בדיקה #4 - אופן החלטה")
        _rtl(ws_dm)
        _header(
            ws_dm,
            [
                "בדיקה",
                "סיבה",
                "מספר קרן",
                "שם קרן",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "שורה בקובץ",
            ] + VALIDATION_COLS,
        )
        for ex in exceptions_decision:
            ws_dm.append([ex.check_id, ex.reason, *_txn_to_basic_list(ex.row), ex.row.row_num, "", ""])
        optional_sheets.append(ws_dm)

    # Exceptions - price/type consistency - only create if there are exceptions
    ws_price_type = None
    if exceptions_price_type:
        ws_price_type = wb.create_sheet("בדיקה #4c - מחיר/סוג")
        _rtl(ws_price_type)
        _header(
            ws_price_type,
            [
                "בדיקה",
                "סיבה",
                "מפתח קבוצה",
                "מספר קרן",
                "שם קרן",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "שורה בקובץ",
            ] + VALIDATION_COLS,
        )
        for ex in exceptions_price_type:
            ws_price_type.append([ex.check_id, ex.reason, ex.group_key, *_txn_to_basic_list(ex.row), ex.row.row_num, "", ""])
        optional_sheets.append(ws_price_type)

    # Spec #6: Price checks (TASE price variance and price > 100) - merged into single sheet
    ws_price = None
    price_exceptions = [r for r in (price_check_results or []) if r.is_exception]
    has_price_exceptions = price_exceptions or price_limit_results
    if has_price_exceptions:
        ws_price = wb.create_sheet("בדיקה #6 - חריגות מחיר")
        _rtl(ws_price)
        _header(
            ws_price,
            [
                "סוג בדיקה",
                "מספר קרן",
                "שם קרן",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר בעסקה",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "מחיר סגירה בורסה",
                "סטייה באחוזים",
                "שורה בקובץ",
                "הערה",
            ] + VALIDATION_COLS,
        )
        # Add TASE price variance exceptions
        for r in price_exceptions:
            ws_price.append([
                "סטיית מחיר מבורסה",
                r.row.fund_no,
                r.row.fund_name,
                r.row.security_name,
                r.row.security_no,
                r.row.quantity,
                r.row.price,
                _fmt_date(r.row.tx_date),
                _fmt_time(r.row.tx_time),
                r.row.tx_type,
                r.row.decision_method,
                r.tase_closing_price,
                f"{r.variance_pct:.2f}%" if r.variance_pct is not None else "",
                r.row.row_num,
                r.error_message or "",
                "", ""
            ])
        # Add price > 100 exceptions
        for r in (price_limit_results or []):
            ws_price.append([
                "מחיר מעל 100",
                r.row.fund_no,
                r.row.fund_name,
                r.row.security_name,
                r.row.security_no,
                r.row.quantity,
                r.row.price,
                _fmt_date(r.row.tx_date),
                _fmt_time(r.row.tx_time),
                r.row.tx_type,
                r.row.decision_method,
                "",  # No TASE price for this check
                "",  # No variance for this check
                r.row.row_num,
                "",
                "", ""
            ])
        optional_sheets.append(ws_price)

    # Spec #7: Problematic securities - only create if there are exceptions
    ws_prob = None
    if problematic_security_results:
        ws_prob = wb.create_sheet("בדיקה #7 - ניירות בעייתיים")
        _rtl(ws_prob)
        _header(
            ws_prob,
            [
                "מספר קרן",
                "שם קרן",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "רשימות בעייתיות",
                "שורה בקובץ",
            ] + VALIDATION_COLS,
        )
        for r in problematic_security_results:
            ws_prob.append([
                *_txn_to_basic_list(r.row),
                ", ".join(r.matched_lists),
                r.row.row_num,
                "", ""
            ])
        optional_sheets.append(ws_prob)

    # Samples - only create if there are samples
    ws_s = None
    has_samples = samples.decision_1 is not None or samples.decision_2 is not None
    if has_samples:
        ws_s = wb.create_sheet("בדיקה #5 - דגימות לבדיקה")
        _rtl(ws_s)
        _header(
            ws_s,
            [
                "קבוצה",
                "מספר קרן",
                "שם קרן",
                "שם נייר",
                "מספר נייר",
                "כמות",
                "מחיר",
                "תאריך",
                "שעה",
                "סוג",
                "אופן החלטה",
                "תאריך החלטת דירקטוריון",
                "סבירות החלטה",
                "ציות לנוהל מנהל",
            ] + VALIDATION_COLS,
        )

        def add_sample(label: str, row: Optional[TxnRow]) -> None:
            if row:
                ws_s.append([label, *_txn_to_basic_list(row), "", "", "", "", ""])

        # Only add samples that exist (no empty rows)
        add_sample("אופן החלטה = 1", samples.decision_1)
        add_sample("אופן החלטה = 2", samples.decision_2)
        optional_sheets.append(ws_s)

    # Apply styling to all optional data sheets that were created
    for ws in optional_sheets:
        _style_cells(ws)

    # Sheet: פירוט בדיקות (Specification details) - copy with original styling preserved
    ws_spec = None
    if spec_file_path and spec_file_path.exists():
        _copy_spec_sheet_to_workbook(spec_file_path, wb, "פירוט בדיקות")
        ws_spec = wb["פירוט בדיקות"]

    # Set Calibri font BEFORE measuring widths/heights (measurement is tuned for Calibri)
    # Then auto-fit columns and rows for all sheets (only include sheets that were created)
    all_sheets = [ws_sum, ws_checks] + optional_sheets
    for ws in all_sheets:
        # 1. Set font first (measurement is calibrated for Calibri)
        _set_font_calibri(ws)
        # 2. Calculate and set column widths, get the widths dict
        column_widths = _auto_fit_columns(ws)
        # 3. Calculate row heights using the column widths
        _auto_fit_rows(ws, column_widths=column_widths)

    # Apply auto-fit to פירוט בדיקות (keeps original styling but adjusts dimensions)
    if ws_spec:
        _set_font_calibri(ws_spec)
        spec_column_widths = _auto_fit_columns(ws_spec)
        _auto_fit_rows(ws_spec, column_widths=spec_column_widths)

    # Reorder sheets according to specification:
    # 1. סיכום, 2. פירוט בדיקות, 3. סטטוס בדיקות, then rest in spec order
    # Note: Spec #2 (out-of-scope funds) count is shown in summary table, no separate sheet
    desired_order = [
        "סיכום",                              # Summary
        "פירוט בדיקות",                       # Specification details
        "סטטוס בדיקות",                       # Check status
        "בדיקה #1 - עסקאות בין קרנות",        # Spec #1
        "בדיקה #3 - תאריך",                   # Spec #3
        "בדיקה #4 - אופן החלטה",              # Spec #4
        "בדיקה #4c - מחיר/סוג",               # Spec #4c
        "בדיקה #5 - דגימות לבדיקה",           # Spec #5
        "בדיקה #6 - חריגות מחיר",             # Spec #6 (merged price checks)
        "בדיקה #7 - ניירות בעייתיים",         # Spec #7
    ]

    # Move sheets to correct positions
    for idx, sheet_name in enumerate(desired_order):
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            wb.move_sheet(sheet, offset=idx - wb.sheetnames.index(sheet_name))

    wb.save(output_path)
    wb.close()


# -----------------------------
# CLI
# -----------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Mizrahi Special Transactions - CLI processor (checks #1–#7).")
    p.add_argument("--mutual-funds-list", required=True, type=Path, help="Path to 'Mutual Funds List.xlsx'")
    p.add_argument("--input-report", required=True, type=Path, help="Path to manager report CSV or XLSX (e.g., 1702431.csv)")
    p.add_argument("--output-xlsx", required=True, type=Path, help="Path to write output XLSX")
    p.add_argument("--email-json", required=True, type=Path, help="Path to write email JSON payload (two objects only)")
    p.add_argument("--report-month", type=str, default=None, help="Optional report month in YYYY-MM (otherwise inferred from ת.דוח)")
    p.add_argument("--seed", type=int, default=None, help="Optional RNG seed for sampling")
    p.add_argument("--trustee-name", type=str, default=MIZRAHI_TRUSTEE_NAME_DEFAULT, help="Trustee name filter (default: Mizrahi)")
    p.add_argument("--manager-name", type=str, default=None, help="Fund manager name for report header (optional, column left empty if not provided)")
    p.add_argument("--spec-file", type=Path, default=None, help="Path to specification table Excel file (for פירוט בדיקות sheet)")
    p.add_argument("--cache-lists", type=Path, default=None, help="Path to cache/load problematic securities lists JSON")
    p.add_argument("--skip-tase-prices", action="store_true", help="Skip TASE price checks (spec #6 part 1)")
    p.add_argument(
        "--price-threshold",
        type=float,
        default=TASE_VARIANCE_THRESHOLD_DEFAULT,
        help=f"Price variance threshold in percent for TASE price checks (default: {TASE_VARIANCE_THRESHOLD_DEFAULT}%%)"
    )
    p.add_argument(
        "--max-exceptions",
        type=int,
        default=MAX_EXCEPTIONS_BEFORE_SAMPLING,
        help=f"Maximum exceptions before smart sampling is applied (default: {MAX_EXCEPTIONS_BEFORE_SAMPLING})"
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()

    # Set up logging with separate files for each specification
    log_dir = setup_logging()

    # Log startup info
    logger.info("=" * 60)
    logger.info("Mizrahi Special Transactions Processor - Starting")
    logger.info("=" * 60)
    logger.info("Log directory: %s", log_dir)
    logger.info("Price variance threshold: %.2f%%", args.price_threshold)
    logger.info("Manager name: %s", args.manager_name if args.manager_name else "(not specified - column will be empty)")

    for path in [args.mutual_funds_list, args.input_report]:
        if not path.exists():
            raise SystemExit(f"File not found: {path}")

    # Fund scope = Mizrahi trustee funds only (Maya report removed for now)
    in_scope_funds = load_mizrahi_fund_ids(args.mutual_funds_list, args.trustee_name)
    if not in_scope_funds:
        logger.warning("Filtered Mizrahi fund list is empty. Check trustee name / input file.")

    rows, meta = load_manager_report(args.input_report)

    # Determine report month: use provided value, or calculate previous month from today
    if args.report_month:
        report_month = args.report_month
    else:
        # Calculate previous month (current month - 1)
        today = datetime_module.now()
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year
        report_month = f"{prev_year:04d}-{prev_month:02d}"
        logger.info("Report month not provided, using previous month: %s", report_month)

    logger.info("Report month: %s", report_month)
    logger.info("Total rows loaded: %d", len(rows))

    # Check #1: all rows (inter-fund transactions - same abs quantity, opposite signs)
    ex_dup = check_1_abs_quantity_pairs(rows)

    # Check #2: filter to in-scope rows (fund exists in Mizrahi filtered list)
    in_scope_rows = [r for r in rows if r.fund_no in in_scope_funds]
    out_scope_rows = [r for r in rows if r.fund_no is not None and r.fund_no not in in_scope_funds]

    logger.info("In-scope rows: %d, Out-of-scope rows: %d", len(in_scope_rows), len(out_scope_rows))

    out_of_scope_funds: dict[int, dict[str, Any]] = {}
    if out_scope_rows:
        counts = Counter([r.fund_no for r in out_scope_rows if r.fund_no is not None])
        for fid, cnt in counts.items():
            out_of_scope_funds[int(fid)] = {
                "count_rows": int(cnt),
                "fund_name": next((r.fund_name for r in out_scope_rows if r.fund_no == fid and r.fund_name), None),
                "reason": "לא ברשימת קרנות מזרחי",
            }

    # Check #3, #4: in-scope only
    ex_date = check_3_dates_in_report_month(in_scope_rows, report_month)
    ex_decision = check_4_decision_method_rules(in_scope_rows)
    ex_price_type = check_4c_price_type_consistency(in_scope_rows)

    # Valid lines for sampling: in-scope rows NOT present in any exception list (including duplicates)
    ex_row_nums = {e.row.row_num for e in (ex_dup + ex_date + ex_decision + ex_price_type)}
    valid_rows = [r for r in in_scope_rows if r.row_num not in ex_row_nums]

    # Check #5: sampling
    samples = pick_samples(valid_rows, seed=args.seed)

    # Check #5.1: email JSON (two objects only)
    email_payload = build_email_json(samples)
    args.email_json.parent.mkdir(parents=True, exist_ok=True)
    args.email_json.write_text(json.dumps(email_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Check #6 Part 1: TASE price checks (sampled)
    price_check_results: list[PriceCheckResult] = []
    if not args.skip_tase_prices:
        logger.info("Running TASE price checks (spec #6 part 1) with threshold %.2f%%...", args.price_threshold)
        price_check_results = check_6_tase_prices(in_scope_rows, seed=args.seed, variance_threshold_pct=args.price_threshold)
        price_exceptions = [r for r in price_check_results if r.is_exception]
        logger.info("Sampled %d transactions, %d exceptions (>%.2f%% variance)", len(price_check_results), len(price_exceptions), args.price_threshold)

    # Check #6 Part 2: Price > 100 for types 31-36
    logger.info("Running price > 100 check (spec #6 part 2)...")
    price_limit_results = check_6_price_limits(in_scope_rows)
    logger.info("Found %d exceptions with price > 100", len(price_limit_results))

    # Check #7: Problematic securities
    logger.info("Running problematic securities check (spec #7)...")
    problematic_lists = fetch_problematic_lists(cache_path=args.cache_lists)
    for list_type, securities in problematic_lists.items():
        logger.info("  %s: %d ניירות", PROBLEMATIC_LISTS_CONFIG[list_type]['name_he'], len(securities))
    problematic_security_results = check_7_problematic_securities(in_scope_rows, problematic_lists)
    logger.info("Found %d transactions with problematic securities", len(problematic_security_results))

    # Apply smart sampling to exception lists if they exceed max_exceptions
    max_ex = args.max_exceptions
    ex_dup_original = len(ex_dup)
    ex_date_original = len(ex_date)
    ex_decision_original = len(ex_decision)
    ex_price_type_original = len(ex_price_type)

    ex_dup, _ = smart_sample_exceptions(ex_dup, max_ex, seed=args.seed)
    ex_date, _ = smart_sample_exceptions(ex_date, max_ex, seed=args.seed)
    ex_decision, _ = smart_sample_exceptions(ex_decision, max_ex, seed=args.seed)
    ex_price_type, _ = smart_sample_exceptions(ex_price_type, max_ex, seed=args.seed)

    if ex_dup_original > max_ex or ex_date_original > max_ex or ex_decision_original > max_ex or ex_price_type_original > max_ex:
        logger.info("Smart sampling applied: dup=%d->%d, date=%d->%d, decision=%d->%d, price_type=%d->%d",
                   ex_dup_original, len(ex_dup), ex_date_original, len(ex_date),
                   ex_decision_original, len(ex_decision), ex_price_type_original, len(ex_price_type))

    # Manager name: use empty string if not provided (column header will still appear)
    manager_name = args.manager_name if args.manager_name else ""

    # Count unique funds in input file
    unique_funds_in_input = len({r.fund_no for r in rows if r.fund_no is not None})
    # Count unique mizrahi funds in input file (intersection of input funds and in_scope_funds)
    unique_mizrahi_funds_in_input = len({r.fund_no for r in in_scope_rows if r.fund_no is not None})

    # Format exception counts - show "sampled/original" if sampling was applied
    def fmt_count(sampled: int, original: int) -> str:
        if sampled < original:
            return f"{sampled} (מתוך {original})"
        return str(sampled)

    summary = {
        "חודש דוח": report_month,
        "סיבת סינון": "קרנות מזרחי בלבד",
        "מספר קרנות של מנהל הקרן": unique_funds_in_input,
        "מספר קרנות של מנהל הקרן – בנאמנות מזרחי": unique_mizrahi_funds_in_input,
        "שורות בתחום": len(in_scope_rows),
        "קרנות מחוץ לתחום": len(out_of_scope_funds),
        "חריגות עסקאות בין קרנות": fmt_count(len(ex_dup), ex_dup_original),
        "חריגות תאריך": fmt_count(len(ex_date), ex_date_original),
        "חריגות אופן החלטה": fmt_count(len(ex_decision), ex_decision_original),
        "חריגות מחיר/סוג לא עקביים": fmt_count(len(ex_price_type), ex_price_type_original),
        "חריגות מחיר מעל 100": len(price_limit_results),
        "חריגות ניירות בעייתיים": len(problematic_security_results),
        "שורות תקינות לדגימה": len(valid_rows),
        "דגימה אופן החלטה 1 - שורה": samples.decision_1.row_num if samples.decision_1 else None,
        "דגימה אופן החלטה 2 - שורה": samples.decision_2.row_num if samples.decision_2 else None,
        "סף סטייה במחיר": f"{args.price_threshold}%",
        "סף דגימה חריגות": args.max_exceptions,
    }

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    write_output_xlsx(
        args.output_xlsx,
        report_month=report_month,
        manager_name=manager_name,
        trustee_name="מזרחי טפחות",
        summary=summary,
        exceptions_duplicates=ex_dup,
        exceptions_date=ex_date,
        exceptions_decision=ex_decision,
        exceptions_price_type=ex_price_type,
        samples=samples,
        in_scope_funds=in_scope_funds,
        price_check_results=price_check_results,
        price_limit_results=price_limit_results,
        problematic_security_results=problematic_security_results,
        spec_file_path=args.spec_file,
    )

    logger.info("=" * 60)
    logger.info("Processing complete")
    logger.info("=" * 60)
    logger.info("Summary: %s", json.dumps(summary, ensure_ascii=False, indent=2))
    logger.info("Output file: %s", args.output_xlsx)
    logger.info("Email JSON: %s", args.email_json)
    logger.info("Log directory: %s", log_dir)
    logger.info("")
    logger.info("Log files created:")
    logger.info("  - main.log              : General processing log")
    logger.info("  - chk1_inter_fund.log   : CHK_1 - Inter-fund transactions")
    logger.info("  - chk3_date.log         : CHK_3 - Date validation")
    logger.info("  - chk4_decision.log     : CHK_4 - Decision method rules")
    logger.info("  - chk6_tase_price.log   : CHK_6 - TASE price checks")
    logger.info("  - chk6_price_limit.log  : CHK_6 - Price > 100 checks")
    logger.info("  - chk6_failed_urls.log  : CHK_6 - Failed URL fetches (manual verification)")
    logger.info("  - chk7_problematic.log  : CHK_7 - Problematic securities")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
