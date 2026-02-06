#!/usr/bin/env python3
"""
Mizrahi_4 Automation - Daily Tracking Data Generator

Generates an Excel workbook with daily tracking data for Mizrahi-scope funds.
Creates one sheet per fund with dates from January 1 of the current year to today.

Inputs:
  - Mutual_Funds_List.xlsx - Worksheet.csv: Fund metadata (fees, trustee info)
  - Mizrahi_4/input/fund-index-table.xlsx: Fund-to-index mapping (86 funds)
  - Mizrahi_4/input/BFIX PRICE BLOOMBERG.xlsx: BFIX currency rates (optional)
  - Mizrahi_4/input/bloomberg-index.xlsx: Bloomberg index data (optional)

Outputs:
  - Mizrahi_4/output/output.xlsx: XLSX workbook with one sheet per fund

Usage:
    python Mizrahi_4/mizrahi_4_logic.py \
        --mutual-funds-list "Mutual_Funds_List.xlsx - Worksheet.csv" \
        --fund-index-table "Mizrahi_4/input/fund-index-table.xlsx" \
        --bfix-prices "Mizrahi_4/input/BFIX PRICE BLOOMBERG.xlsx" \
        --bloomberg-index "Mizrahi_4/input/bloomberg-index.xlsx" \
        --output-xlsx "Mizrahi_4/output/output.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import logging
import os
import sys
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Global log directory for current run
LOG_RUN_DIR: Optional[Path] = None

# Logger
logger = logging.getLogger(__name__)

# Global counter for unique table names
_table_counter = 0

# Excel styling constants
HEADER_FONT = Font(name='Arial', bold=True)
DATA_ALIGNMENT = Alignment(horizontal='right', vertical='center')

# Column definitions (1-indexed)
# (column_letter, header_text, number_format)
COLUMN_HEADERS = {
    1: ("A", "תאריך", 'd"/"m"/"yyyy'),
    2: ("B", "מסחר בקרן", None),
    3: ("C", "יום ללא חישוב", None),
    4: ("D", "מדד", "#,##0.00"),
    5: ("E", "שע\"ח", "#,##0.0000"),
    6: ("F", "שיעור שכר מנהל", "0.000%"),
    7: ("G", "שיעור שכר נאמן", "0.000%"),
    8: ("H", "מקדם שכר מצטבר", None),
    9: ("I", "שווי נכסי קרן לפני ניקויים", None),
    10: ("J", "יחידות בידי ציבור", None),
    11: ("K", "יחידות SWAP כשרה", None),
    12: ("L", "שער קרן לפני ניקויים", None),
    13: ("M", "שווי שכר מנהל יומי", None),
    14: ("N", "שווי שכר נאמן יומי", None),
    15: ("O", "שווי נכסי קרן בניקוי שכר מנהל ונאמן", None),
    16: ("P", "שער קרן בניקוי שכר מנהל ונאמן", None),
    17: ("Q", "מדד יחס", None),
    18: ("R", "הפרש עקיבה יומי", None),
    19: ("S", "שיעור דמי ניהול משתנים", "0.000%"),
    20: ("T", "דמי ניהול משתנים יומי", None),
    21: ("U", "דמי ניהול משתנים מצטבר", None),
    22: ("V", "שער קרן לאחר ניקויים", None),
    23: ("W", "עמלת פדיון ויצירה", None),
    24: ("X", "שער יצירה", None),
    25: ("Y", "שער פדיון", None),
    26: ("Z", "שווי דמי ניהול משתנים יומי", None),
    27: ("AA", "שווי נכסים לאחר ניקויים", None),
    28: ("AB", "שווי דמי ניהול משתנים מצטבר", None),
    29: ("AC", "שיעור ערבות בנקאית", None),
    30: ("AD", "שווי ערבות בנקאית", None),
    31: ("AE", "רצועת ביטחון", None),
}


# -----------------------------
# Data structures
# -----------------------------

@dataclass
class MutualFund:
    """A fund from the mutual funds list."""
    fund_id: int
    fund_name: str
    trustee_name: str
    manager_fee: float  # שכר המנהל (as percentage, e.g., 0.65 for 0.65%)
    trustee_fee: float  # שכר הנאמן (as percentage)
    variable_fee: float  # דמי ניהול משתנים (as percentage)


@dataclass
class FundData:
    """A fund with its index mapping for output."""
    fund_id: int
    fund_name: str
    index_id: str  # Index identifier from column B (name for אינדקס, code for others)
    manager_fee: float  # Already divided by 100 (e.g., 0.0065 for 0.65%)
    trustee_fee: float  # Already divided by 100
    variable_fee: float  # Already divided by 100
    currency_code: Optional[str] = None  # סוג מטבע (קוד bfix) - None means use 1
    bfix_codes: list[str] = None  # List of BFIX codes for this fund
    data_source: Optional[str] = None  # מקור נתונים - "בלומברג", "מאי"ה", "אינדקס", etc.
    index_name: Optional[str] = None  # Index name from column C (לינק) - used for מאי"ה header
    indx_code: Optional[str] = None  # INDX code extracted from URL for אינדקס funds


@dataclass
class BfixColumn:
    """A column from BFIX PRICE BLOOMBERG.xlsx."""
    code: str  # e.g., "ILS F103 Curncy"
    description: str  # e.g., "שער 17:30 - כל הימים BFIX USD"
    header: str  # Combined: "description (code)"
    col_idx: int  # Column index in the source file


@dataclass
class BfixData:
    """BFIX price data loaded from Bloomberg file."""
    columns: list[BfixColumn]  # All available columns
    # Map: date -> {code -> {col_idx -> value}}
    data: dict[dt.date, dict[str, dict[int, float]]]


@dataclass
class BloombergIndexColumn:
    """A column from bloomberg-index.xlsx."""
    code: str  # First part of the code, e.g., "SPTR500N"
    full_code: str  # Full code with suffix, e.g., "SPTR500N EQUITY"
    header: str  # Combined: "מדד (full_code)"
    col_idx: int  # Column index in the source file


@dataclass
class BloombergIndexData:
    """Bloomberg index data loaded from Excel file."""
    columns: list[BloombergIndexColumn]  # All available columns
    # Map: date -> {code (lowercase) -> value}
    data: dict[dt.date, dict[str, float]]


@dataclass
class IndxIndexData:
    """INDX index data loaded from individual Excel files."""
    # Map: index_id (str) -> {date -> value}
    data: dict[str, dict[dt.date, float]] = field(default_factory=dict)


@dataclass
class TaseIndexData:
    """TASE (Maya) index data fetched from TASE Data Hub API."""
    # Map: index_id (str) -> {date -> closingIndexPrice}
    data: dict[str, dict[dt.date, float]] = field(default_factory=dict)


# -----------------------------
# Helpers
# -----------------------------

def _to_str(v) -> Optional[str]:
    """Convert value to string, handling None and pandas NaN."""
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    # Also check for string "nan" from pandas
    if s.lower() == "nan" or not s:
        return None
    return s


def _to_int(v) -> Optional[int]:
    """Convert value to int, handling None and non-numeric values."""
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _to_float(v) -> Optional[float]:
    """Convert value to float, handling None and non-numeric values."""
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except Exception:
        return None


# -----------------------------
# Logging setup
# -----------------------------

def setup_logging(log_base_dir: Path = Path(__file__).parent.parent / "log") -> Path:
    """Set up logging with timestamped directory."""
    global LOG_RUN_DIR

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    short_uuid = str(uuid.uuid4())[:8]
    run_id = f"{timestamp}_{short_uuid}"

    run_dir = log_base_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    LOG_RUN_DIR = run_dir

    log_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(log_format)

    # File handler
    file_handler = logging.FileHandler(run_dir / "mizrahi_4.log", encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(log_format)

    logger.setLevel(logging.DEBUG)
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    logger.info("Log directory created: %s", run_dir)
    return run_dir


# -----------------------------
# Data loaders
# -----------------------------

def load_mutual_funds(csv_path: Path) -> dict[int, MutualFund]:
    """
    Load mutual funds list from CSV.

    Returns dict keyed by fund ID (מספר בורסה).
    """
    funds: dict[int, MutualFund] = {}

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [h.strip() for h in reader.fieldnames]

        for row in reader:
            fund_id = _to_int(row.get("מספר בורסה"))
            if fund_id is None:
                continue

            fund = MutualFund(
                fund_id=fund_id,
                fund_name=_to_str(row.get("שם קרן בעברית")) or "",
                trustee_name=_to_str(row.get("שם נאמן")) or "",
                manager_fee=_to_float(row.get("שכר המנהל")) or 0.0,
                trustee_fee=_to_float(row.get("שכר הנאמן")) or 0.0,
                variable_fee=_to_float(row.get("דמי ניהול משתנים")) or 0.0,
            )
            funds[fund_id] = fund

    logger.info("Loaded %d funds from mutual funds list", len(funds))
    return funds


def _extract_indx_code_from_url(url: Optional[str]) -> Optional[str]:
    """
    Extract INDX code from URL like 'https://indx.co.il/index/2123-index/'.
    Returns '2123' from the example above.
    """
    import re
    if not url:
        return None
    # Match pattern like /index/2123-index/ or /index/2123/
    match = re.search(r'/index/(\d+)', url)
    return match.group(1) if match else None


def load_fund_index_table(xlsx_path: Path) -> dict[int, tuple[str, Optional[str], list[str], Optional[str], Optional[str], Optional[str]]]:
    """
    Load fund-to-index mapping from Excel file.

    Returns dict mapping fund_id -> (index_id, currency_code, bfix_codes, data_source, index_name, indx_code).
    currency_code is None if not specified (use default 1).
    bfix_codes is a list of BFIX codes from column F (comma-separated).
    data_source is the מקור נתונים value ("בלומברג", "מאי"ה", "אינדקס", etc.).
    index_name is the לינק value (index name for מאי"ה, URL for אינדקס).
    indx_code is the INDX code extracted from the URL for אינדקס funds.
    """
    import pandas as pd

    df = pd.read_excel(xlsx_path)

    # Column names from the file
    fund_col = "מס' קרן"
    index_col = "מספר מדד"
    link_col = "לינק"
    currency_col = "סוג מטבע (קוד bfix)"
    data_source_col = "מקור נתונים"

    fund_index_map: dict[int, tuple[str, Optional[str], list[str], Optional[str], Optional[str], Optional[str]]] = {}

    for _, row in df.iterrows():
        fund_id = _to_int(row.get(fund_col))
        index_id = _to_str(row.get(index_col))
        index_name = _to_str(row.get(link_col))
        currency_code_raw = _to_str(row.get(currency_col))
        data_source = _to_str(row.get(data_source_col))

        # Extract INDX code from URL for אינדקס funds
        indx_code: Optional[str] = None
        if data_source == "אינדקס" and index_name:
            indx_code = _extract_indx_code_from_url(index_name)

        # Parse BFIX codes (comma-separated)
        bfix_codes: list[str] = []
        if currency_code_raw:
            bfix_codes = [code.strip() for code in currency_code_raw.split(',') if code.strip()]

        if fund_id is not None and index_id is not None:
            fund_index_map[fund_id] = (index_id, currency_code_raw, bfix_codes, data_source, index_name, indx_code)

    logger.info("Loaded %d fund-index mappings", len(fund_index_map))
    return fund_index_map


def load_bfix_data(xlsx_path: Path) -> BfixData:
    """
    Load BFIX price data from Bloomberg Excel file.

    File structure (sheet "ערכים"):
    - Row 0: Descriptions (e.g., "שער 17:30 - כל הימים BFIX USD")
    - Row 1: BFIX codes (e.g., "ILS F103 Curncy")
    - Row 2: Field types (WEEKDAY, date, LAST_PRICE, etc.) - skipped
    - Row 3+: Data rows
    - Column 2 (C): Dates

    Returns BfixData with all columns and date-indexed data.
    """
    import pandas as pd

    df_raw = pd.read_excel(xlsx_path, sheet_name='ערכים', header=None)

    # Parse columns (starting from column 3, which is index 3)
    columns: list[BfixColumn] = []
    code_to_columns: dict[str, list[BfixColumn]] = {}

    for col_idx in range(3, df_raw.shape[1]):
        code = df_raw.iloc[1, col_idx]
        desc = df_raw.iloc[0, col_idx]

        if pd.isna(code) or not str(code).strip():
            continue

        code = str(code).strip()
        desc = str(desc).strip() if pd.notna(desc) and str(desc).strip() else code

        # Create header: "description (code)"
        header = f"{desc} ({code})"

        bfix_col = BfixColumn(
            code=code,
            description=desc,
            header=header,
            col_idx=col_idx
        )
        columns.append(bfix_col)

        if code not in code_to_columns:
            code_to_columns[code] = []
        code_to_columns[code].append(bfix_col)

    # Parse data rows (starting from row 3, which is index 3)
    data: dict[dt.date, dict[str, dict[int, float]]] = {}

    for row_idx in range(3, df_raw.shape[0]):
        date_val = df_raw.iloc[row_idx, 2]  # Column C (index 2) has dates

        if pd.isna(date_val):
            continue

        # Convert to date
        if isinstance(date_val, dt.datetime):
            date_key = date_val.date()
        elif isinstance(date_val, dt.date):
            date_key = date_val
        else:
            try:
                date_key = pd.to_datetime(date_val).date()
            except Exception:
                continue

        data[date_key] = {}

        for bfix_col in columns:
            value = df_raw.iloc[row_idx, bfix_col.col_idx]

            if bfix_col.code not in data[date_key]:
                data[date_key][bfix_col.code] = {}

            # Store value (could be NaN)
            if pd.notna(value):
                try:
                    data[date_key][bfix_col.code][bfix_col.col_idx] = float(value)
                except (ValueError, TypeError):
                    # Non-numeric value, skip
                    pass

    logger.info("Loaded BFIX data: %d columns, %d dates", len(columns), len(data))
    return BfixData(columns=columns, data=data)


def load_bloomberg_index_data(xlsx_path: Path) -> BloombergIndexData:
    """
    Load Bloomberg index data from Excel file.

    File structure (sheet "ערכים"):
    - Row 0-1: Headers/labels (skipped)
    - Row 2: Index codes with suffixes (e.g., "SPTR500N EQUITY", "M1WO INDEX")
    - Row 3: Field types (date, LAST_PRICE, etc.) - skipped
    - Row 4+: Data rows
    - Column 1 (B): Dates

    Returns BloombergIndexData with columns and date-indexed data.
    """
    import pandas as pd

    df_raw = pd.read_excel(xlsx_path, sheet_name='ערכים', header=None)

    # Parse columns (starting from column 2, which is index 2)
    columns: list[BloombergIndexColumn] = []

    for col_idx in range(2, df_raw.shape[1]):
        full_code = df_raw.iloc[2, col_idx]  # Row 2 has the full index codes

        if pd.isna(full_code) or not str(full_code).strip():
            continue

        full_code = str(full_code).strip()

        # Skip if it looks like a date (from column 1 bleeding over)
        if 'datetime' in str(type(full_code)).lower() or full_code.startswith('20'):
            continue

        # Extract the first part (before space) as the code for matching
        code = full_code.split()[0] if ' ' in full_code else full_code

        # Create header: "מדד (full_code)"
        header = f"מדד ({full_code})"

        bloomberg_col = BloombergIndexColumn(
            code=code,
            full_code=full_code,
            header=header,
            col_idx=col_idx
        )
        columns.append(bloomberg_col)

    # Parse data rows (starting from row 4, which is index 4)
    data: dict[dt.date, dict[str, float]] = {}

    for row_idx in range(4, df_raw.shape[0]):
        date_val = df_raw.iloc[row_idx, 1]  # Column B (index 1) has dates

        if pd.isna(date_val):
            continue

        # Convert to date
        if isinstance(date_val, dt.datetime):
            date_key = date_val.date()
        elif isinstance(date_val, dt.date):
            date_key = date_val
        else:
            try:
                date_key = pd.to_datetime(date_val).date()
            except Exception:
                continue

        data[date_key] = {}

        for bloomberg_col in columns:
            value = df_raw.iloc[row_idx, bloomberg_col.col_idx]

            # Handle various error/missing values
            if pd.isna(value):
                continue

            # Check for Excel errors like #NAME?, #N/A, etc.
            if isinstance(value, str):
                if value.startswith('#'):
                    continue
                try:
                    value = float(value)
                except ValueError:
                    continue

            try:
                # Store using lowercase code for case-insensitive matching
                data[date_key][bloomberg_col.code.lower()] = float(value)
            except (ValueError, TypeError):
                pass

    logger.info("Loaded Bloomberg index data: %d columns, %d dates", len(columns), len(data))
    return BloombergIndexData(columns=columns, data=data)


def load_indx_index_data(indx_records_dir: Path) -> IndxIndexData:
    """
    Load INDX historical data from Excel files.

    Files expected: {index_id}_Historical_Data.xlsx
    Sheet: index_levels
    Columns: Date, EoD Price
    """
    import pandas as pd

    result = IndxIndexData()

    for file in indx_records_dir.glob("*_Historical_Data.xlsx"):
        # Extract index_id from filename: "2123_Historical_Data.xlsx" → "2123"
        index_id = file.stem.replace("_Historical_Data", "")

        try:
            df = pd.read_excel(file, sheet_name="index_levels")
            date_values = {}

            for _, row in df.iterrows():
                date_val = row["Date"]
                price_val = row["EoD Price"]

                # Convert date
                if isinstance(date_val, dt.datetime):
                    date_key = date_val.date()
                elif isinstance(date_val, dt.date):
                    date_key = date_val
                else:
                    continue

                # Convert price
                if pd.notna(price_val):
                    date_values[date_key] = float(price_val)

            result.data[index_id] = date_values
            logger.info("Loaded INDX data for index %s: %d dates", index_id, len(date_values))

        except Exception as e:
            logger.warning("Failed to load INDX file %s: %s", file, e)

    return result


def fetch_tase_index_data(
    index_ids: list[str],
    api_key: str,
    from_date: dt.date,
    to_date: dt.date
) -> TaseIndexData:
    """
    Fetch index data from TASE Data Hub API for specified indices.

    Uses endpoint: /v1/indices/eod/history/five-years/by-index
    Returns closingIndexPrice for each date.
    """
    import requests
    import time

    result = TaseIndexData()
    base_url = "https://datawise.tase.co.il/v1/indices/eod/history/five-years/by-index"

    headers = {
        "accept": "application/json",
        "accept-language": "he-IL",
        "apikey": api_key
    }

    for index_id in index_ids:
        try:
            # Convert index_id to integer for API
            index_int = int(index_id)

            params = {
                "indexId": index_int,
                "fromDate": from_date.strftime("%Y-%m-%d"),
                "toDate": to_date.strftime("%Y-%m-%d")
            }

            logger.debug("Fetching TASE data for index %s (from %s to %s)",
                        index_id, from_date, to_date)

            response = requests.get(base_url, headers=headers, params=params, timeout=30)

            if response.status_code == 429:
                # Rate limited - wait and retry
                logger.warning("Rate limited by TASE API, waiting 3 seconds...")
                time.sleep(3)
                response = requests.get(base_url, headers=headers, params=params, timeout=30)

            if response.status_code != 200:
                logger.warning("TASE API error for index %s: HTTP %d - %s",
                              index_id, response.status_code, response.text[:200])
                continue

            data = response.json()
            eod_data = data.get("indexEndOfDay", {})
            results = eod_data.get("result", [])

            date_values: dict[dt.date, float] = {}
            for item in results:
                trade_date_str = item.get("tradeDate")
                closing_price = item.get("closingIndexPrice")

                if trade_date_str and closing_price is not None:
                    # Parse date from format "2023-03-01T00:00:00"
                    try:
                        trade_date = dt.datetime.strptime(
                            trade_date_str.split("T")[0], "%Y-%m-%d"
                        ).date()
                        date_values[trade_date] = float(closing_price)
                    except (ValueError, TypeError) as e:
                        logger.debug("Failed to parse TASE date/price: %s", e)
                        continue

            result.data[index_id] = date_values
            logger.info("Fetched TASE data for index %s: %d dates", index_id, len(date_values))

            # Respect rate limit: 10 requests per 2 seconds
            time.sleep(0.25)

        except Exception as e:
            logger.warning("Failed to fetch TASE data for index %s: %s", index_id, e)
            continue

    return result


def is_fund_in_scope(fund: MutualFund) -> bool:
    """
    Check if a fund is in scope based on:
    1. Mizrahi trustee
    2. Has non-zero variable management fees (דמי ניהול משתנים)
    """
    # Check Mizrahi trustee
    if "מזרחי טפחות" not in fund.trustee_name:
        return False

    # Check variable fees - must be non-zero and non-blank
    if fund.variable_fee is None or fund.variable_fee == 0:
        return False

    return True


def get_in_scope_funds(
    mutual_funds: dict[int, MutualFund],
    fund_index_map: dict[int, tuple[str, Optional[str], list[str], Optional[str], Optional[str], Optional[str]]]
) -> list[FundData]:
    """
    Filter to in-scope funds: Mizrahi trustee + has index mapping + has variable fees.

    Returns list of FundData objects with fees divided by 100.
    """
    in_scope: list[FundData] = []

    for fund_id, (index_id, currency_code, bfix_codes, data_source, index_name, indx_code) in fund_index_map.items():
        fund = mutual_funds.get(fund_id)

        if fund is None:
            logger.warning("Fund %d in index table but not in mutual funds list", fund_id)
            continue

        # Check if fund is in scope (Mizrahi + has variable fees)
        if not is_fund_in_scope(fund):
            logger.debug("Fund %d (%s) excluded: not in scope (trustee=%s, variable_fee=%s)",
                        fund_id, fund.fund_name, fund.trustee_name, fund.variable_fee)
            continue

        # Create FundData with fees divided by 100
        fund_data = FundData(
            fund_id=fund_id,
            fund_name=fund.fund_name,
            index_id=index_id,
            manager_fee=fund.manager_fee / 100.0,
            trustee_fee=fund.trustee_fee / 100.0,
            variable_fee=fund.variable_fee / 100.0,
            currency_code=currency_code,
            bfix_codes=bfix_codes if bfix_codes else [],
            data_source=data_source,
            index_name=index_name,
            indx_code=indx_code,
        )
        in_scope.append(fund_data)

    logger.info("Found %d in-scope funds (Mizrahi + index mapping + variable fees)", len(in_scope))
    return in_scope


def generate_date_range() -> list[dt.date]:
    """
    Generate list of dates from January 1 of current year to today.
    """
    today = dt.date.today()
    start_date = dt.date(today.year, 1, 1)

    dates: list[dt.date] = []
    current = start_date
    while current <= today:
        dates.append(current)
        current += dt.timedelta(days=1)

    logger.info("Generated date range: %s to %s (%d days)",
               start_date, today, len(dates))
    return dates


# -----------------------------
# Excel generation
# -----------------------------

def get_bfix_columns_for_fund(fund: FundData, bfix_data: BfixData) -> list[BfixColumn]:
    """
    Get all BFIX columns that match the fund's BFIX codes.

    Returns columns in the order they appear in the BFIX file,
    including all columns for each code (not just the first).
    """
    if not fund.bfix_codes or not bfix_data:
        return []

    # Normalize fund codes for case-insensitive matching
    fund_codes_normalized = {code.upper(): code for code in fund.bfix_codes}

    matching_columns: list[BfixColumn] = []
    for bfix_col in bfix_data.columns:
        if bfix_col.code.upper() in fund_codes_normalized:
            matching_columns.append(bfix_col)

    return matching_columns


def get_bloomberg_column_for_fund(
    fund: FundData,
    bloomberg_data: BloombergIndexData
) -> Optional[BloombergIndexColumn]:
    """
    Find the Bloomberg index column matching the fund's index_id.

    Uses case-insensitive matching on the first part of the index code.
    """
    if not bloomberg_data or not fund.index_id:
        return None

    fund_index_normalized = fund.index_id.upper()

    for bloomberg_col in bloomberg_data.columns:
        if bloomberg_col.code.upper() == fund_index_normalized:
            return bloomberg_col

    return None


def create_fund_sheet(
    ws,
    fund: FundData,
    dates: list[dt.date],
    bfix_data: Optional[BfixData] = None,
    bloomberg_data: Optional[BloombergIndexData] = None,
    indx_data: Optional[IndxIndexData] = None,
    tase_data: Optional[TaseIndexData] = None
) -> None:
    """
    Populate a worksheet for a single fund.

    Structure:
    - Row 1: Title with fund name and code (merged across columns)
    - Row 2: Headers (תאריך, מסחר בקרן, מדד, etc.) + BFIX columns
    - Row 3+: Data rows (one per date)

    For funds with data_source == "בלומברג", column D is populated with
    Bloomberg index values instead of just the index ID.
    For funds with data_source == "אינדקס", column D is populated with
    INDX index values.
    For funds with data_source == "מאי"ה", column D is populated with
    TASE index values from the API.
    """
    global _table_counter

    # Get BFIX columns for this fund
    bfix_columns = get_bfix_columns_for_fund(fund, bfix_data) if bfix_data else []
    total_columns = 31 + len(bfix_columns)  # Base 31 columns + BFIX columns

    # Check if fund uses Bloomberg data source
    use_bloomberg = fund.data_source == "בלומברג" and bloomberg_data is not None
    bloomberg_col = None
    if use_bloomberg:
        bloomberg_col = get_bloomberg_column_for_fund(fund, bloomberg_data)
        if bloomberg_col is None:
            logger.warning(
                "Fund %d (%s) has Bloomberg data source but no matching index column found for '%s'",
                fund.fund_id, fund.fund_name, fund.index_id
            )
            use_bloomberg = False

    # Check if fund uses INDX data source
    use_indx = fund.data_source == "אינדקס" and indx_data is not None and fund.indx_code is not None
    indx_values = None
    if use_indx:
        if fund.indx_code in indx_data.data:
            indx_values = indx_data.data[fund.indx_code]
        else:
            logger.warning(
                "Fund %d (%s) has INDX data source but no matching data for index code '%s'",
                fund.fund_id, fund.fund_name, fund.indx_code
            )
            use_indx = False

    # Check if fund uses TASE (מאי"ה) data source
    use_tase = fund.data_source == 'מאי"ה' and tase_data is not None
    tase_values = None
    if use_tase:
        if fund.index_id in tase_data.data:
            tase_values = tase_data.data[fund.index_id]
        else:
            logger.warning(
                "Fund %d (%s) has TASE data source but no matching data for index '%s'",
                fund.fund_id, fund.fund_name, fund.index_id
            )
            use_tase = False

    # Set RTL
    ws.sheet_view.rightToLeft = True

    # Row 1: Title with fund name and code (merged across all columns)
    title_text = f"{fund.fund_name} ({fund.fund_id})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(1, 1, title_text)
    title_cell.font = Font(name='Arial', bold=True, size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='right', vertical='center')
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Row 2: Headers
    current_year = dt.date.today().year
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, (_, header_text, _) in COLUMN_HEADERS.items():
        # Special case for date column - show current year
        if col_idx == 1:
            header_text = f"תאריך ({current_year})"
        # Special case for מדד column when using Bloomberg
        elif col_idx == 4 and use_bloomberg and bloomberg_col:
            header_text = bloomberg_col.header  # "מדד (FULL_INDEX_CODE)"
        # Special case for מדד column when using INDX
        elif col_idx == 4 and use_indx and indx_values is not None:
            header_text = f"מדד ({fund.index_id})"
        # Special case for מדד column when using TASE (מאי"ה)
        elif col_idx == 4 and use_tase and tase_values is not None:
            # Use index_name from column C (e.g., "תל בונד-שקלי 1-3")
            index_display = fund.index_name if fund.index_name else fund.index_id
            header_text = f"מדד ({index_display})"
        cell = ws.cell(2, col_idx, header_text if header_text else None)
        cell.font = HEADER_FONT
        cell.alignment = header_alignment

    # Add BFIX column headers (starting at column 32)
    for i, bfix_col in enumerate(bfix_columns):
        col_idx = 32 + i
        cell = ws.cell(2, col_idx, bfix_col.header)
        cell.font = HEADER_FONT
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, date in enumerate(dates, start=3):
        # Column A: תאריך
        cell = ws.cell(row_idx, 1, date)
        cell.number_format = 'd"/"m"/"yyyy'
        cell.alignment = DATA_ALIGNMENT

        # Column D: מדד - either Bloomberg data or index_id placeholder
        cell = ws.cell(row_idx, 4)
        cell.alignment = DATA_ALIGNMENT
        if use_bloomberg and bloomberg_col:
            # Get Bloomberg index value for this date
            date_data = bloomberg_data.data.get(date, {})
            value = date_data.get(bloomberg_col.code.lower())

            if value is not None:
                cell.value = value
                cell.number_format = "#,##0.00"
            else:
                # Missing value - use Excel #N/A error with red font
                cell.value = '#N/A'
                cell.data_type = 'e'  # Excel error type
                cell.font = Font(color="FF0000")
        elif use_indx and indx_values is not None:
            # Get INDX index value for this date
            if date in indx_values:
                cell.value = indx_values[date]
                cell.number_format = "#,##0.00"
            else:
                # Missing value - use Excel #N/A error with red font
                cell.value = '#N/A'
                cell.data_type = 'e'  # Excel error type
                cell.font = Font(color="FF0000")
        elif use_tase and tase_values is not None:
            # Get TASE (מאי"ה) index value for this date
            if date in tase_values:
                cell.value = tase_values[date]
                cell.number_format = "#,##0.00"
            else:
                # Missing value - use Excel #N/A error with red font
                cell.value = '#N/A'
                cell.data_type = 'e'  # Excel error type
                cell.font = Font(color="FF0000")
        else:
            # Default: just show index_id as text placeholder
            cell.value = fund.index_id
            cell.number_format = "@"  # Text format

        # Column E: שע"ח - always 1.0000
        cell = ws.cell(row_idx, 5, 1.0)
        cell.number_format = "#,##0.0000"
        cell.alignment = DATA_ALIGNMENT

        # Column F: שיעור שכר מנהל
        cell = ws.cell(row_idx, 6, fund.manager_fee)
        cell.number_format = "0.000%"
        cell.alignment = DATA_ALIGNMENT

        # Column G: שיעור שכר נאמן
        cell = ws.cell(row_idx, 7, fund.trustee_fee)
        cell.number_format = "0.000%"
        cell.alignment = DATA_ALIGNMENT

        # Column S: שיעור דמי ניהול משתנים
        cell = ws.cell(row_idx, 19, fund.variable_fee)
        cell.number_format = "0.000%"
        cell.alignment = DATA_ALIGNMENT

        # Add BFIX data columns
        if bfix_data and bfix_columns:
            # Check if date exists in BFIX data at all
            date_exists_in_bfix = date in bfix_data.data
            date_data = bfix_data.data.get(date, {})

            for i, bfix_col in enumerate(bfix_columns):
                col_idx = 32 + i
                cell = ws.cell(row_idx, col_idx)
                cell.alignment = DATA_ALIGNMENT

                if not date_exists_in_bfix:
                    # Date doesn't exist in BFIX data at all - use Excel #N/A error with red font
                    cell.value = '#N/A'
                    cell.data_type = 'e'  # Excel error type
                    cell.font = Font(color="FF0000")  # Red to indicate missing date
                else:
                    # Date exists - check for value
                    code_data = date_data.get(bfix_col.code, {})
                    value = code_data.get(bfix_col.col_idx)

                    if value is not None:
                        cell.value = value
                        cell.number_format = "#,##0.0000"
                    else:
                        # Value is NaN in source - use Excel #N/A error
                        cell.value = '#N/A'
                        cell.data_type = 'e'  # Excel error type

    # Set column widths
    column_widths = {
        1: 15,   # A - תאריך
        2: 12,   # B - מסחר בקרן
        3: 14,   # C - יום ללא חישוב
        4: 12,   # D - מדד
        5: 10,   # E - שע"ח
        6: 16,   # F - שכר מנהל
        7: 16,   # G - שכר נאמן
        8: 16,   # H - מקדם שכר
        9: 22,   # I - שווי נכסי קרן
        10: 16,  # J - יחידות ציבור
        11: 16,  # K - יחידות SWAP
        12: 18,  # L - שער קרן לפני
        16: 26,  # M - שווי שכר מנהל
        14: 16,  # N - שווי שכר נאמן
        15: 28,  # O - שווי נכסי קרן בניקוי
        16: 26,  # P - שער קרן בניקוי
        17: 12,  # Q - מדד יחס
        18: 16,  # R - הפרש עקיבה
        19: 22,  # S - דמי ניהול משתנים
        20: 20,  # T - דמי ניהול יומי
        21: 22,  # U - דמי ניהול מצטבר
        22: 20,  # V - שער קרן לאחר
        23: 16,  # W - עמלת פדיון
        24: 12,  # X - שער יצירה
        25: 12,  # Y - שער פדיון
        26: 22,  # Z - שווי דמי ניהול יומי
        27: 22,  # AA - שווי נכסים לאחר
        28: 24,  # AB - שווי דמי ניהול מצטבר
        29: 18,  # AC - שיעור ערבות
        30: 18,  # AD - שווי ערבות
        31: 16,  # AE - רצועת ביטחון
    }

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set BFIX column widths
    for i in range(len(bfix_columns)):
        col_idx = 32 + i
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    # Apply header styling (row 2) - blue background, white bold text
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_styled = Font(name='Arial', bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )

    for col_idx in range(1, total_columns + 1):
        cell = ws.cell(2, col_idx)
        cell.fill = header_fill
        cell.font = header_font_styled
        cell.border = thin_border

    # Apply alternating row stripes to data rows
    stripe_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    last_row = 2 + len(dates)

    for row_idx in range(3, last_row + 1):
        row_offset = row_idx - 3  # 0-indexed from first data row
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border
            # Apply stripe to odd rows (0-indexed)
            if row_offset % 2 == 1:
                # Only set fill if cell doesn't have special formatting (like red font for MISSING)
                if cell.font.color is None or cell.font.color.rgb != "00FF0000":
                    cell.fill = stripe_fill
                else:
                    # Keep red font but add stripe background
                    cell.fill = stripe_fill

    # Enable auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(total_columns)}{last_row}"


def create_missing_funds_sheet(ws, missing_funds: list[MutualFund]) -> None:
    """
    Create a sheet listing Mizrahi funds that are missing from the index table.
    """
    global _table_counter

    # Set RTL
    ws.sheet_view.rightToLeft = True

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(1, 1, "קרנות חסרות במקרא")
    title_cell.font = Font(name='Arial', bold=True, size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='right', vertical='center')
    title_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    # Row 2: Headers
    headers = ["מספר בורסה", "שם קרן בעברית", "שם נאמן", "מצב הקרן"]
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(2, col_idx, header)
        cell.font = HEADER_FONT
        cell.alignment = header_alignment

    # Data rows
    for row_idx, fund in enumerate(sorted(missing_funds, key=lambda f: f.fund_id), start=3):
        ws.cell(row_idx, 1, fund.fund_id).alignment = DATA_ALIGNMENT
        ws.cell(row_idx, 2, fund.fund_name).alignment = DATA_ALIGNMENT
        ws.cell(row_idx, 3, fund.trustee_name).alignment = DATA_ALIGNMENT
        # We don't have status in MutualFund, so leave empty or add it

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12

    # Add table
    if len(missing_funds) > 0:
        _table_counter += 1
        table_name = f"MissingFunds_{_table_counter}"
        last_row = 2 + len(missing_funds)
        table_range = f"A2:D{last_row}"

        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium3",  # Red-ish style
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)


def generate_output_workbook(
    funds: list[FundData],
    dates: list[dt.date],
    output_path: Path,
    missing_funds: list[MutualFund] = None,
    bfix_data: Optional[BfixData] = None,
    bloomberg_data: Optional[BloombergIndexData] = None,
    indx_data: Optional[IndxIndexData] = None,
    tase_data: Optional[TaseIndexData] = None
) -> None:
    """
    Generate the output Excel workbook with one sheet per fund.
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Sort funds by fund_id for consistent ordering
    sorted_funds = sorted(funds, key=lambda f: f.fund_id)

    for fund in sorted_funds:
        # Sheet name is the fund's מספר בורסה
        sheet_name = str(fund.fund_id)

        # Excel sheet names max 31 chars
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)
        create_fund_sheet(ws, fund, dates, bfix_data, bloomberg_data, indx_data, tase_data)

        logger.debug("Created sheet for fund %d (%s)", fund.fund_id, fund.fund_name)

    # Add missing funds sheet at the end
    if missing_funds:
        ws_missing = wb.create_sheet(title="קרנות חסרות במקרא")
        create_missing_funds_sheet(ws_missing, missing_funds)
        logger.info("Added missing funds sheet with %d funds", len(missing_funds))

    # Save workbook
    wb.save(output_path)
    logger.info("Saved workbook to: %s", output_path)
    logger.info("Total sheets: %d", len(wb.sheetnames))


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate daily tracking Excel workbook for Mizrahi-scope funds"
    )
    parser.add_argument(
        "--mutual-funds-list",
        required=True,
        help="Path to Mutual_Funds_List CSV file"
    )
    parser.add_argument(
        "--fund-index-table",
        required=True,
        help="Path to fund-index-table.xlsx file"
    )
    parser.add_argument(
        "--bfix-prices",
        required=False,
        help="Path to BFIX PRICE BLOOMBERG.xlsx file (optional)"
    )
    parser.add_argument(
        "--bloomberg-index",
        required=False,
        help="Path to bloomberg-index.xlsx file (optional)"
    )
    parser.add_argument(
        "--indx-records",
        required=False,
        help="Directory containing INDX historical Excel files (*_Historical_Data.xlsx)"
    )
    parser.add_argument(
        "--skip-tase-data",
        action="store_true",
        help="Skip fetching TASE (מאי\"ה) index data from TASE Data Hub API"
    )
    parser.add_argument(
        "--output-xlsx",
        required=True,
        help="Path for output Excel file"
    )

    args = parser.parse_args()

    # Setup logging
    setup_logging()

    logger.info("=" * 60)
    logger.info("Mizrahi_4 Daily Tracking Data Generator")
    logger.info("=" * 60)

    # Load data
    mutual_funds_path = Path(args.mutual_funds_list)
    fund_index_path = Path(args.fund_index_table)
    output_path = Path(args.output_xlsx)

    logger.info("Loading mutual funds from: %s", mutual_funds_path)
    mutual_funds = load_mutual_funds(mutual_funds_path)

    logger.info("Loading fund-index table from: %s", fund_index_path)
    fund_index_map = load_fund_index_table(fund_index_path)

    # Load BFIX data if provided
    bfix_data: Optional[BfixData] = None
    if args.bfix_prices:
        bfix_path = Path(args.bfix_prices)
        logger.info("Loading BFIX prices from: %s", bfix_path)
        try:
            bfix_data = load_bfix_data(bfix_path)
        except Exception as e:
            logger.warning("Failed to load BFIX data: %s", e)
            logger.warning("Continuing without BFIX data")

    # Load Bloomberg index data if provided
    bloomberg_data: Optional[BloombergIndexData] = None
    if args.bloomberg_index:
        bloomberg_path = Path(args.bloomberg_index)
        logger.info("Loading Bloomberg index data from: %s", bloomberg_path)
        try:
            bloomberg_data = load_bloomberg_index_data(bloomberg_path)
        except Exception as e:
            logger.warning("Failed to load Bloomberg index data: %s", e)
            logger.warning("Continuing without Bloomberg index data")

    # Load INDX index data if provided
    indx_data: Optional[IndxIndexData] = None
    if args.indx_records:
        indx_records_path = Path(args.indx_records)
        logger.info("Loading INDX index data from: %s", indx_records_path)
        try:
            indx_data = load_indx_index_data(indx_records_path)
            logger.info("Loaded INDX data for %d indices", len(indx_data.data))
        except Exception as e:
            logger.warning("Failed to load INDX index data: %s", e)
            logger.warning("Continuing without INDX index data")

    # Get in-scope funds
    in_scope_funds = get_in_scope_funds(mutual_funds, fund_index_map)

    # Fetch TASE (מאי"ה) index data (enabled by default, skip with --skip-tase-data)
    tase_data: Optional[TaseIndexData] = None
    tase_fund_count = sum(1 for f in in_scope_funds if f.data_source == 'מאי"ה')

    if args.skip_tase_data:
        if tase_fund_count > 0:
            logger.info("Skipping TASE data fetch (--skip-tase-data). %d funds will show placeholder values.", tase_fund_count)
    elif tase_fund_count > 0:
        # Get API key from environment
        tase_api_key = os.environ.get("TASE_API_KEY")

        if not tase_api_key:
            logger.warning("TASE_API_KEY not found in environment. Set it in .env file.")
            logger.warning("Continuing without TASE data")
        else:
            # Find unique TASE index IDs from in-scope funds
            tase_index_ids = list(set(
                fund.index_id for fund in in_scope_funds
                if fund.data_source == 'מאי"ה'
            ))

            logger.info("Fetching TASE data for %d unique indices", len(tase_index_ids))
            # Generate date range for TASE API
            today = dt.date.today()
            from_date = dt.date(today.year, 1, 1)

            try:
                tase_data = fetch_tase_index_data(
                    tase_index_ids,
                    tase_api_key,
                    from_date,
                    today
                )
                logger.info("Fetched TASE data for %d indices", len(tase_data.data))
            except Exception as e:
                logger.warning("Failed to fetch TASE data: %s", e)
                logger.warning("Continuing without TASE data")

    if not in_scope_funds:
        logger.error("No in-scope funds found!")
        sys.exit(1)

    # Find in-scope funds missing from index table (Mizrahi + has variable fees)
    index_fund_ids = set(fund_index_map.keys())
    missing_funds: list[MutualFund] = []
    for fund_id, fund in mutual_funds.items():
        if is_fund_in_scope(fund) and fund_id not in index_fund_ids:
            missing_funds.append(fund)
    logger.info("Found %d in-scope funds missing from index table", len(missing_funds))

    # Generate date range
    dates = generate_date_range()

    # Generate output
    logger.info("Generating output workbook...")
    generate_output_workbook(in_scope_funds, dates, output_path, missing_funds, bfix_data, bloomberg_data, indx_data, tase_data)

    logger.info("=" * 60)
    logger.info("Complete!")
    logger.info("  Funds processed: %d", len(in_scope_funds))
    logger.info("  Date range: %s to %s (%d days)",
               dates[0], dates[-1], len(dates))
    logger.info("  Output file: %s", output_path)
    if bfix_data:
        logger.info("  BFIX columns: %d", len(bfix_data.columns))
    if bloomberg_data:
        logger.info("  Bloomberg index columns: %d", len(bloomberg_data.columns))
    if indx_data:
        logger.info("  INDX indices: %d", len(indx_data.data))
    if tase_data:
        logger.info("  TASE indices: %d", len(tase_data.data))
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
