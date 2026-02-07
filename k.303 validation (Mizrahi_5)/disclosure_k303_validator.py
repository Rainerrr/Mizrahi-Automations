#!/usr/bin/env python3
"""
Disclosure Report K.303 Validator (Migdal Variant)

Validates K.303 disclosure reports and produces an Excel report with:
- Summary sheet
- Check specifications details
- Check status summary
- Per-check exception sheets

Inputs:
  - Mutual Funds List CSV: filtered to Mizrahi trustee funds
  - Current month disclosure report CSV
  - Previous month disclosure report CSV
  - K.303 checklist CSV (for פירוט בדיקות sheet)

Outputs:
  - XLSX workbook with validation results

Example:
  python disclosure_k303_validator.py \
    --mutual-funds-list "Mutual_Funds_List.xlsx - Worksheet.csv" \
    --current-report "disclosure_migdal_current_month.csv" \
    --previous-report "disclosure_migdal_previous_month.xlsx - Sheet1.csv" \
    --spec-file "בדיקת דוח גילוי נאות ק.303.xlsx - גיליון1.csv" \
    --output-xlsx "disclosure_validation_output.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import logging
import sys
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Add parent directory to path for shared imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.constants import FUND_MANAGER_CODES_K303

# Global log directory for current run
LOG_RUN_DIR: Optional[Path] = None

# Path to code index file (relative to this script)
CODE_INDEX_PATH = Path(__file__).parent / "k303_code_index.xlsx"

# Path to default spec file (relative to this script)
DEFAULT_SPEC_PATH = Path(__file__).parent / "k303_spec.xlsx"

# K.303 Code Index - hierarchical code definitions
# Level 1 (2 digits), Level 2 (4 digits), Level 3 (6 digits), Level 4 (8 digits)
K303_CODE_INDEX: dict[str, str] = {
    # Level 1 - רמה 1
    "01": "מניות",
    "02": "סחורות",
    "03": 'אג"ח',
    "04": "אחר",
    "05": "מזומנים ופקדונות",
    "06": 'מט"ח',
    "07": 'התפלגות אג"ח לפי דירוגים',
    "08": 'מח"מ תיק האג"ח',
    "09": "סימול אודות מנהל השקעות חיצוני בקרן",
    "10": "השווי הנקי של נכסי הקרן",
    # Level 2 - רמה 2
    "0101": "מניות הנסחרות בארץ",
    "0102": 'מניות הנסחרות בחו"ל',
    "0301": 'אג"ח הנסחרות בארץ',
    "0302": 'אג"ח הנסחרות בחו"ל',
    "0501": 'בש"ח',
    "0502": 'במט"ח',
    "0601": 'חשיפה כוללת למט"ח',
    "0602": "פירוט חשיפה למטבעות הבאים",
    "0701": 'אג"ח מדורגות בדירוג השקעה',
    "0702": 'אג"ח שאינן מדורגות בדירוג השקעה או שאינן מדורגות כלל',
    "0801": 'מח"מ תיק האג"ח',
    "0802": 'מח"מ תיק אג"ח בארץ',
    "0803": 'מח"מ תיק אג"ח בחו"ל',
    # Level 3 - רמה 3
    "010101": 'ת"א- 125',
    "010102": "יתר מניות והמירים",
    "010201": "צפון אמריקה (DM)",
    "010202": "צפון אמריקה (None-DM)",
    "010203": "מרכז ודרום אמריקה (DM)",
    "010204": "מרכז ודרום אמריקה (None-DM)",
    "010205": "אפריקה (DM)",
    "010206": "אפריקה (None-DM)",
    "010207": "אירופה (DM)",
    "010208": "אירופה (None-DM)",
    "010209": "אסיה (DM)",
    "010210": "אסיה (None-DM)",
    "010211": "ישראל",
    "010212": "אוסטרליה וניו-זילנד (DM)",
    "010213": "גלובאלי/אחר (DM)",
    "010214": "גלובאלי/אחר (None-DM)",
    "030101": 'אג"ח ממשלתי',
    "030102": 'אג"ח קונצרני',
    "030103": "תעודות חוב",
    "030104": "תעודות פיקדון",
    "030201": 'אג"ח מדינה ובערבות מדינה',
    "030202": 'אג"ח קונצרני בחו"ל',
    "030203": 'אג"ח שהונפקו על ידי יישויות אזוריות שאינן בערבות מדינה',
    "060201": "דולר של ארצות הברית של אמריקה",
    "060202": "יורו",
    "060203": "לירה שטרלינג",
    "060204": "כתר שבדי",
    "060205": "פרנק שוויצרי",
    "060206": "דולר קנדי",
    "060207": "כתר דני",
    "060208": "רנד דרום אפריקאי",
    "060209": "דולר אוסטרלי",
    "060210": "כתר נורבגי",
    "060211": "יין יפני",
    "060212": "דולר ניו זילנדי",
    "060213": "דולר הונג קונג",
    "060214": "יואן סיני",
    "060215": "דולר סנגפורי",
    "060216": "דולר טאיווני",
    "060217": "באת תאילנדי",
    "060218": "וואן דרום קוראיני",
    "060219": "רופיה הודית",
    "060220": "רובל רוסי",
    "060221": "פזו מקסיקני",
    "060222": "פורינט הונגרי",
    "060223": "ריאל ברזילאי",
    "060224": "לירה תורכית",
    "060225": "זלוטי פולני",
    "060226": "קרונה איסלנדית",
    "060227": "קרונה צ'כית (כתר צ'כי)",
    "060228": "פזו פיליפיני",
    "060229": "לירה מצרית",
    "060230": "דונג וייטנאמי",
    "070101": "גבוה (AA ומעלה)",
    "070102": "בינוני (BBB עד למטה מ- AA)",
    "070201": "נמוך (למטה מ- BBB)",
    "070202": 'אג"ח שאינן מדורגות',
    "080201": "ממשלתי שקלי",
    "080202": "ממשלתי צמוד",
    "080203": 'ממשלתי צמוד מט"ח',
    "080204": "קונצרני שקלי",
    "080205": "קונצרני צמוד",
    "080206": 'קונצרני צמוד מט"ח',
    # Level 4 - רמה 4
    "03010101": "ממשלתי צמוד מדד",
    "03010102": "ממשלתי לא צמוד",
    "03010103": 'ממשלתי צמוד מט"ח',
    "03010104": "ממשלתי אחר",
    "03010201": "קונצרני צמוד מדד",
    "03010202": "קונצרני שקלי ריבית קבועה",
    "03010203": "קונצרני שקלי ריבית משתנה",
    "03010204": 'קונצרני -צמוד מט"ח/אחר',
    "03010301": "שקלי",
    "03010302": "צמוד",
    "03010303": 'מט"ח',
    "03010401": "שקלי",
    "03010402": "צמוד",
    "03010403": 'מט"ח',
    "03020101": "צפון אמריקה (DM)",
    "03020102": "צפון אמריקה (None-DM)",
    "03020103": "מרכז ודרום אמריקה (DM)",
    "03020104": "מרכז ודרום אמריקה (None-DM)",
    "03020105": "אפריקה (DM)",
    "03020106": "אפריקה (None-DM)",
    "03020107": "אירופה (DM)",
    "03020108": "אירופה (None-DM)",
    "03020109": "אסיה (DM)",
    "03020110": "אסיה (None-DM)",
    "03020111": "ישראל",
    "03020112": "אוסטרליה וניו-זילנד (DM)",
    "03020113": "גלובאלי/אחר (DM)",
    "03020114": "גלובאלי/אחר (None-DM)",
    "03020201": "צפון אמריקה (DM)",
    "03020202": "צפון אמריקה (None-DM)",
    "03020203": "מרכז ודרום אמריקה (DM)",
    "03020204": "מרכז ודרום אמריקה (None-DM)",
    "03020205": "אפריקה (DM)",
    "03020206": "אפריקה (None-DM)",
    "03020207": "אירופה (DM)",
    "03020208": "אירופה (None-DM)",
    "03020209": "אסיה (DM)",
    "03020210": "אסיה (None-DM)",
    "03020211": "ישראל",
    "03020212": "אוסטרליה וניו-זילנד (DM)",
    "03020213": "גלובאלי/אחר (DM)",
    "03020214": "גלובאלי/אחר (None-DM)",
    "03020301": "צפון אמריקה (DM)",
    "03020302": "צפון אמריקה (None-DM)",
    "03020303": "מרכז ודרום אמריקה (DM)",
    "03020304": "מרכז ודרום אמריקה (None-DM)",
    "03020305": "אפריקה (DM)",
    "03020306": "אפריקה (None-DM)",
    "03020307": "אירופה (DM)",
    "03020308": "אירופה (None-DM)",
    "03020309": "אסיה (DM)",
    "03020310": "אסיה (None-DM)",
    "03020311": "ישראל",
    "03020312": "אוסטרליה וניו-זילנד (DM)",
    "03020313": "גלובאלי/אחר (DM)",
    "03020314": "גלובאלי/אחר (None-DM)",
}


def get_full_code_description(code: str) -> str:
    """
    Get the full hierarchical description for a K.303 code.

    Combines descriptions from all parent levels, removing redundant words.
    Example: 0501 -> "מזומנים ופקדונות בש"ח" (combines 05 + 0501)
    Example: 03010101 -> "אג"ח הנסחרות בארץ ממשלתי צמוד מדד"

    Args:
        code: The K.303 code (2, 4, 6, or 8 digits)

    Returns:
        Combined description string, or empty string if code not found
    """
    if not code:
        return ""

    code = str(code).strip()

    # Determine parent codes based on code length
    # 2 digits: just level 1
    # 4 digits: level 1 (first 2) + level 2
    # 6 digits: level 1 + level 2 (first 4) + level 3
    # 8 digits: level 1 + level 2 + level 3 (first 6) + level 4
    parent_codes = []
    if len(code) >= 2:
        parent_codes.append(code[:2])
    if len(code) >= 4:
        parent_codes.append(code[:4])
    if len(code) >= 6:
        parent_codes.append(code[:6])
    if len(code) >= 8:
        parent_codes.append(code[:8])

    # Build combined description
    combined = ""
    for parent_code in parent_codes:
        desc = K303_CODE_INDEX.get(parent_code, "")
        if not desc:
            continue

        if not combined:
            combined = desc
        else:
            # Remove redundant leading words from the new description
            combined = _merge_descriptions(combined, desc)

    # Remove any non-consecutive duplicate words
    combined = _remove_duplicate_words(combined)

    return combined


def _merge_descriptions(current: str, new_desc: str) -> str:
    """
    Merge two descriptions, removing redundant leading words from new_desc.

    If new_desc starts with words that are at the end of current,
    those words are skipped to avoid repetition.

    Example: "אג"ח ממשלתי" + "ממשלתי צמוד מדד" -> "אג"ח ממשלתי צמוד מדד"
    """
    if not new_desc:
        return current
    if not current:
        return new_desc

    current_words = current.split()
    new_words = new_desc.split()

    # Find how many leading words of new_desc match trailing words of current
    max_overlap = min(len(current_words), len(new_words))
    overlap = 0

    for i in range(1, max_overlap + 1):
        # Check if last i words of current match first i words of new_desc
        if current_words[-i:] == new_words[:i]:
            overlap = i

    # If there's overlap, skip those words from new_desc
    if overlap > 0:
        remaining_words = new_words[overlap:]
        if remaining_words:
            return current + " " + " ".join(remaining_words)
        else:
            return current
    else:
        return current + " " + new_desc


def _remove_duplicate_words(text: str) -> str:
    """
    Remove duplicate words from text, keeping only the first occurrence.

    Example: "אג"ח הנסחרות בארץ אג"ח ממשלתי" -> "אג"ח הנסחרות בארץ ממשלתי"
    """
    if not text:
        return text

    words = text.split()
    seen: set[str] = set()
    result: list[str] = []

    for word in words:
        if word not in seen:
            seen.add(word)
            result.append(word)

    return " ".join(result)


def code_desc(code: str) -> str:
    """
    Format a code with its full hierarchical description.

    Example: "03010101" -> "03010101 (אג"ח הנסחרות בארץ ממשלתי צמוד מדד)"
    """
    desc = get_full_code_description(code)
    if desc:
        return f"{code} ({desc})"
    return code


def codes_desc(codes: list[str], separator: str = "/") -> str:
    """
    Format multiple codes with their descriptions.

    Example: ["03010202", "03010203"] -> "03010202/03010203 (קונצרני שקלי ריבית קבועה/משתנה)"
    """
    if not codes:
        return ""
    if len(codes) == 1:
        return code_desc(codes[0])

    codes_str = separator.join(codes)
    descs = [get_full_code_description(c) for c in codes]
    # Filter out empty descriptions
    descs = [d for d in descs if d]
    if descs:
        return f"{codes_str} ({separator.join(descs)})"
    return codes_str


# Loggers
logger = logging.getLogger(__name__)
logger_chk1a = logging.getLogger('CHK_1A')
logger_chk1b = logging.getLogger('CHK_1B')
logger_chk2a = logging.getLogger('CHK_2A')
logger_chk2b = logging.getLogger('CHK_2B')
logger_chk3 = logging.getLogger('CHK_3')


def setup_logging(log_base_dir: Path = Path(__file__).parent.parent / "log") -> Path:
    """Set up logging with separate files for each check."""
    global LOG_RUN_DIR

    from datetime import datetime as datetime_module
    timestamp = datetime_module.now().strftime("%Y%m%d_%H%M%S")
    short_uuid = str(uuid.uuid4())[:8]
    run_id = f"{timestamp}_{short_uuid}"

    run_dir = log_base_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    LOG_RUN_DIR = run_dir

    log_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    detailed_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(name)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(log_format)

    logger.setLevel(logging.DEBUG)
    logger.addHandler(console_handler)
    main_file_handler = logging.FileHandler(run_dir / "main.log", encoding='utf-8')
    main_file_handler.setLevel(logging.DEBUG)
    main_file_handler.setFormatter(log_format)
    logger.addHandler(main_file_handler)

    spec_loggers = [
        (logger_chk1a, "chk1a_fund_completeness.log", "CHK_1A - Fund Completeness"),
        (logger_chk1b, "chk1b_date_validity.log", "CHK_1B - Date Validity"),
        (logger_chk2a, "chk2a_prev_month.log", "CHK_2A - Previous Month Comparison"),
        (logger_chk2b, "chk2b_exposure_profile.log", "CHK_2B - Exposure Profile Check"),
        (logger_chk3, "chk3_combinations.log", "CHK_3 - Within-Month Combinations"),
    ]

    for spec_logger, filename, description in spec_loggers:
        spec_logger.setLevel(logging.DEBUG)
        file_handler = logging.FileHandler(run_dir / filename, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(detailed_format)
        spec_logger.addHandler(file_handler)
        spec_logger.addHandler(main_file_handler)
        spec_logger.info("=" * 70)
        spec_logger.info(description)
        spec_logger.info("=" * 70)

    logger.info("Log directory created: %s", run_dir)
    return run_dir


# -----------------------------
# Configuration
# -----------------------------

MIZRAHI_TRUSTEE_NAME = 'מזרחי טפחות חברה לנאמנות בע"מ'

FUND_MANAGER_NAMES = list(FUND_MANAGER_CODES_K303.keys())

# Disclosure report columns
D_COL_FUND_NO = "מספר קרן"
D_COL_FUND_NAME = "שם קרן"
D_COL_LEVEL_1 = "רמה 1"
D_COL_LEVEL_2 = "רמה 2"
D_COL_LEVEL_3 = "רמה 3"
D_COL_LEVEL_4 = "רמה 4"
D_COL_PERCENT = "%מקרן"
D_COL_EXTRA_DATA = "נתונים נוספים"
D_COL_REPORT_DATE = "תאריך דוח"
D_COL_RECORD_NO = "מס.רשומה"
D_COL_TOTAL_RECORDS = "סהכ רשומות"
D_COL_MANAGER_NO = "מס.מנהל ברשם"

# Mutual Funds List columns
MF_COL_FUND_ID = "מספר בורסה"
MF_COL_TRUSTEE = "שם נאמן"
MF_COL_MANAGER = "שם מנהל"
MF_COL_EXPOSURE_PROFILE = "פרופיל החשיפה"
MF_COL_FUND_TYPE = "סוג הקרן"

# Excel styling
HEADER_FONT = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
DEFAULT_FONT = Font(name='Calibri', size=11)

WRAP_HEADERS = {
    "שם קרן",
    "סיבה",
    "הערות",
    "שדה",
    "ערך",
    "פירוט הבדיקה",
    "אופן ביצוע",
}

HEBREW_MONTHS = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר"
}

# Exposure profile mapping - determines equity/FX exposure from profile code
# Profile format: <digit><letter> e.g., "0A", "3B", "4D"
# First digit: max equity exposure (code 01)
# Letter: max FX exposure (code 06)
EQUITY_EXPOSURE_PROFILES = {
    '0': 0,    # 0% equity
    '1': 10,   # up to 10% equity
    '2': 30,   # up to 30% equity
    '3': 50,   # up to 50% equity
    '4': 120,  # up to 120% equity
    '5': 200,  # up to 200% equity
    '6': None, # above 200% (unlimited)
}

FX_EXPOSURE_PROFILES = {
    '0': 0,    # no FX exposure
    'A': 10,   # up to 10% FX
    'B': 30,   # up to 30% FX
    'C': 50,   # up to 50% FX
    'D': 120,  # up to 120% FX
    'E': 200,  # up to 200% FX
    'F': None, # above 200% (unlimited)
}


# -----------------------------
# Data structures
# -----------------------------

@dataclass
class DisclosureRow:
    """A single row from the disclosure report."""
    row_num: int
    fund_no: Optional[int]
    fund_name: Optional[str]
    level_1: Optional[str]
    level_2: Optional[str]
    level_3: Optional[str]
    level_4: Optional[str]
    percent_from_fund: Optional[float]
    extra_data: Optional[str]
    report_date: Optional[dt.date]
    record_no: Optional[int]
    total_records: Optional[int]
    manager_no: Optional[str]

    @property
    def effective_code(self) -> Optional[str]:
        """Get the most granular non-empty level code."""
        for level in [self.level_4, self.level_3, self.level_2, self.level_1]:
            if level and level.strip():
                return level.strip()
        return None


@dataclass
class MutualFund:
    """A fund from the mutual funds list."""
    fund_id: int
    fund_name: str
    trustee_name: str
    manager_name: str
    exposure_profile: Optional[str]
    fund_type: Optional[str] = None  # סוג הקרן (column L in mutual funds list)


@dataclass
class ExceptionRow:
    """An exception found during validation."""
    check_id: str
    reason: str
    fund_no: Optional[int] = None
    fund_name: Optional[str] = None
    effective_code: Optional[str] = None
    percent_from_fund: Optional[float] = None
    report_date: Optional[dt.date] = None
    row_num: Optional[int] = None
    extra_info: dict = field(default_factory=dict)


# -----------------------------
# Helpers
# -----------------------------

# Regex pattern for illegal Excel XML characters
import re
_ILLEGAL_XML_CHARS_RE = re.compile(
    '[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


def _clean_excel_string(s: Optional[str]) -> Optional[str]:
    """Remove illegal XML characters that cause Excel file corruption."""
    if s is None:
        return None
    # Remove illegal XML characters
    return _ILLEGAL_XML_CHARS_RE.sub('', s)


def _norm_spaces(s: str) -> str:
    return " ".join((s or "").strip().split())


def _to_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    s = _clean_excel_string(s)  # Remove illegal XML characters
    return s if s else None


def _to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except Exception:
        return None


def _parse_ddmmyyyy(v: Any) -> Optional[dt.date]:
    """Parse date in DDMMYYYY format (e.g., 30112025 for Nov 30, 2025)."""
    if v is None or v == "":
        return None
    s = str(int(float(str(v).strip()))) if v else ""
    s = s.strip().zfill(8)
    try:
        day = int(s[0:2])
        month = int(s[2:4])
        year = int(s[4:8])
        return dt.date(year, month, day)
    except Exception:
        return None


def _format_report_month_hebrew(report_month: str) -> str:
    """Convert YYYY-MM to Hebrew format like 'נובמבר 2025'."""
    try:
        year, month = report_month.split("-")
        return f"{HEBREW_MONTHS[int(month)]} {year}"
    except (ValueError, KeyError):
        return report_month


def _fmt_date(d: Optional[dt.date]) -> str:
    return d.strftime("%d-%m-%Y") if d else ""


# -----------------------------
# Data loaders
# -----------------------------

def load_mutual_funds_csv(path: Path) -> dict[int, MutualFund]:
    """Load mutual funds list from CSV, return dict keyed by fund ID."""
    funds: dict[int, MutualFund] = {}

    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [h.strip() for h in reader.fieldnames]

        for row in reader:
            fund_id = _to_int(row.get(MF_COL_FUND_ID))
            if fund_id is None:
                continue

            fund = MutualFund(
                fund_id=fund_id,
                fund_name=_to_str(row.get("שם קרן בעברית")) or "",
                trustee_name=_to_str(row.get(MF_COL_TRUSTEE)) or "",
                manager_name=_to_str(row.get(MF_COL_MANAGER)) or "",
                exposure_profile=_to_str(row.get(MF_COL_EXPOSURE_PROFILE)),
                fund_type=_to_str(row.get(MF_COL_FUND_TYPE)),
            )
            funds[fund_id] = fund

    logger.info("Loaded %d funds from mutual funds list", len(funds))
    return funds


def load_disclosure_report_csv(path: Path) -> list[DisclosureRow]:
    """Load disclosure report from CSV."""
    rows: list[DisclosureRow] = []

    # Try different encodings
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1255', 'iso-8859-8', 'windows-1252']

    for encoding in encodings_to_try:
        try:
            with open(path, 'r', encoding=encoding) as f:
                reader = csv.DictReader(f)
                if reader.fieldnames:
                    # Strip whitespace and \r from header names
                    reader.fieldnames = [h.strip().replace('\r', '') for h in reader.fieldnames]

                for row_num, csv_row in enumerate(reader, start=2):
                    row = DisclosureRow(
                        row_num=row_num,
                        fund_no=_to_int(csv_row.get(D_COL_FUND_NO)),
                        fund_name=_to_str(csv_row.get(D_COL_FUND_NAME)),
                        level_1=_to_str(csv_row.get(D_COL_LEVEL_1)),
                        level_2=_to_str(csv_row.get(D_COL_LEVEL_2)),
                        level_3=_to_str(csv_row.get(D_COL_LEVEL_3)),
                        level_4=_to_str(csv_row.get(D_COL_LEVEL_4)),
                        percent_from_fund=_to_float(csv_row.get(D_COL_PERCENT)),
                        extra_data=_to_str(csv_row.get(D_COL_EXTRA_DATA)),
                        report_date=_parse_ddmmyyyy(csv_row.get(D_COL_REPORT_DATE)),
                        record_no=_to_int(csv_row.get(D_COL_RECORD_NO)),
                        total_records=_to_int(csv_row.get(D_COL_TOTAL_RECORDS)),
                        manager_no=_to_str(csv_row.get(D_COL_MANAGER_NO)),
                    )
                    rows.append(row)
            logger.info("Loaded %d rows from disclosure report (encoding: %s): %s", len(rows), encoding, path.name)
            return rows
        except UnicodeDecodeError:
            rows = []
            continue

    raise ValueError(f"Could not decode file {path} with any supported encoding")


def load_disclosure_report_xlsx(path: Path) -> list[DisclosureRow]:
    """Load disclosure report from XLSX."""
    import shutil
    import tempfile

    rows: list[DisclosureRow] = []

    # If file has wrong extension, copy to temp file with .xlsx
    if path.suffix.lower() != '.xlsx':
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            shutil.copy2(path, tmp.name)
            actual_path = Path(tmp.name)
    else:
        actual_path = path

    try:
        wb = openpyxl.load_workbook(actual_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Build header map from row 1
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val:
                headers[str(val).strip().replace('\r', '')] = col

        for row_num in range(2, ws.max_row + 1):
            def get_val(col_name: str, col_idx_cache=headers):
                col_idx = col_idx_cache.get(col_name)
                if col_idx:
                    return ws.cell(row_num, col_idx).value
                return None

            row = DisclosureRow(
                row_num=row_num,
                fund_no=_to_int(get_val(D_COL_FUND_NO)),
                fund_name=_to_str(get_val(D_COL_FUND_NAME)),
                level_1=_to_str(get_val(D_COL_LEVEL_1)),
                level_2=_to_str(get_val(D_COL_LEVEL_2)),
                level_3=_to_str(get_val(D_COL_LEVEL_3)),
                level_4=_to_str(get_val(D_COL_LEVEL_4)),
                percent_from_fund=_to_float(get_val(D_COL_PERCENT)),
                extra_data=_to_str(get_val(D_COL_EXTRA_DATA)),
                report_date=_parse_ddmmyyyy(get_val(D_COL_REPORT_DATE)),
                record_no=_to_int(get_val(D_COL_RECORD_NO)),
                total_records=_to_int(get_val(D_COL_TOTAL_RECORDS)),
                manager_no=_to_str(get_val(D_COL_MANAGER_NO)),
            )
            # Skip empty rows
            if row.fund_no is not None:
                rows.append(row)

        wb.close()
    finally:
        # Clean up temp file if we created one
        if actual_path != path:
            import os
            try:
                os.unlink(actual_path)
            except:
                pass

    logger.info("Loaded %d rows from disclosure report (XLSX): %s", len(rows), path.name)
    return rows


def load_disclosure_report(path: Path) -> list[DisclosureRow]:
    """Load disclosure report - auto-detect format (CSV or XLSX)."""
    # Check file magic bytes to detect actual format
    with open(path, 'rb') as f:
        magic = f.read(4)

    # PK is ZIP (XLSX) magic
    if magic[:2] == b'PK':
        logger.info("Detected XLSX format for: %s", path.name)
        return load_disclosure_report_xlsx(path)
    else:
        logger.info("Detected CSV format for: %s", path.name)
        return load_disclosure_report_csv(path)


def load_code_index(path: Path = CODE_INDEX_PATH) -> dict[str, str]:
    """Load code definitions from k303_code_index.xlsx.

    Returns dict mapping code -> description (e.g., "01" -> "מניות").
    """
    code_definitions: dict[str, str] = {}

    if not path.exists():
        logger.warning("Code index file not found: %s", path)
        return code_definitions

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["כל הקודים"]

        # Data starts at row 4 (row 1 is title, row 2 is empty, row 3 is headers)
        # Column B = code, Column C = description
        for row_idx in range(4, ws.max_row + 1):
            code = ws.cell(row_idx, 2).value  # Column B - קוד
            description = ws.cell(row_idx, 3).value  # Column C - תאור הקוד

            if code and description:
                code_str = str(code).strip()
                desc_str = str(description).strip()
                if code_str and desc_str:
                    code_definitions[code_str] = desc_str

        wb.close()
        logger.info("Loaded %d code definitions from code index", len(code_definitions))
    except Exception as e:
        logger.warning("Failed to load code index: %s", e)

    return code_definitions


def get_in_scope_fund_ids(
    funds: dict[int, MutualFund],
    trustee_name: str,
    manager_name: Optional[str] = None,
) -> set[int]:
    """Get fund IDs filtered by trustee and optionally by manager name."""
    result = set()
    for fund_id, fund in funds.items():
        if trustee_name and trustee_name not in _norm_spaces(fund.trustee_name):
            continue
        if manager_name and manager_name not in _norm_spaces(fund.manager_name):
            continue
        result.add(fund_id)
    logger.info("Found %d in-scope funds (trustee=%s, manager=%s)", len(result), trustee_name, manager_name)
    return result


# -----------------------------
# Checks
# -----------------------------

def check_1a_fund_completeness(
    disclosure_rows: list[DisclosureRow],
    mizrahi_fund_ids: set[int],
    all_funds: dict[int, MutualFund]
) -> list[ExceptionRow]:
    """Check 1א: Validate fund completeness between mutual funds list and disclosure report."""
    logger_chk1a.info("Starting fund completeness check")

    # Get unique fund IDs from disclosure report
    funds_in_report = set(r.fund_no for r in disclosure_rows if r.fund_no is not None)

    exceptions: list[ExceptionRow] = []

    # Funds in mutual list but missing from report
    missing_from_report = mizrahi_fund_ids - funds_in_report
    for fund_id in sorted(missing_from_report):
        fund = all_funds.get(fund_id)
        fund_name = fund.fund_name if fund else ""
        logger_chk1a.warning("Fund %d (%s) is in mutual funds list but missing from report", fund_id, fund_name)
        exceptions.append(ExceptionRow(
            check_id="1א",
            reason="קרן חסרה בדוח",
            fund_no=fund_id,
            fund_name=fund_name,
        ))

    # Funds in report but missing from mutual list (unexpected)
    extra_in_report = funds_in_report - mizrahi_fund_ids
    # Only flag funds that are completely missing from the mutual funds list
    # (funds under different trustees are expected)
    for fund_id in sorted(extra_in_report):
        if fund_id not in all_funds:
            # Get name from disclosure report
            fund_name = ""
            for r in disclosure_rows:
                if r.fund_no == fund_id:
                    fund_name = r.fund_name or ""
                    break
            logger_chk1a.warning("Fund %d (%s) is in report but not in mutual funds list", fund_id, fund_name)
            exceptions.append(ExceptionRow(
                check_id="1א",
                reason="קרן לא קיימת ברשימת קרנות",
                fund_no=fund_id,
                fund_name=fund_name,
            ))

    logger_chk1a.info("Fund completeness check completed: %d exceptions", len(exceptions))
    return exceptions


def check_1b_report_month_validity(
    disclosure_rows: list[DisclosureRow],
    expected_month: str,
    mizrahi_fund_ids: set[int]
) -> list[ExceptionRow]:
    """Check 1ב: Validate that report_date matches expected report month."""
    logger_chk1b.info("Starting date validity check for month: %s", expected_month)

    year, month = expected_month.split("-")
    expected_year = int(year)
    expected_month_num = int(month)

    exceptions: list[ExceptionRow] = []

    for row in disclosure_rows:
        # Only check in-scope funds
        if row.fund_no not in mizrahi_fund_ids:
            continue

        if row.report_date is None:
            logger_chk1b.warning("Row %d: Missing report date for fund %d", row.row_num, row.fund_no)
            exceptions.append(ExceptionRow(
                check_id="1ב",
                reason="תאריך דוח חסר",
                fund_no=row.fund_no,
                fund_name=row.fund_name,
                effective_code=row.effective_code,
                row_num=row.row_num,
            ))
            continue

        if row.report_date.year != expected_year or row.report_date.month != expected_month_num:
            logger_chk1b.warning(
                "Row %d: Date mismatch for fund %d - expected %s, got %s",
                row.row_num, row.fund_no, expected_month, row.report_date
            )
            exceptions.append(ExceptionRow(
                check_id="1ב",
                reason=f"תאריך לא תואם (צפוי: {expected_month})",
                fund_no=row.fund_no,
                fund_name=row.fund_name,
                effective_code=row.effective_code,
                report_date=row.report_date,
                row_num=row.row_num,
            ))

    logger_chk1b.info("Date validity check completed: %d exceptions", len(exceptions))
    return exceptions


def check_2a_prev_month_comparison(
    current_rows: list[DisclosureRow],
    prev_rows: list[DisclosureRow],
    mizrahi_fund_ids: set[int],
    all_funds: dict[int, MutualFund]
) -> list[ExceptionRow]:
    """Check 2א: Compare disclosure codes between current and previous month."""
    logger_chk2a.info("Starting previous month comparison check")

    # Build lookup: (fund_no, effective_code) -> percent
    def build_lookup(rows: list[DisclosureRow], fund_ids: set[int]) -> dict[tuple[int, str], float]:
        lookup: dict[tuple[int, str], float] = {}
        for row in rows:
            if row.fund_no not in fund_ids:
                continue
            code = row.effective_code
            if code and row.percent_from_fund is not None:
                key = (row.fund_no, code)
                # If multiple rows with same code, sum them (shouldn't happen but be safe)
                lookup[key] = lookup.get(key, 0) + row.percent_from_fund
        return lookup

    current_lookup = build_lookup(current_rows, mizrahi_fund_ids)
    prev_lookup = build_lookup(prev_rows, mizrahi_fund_ids)

    # Get fund names from current rows
    fund_names: dict[int, str] = {}
    for row in current_rows:
        if row.fund_no and row.fund_name:
            fund_names[row.fund_no] = row.fund_name
    for row in prev_rows:
        if row.fund_no and row.fund_name and row.fund_no not in fund_names:
            fund_names[row.fund_no] = row.fund_name

    exceptions: list[ExceptionRow] = []

    all_keys = set(current_lookup.keys()) | set(prev_lookup.keys())

    for fund_no, code in sorted(all_keys):
        current_pct = current_lookup.get((fund_no, code))
        prev_pct = prev_lookup.get((fund_no, code))

        # Get fund type from mutual funds list
        fund = all_funds.get(fund_no)
        fund_type = fund.fund_type if fund else None

        if current_pct is None and prev_pct is not None:
            # Code removed
            logger_chk2a.warning("Fund %d: Code %s removed (was %.2f%%)", fund_no, code, prev_pct)
            exceptions.append(ExceptionRow(
                check_id="2א",
                reason=f"קוד נעלם (היה: {prev_pct:.2f}%)",
                fund_no=fund_no,
                fund_name=fund_names.get(fund_no),
                effective_code=code,
                percent_from_fund=None,
                extra_info={"prev_pct": prev_pct, "fund_type": fund_type},
            ))
        elif current_pct is not None and prev_pct is None:
            # Code added
            logger_chk2a.warning("Fund %d: Code %s added (now %.2f%%)", fund_no, code, current_pct)
            exceptions.append(ExceptionRow(
                check_id="2א",
                reason=f"קוד חדש (כעת: {current_pct:.2f}%)",
                fund_no=fund_no,
                fund_name=fund_names.get(fund_no),
                effective_code=code,
                percent_from_fund=current_pct,
                extra_info={"prev_pct": None, "fund_type": fund_type},
            ))
        elif current_pct is not None and prev_pct is not None:
            # Check delta
            delta = abs(current_pct - prev_pct)
            if delta > 10.0:
                logger_chk2a.warning(
                    "Fund %d: Code %s changed by %.2f%% (%.2f%% -> %.2f%%)",
                    fund_no, code, delta, prev_pct, current_pct
                )
                exceptions.append(ExceptionRow(
                    check_id="2א",
                    reason=f"סטייה > 10% (שינוי: {delta:.2f}%)",
                    fund_no=fund_no,
                    fund_name=fund_names.get(fund_no),
                    effective_code=code,
                    percent_from_fund=current_pct,
                    extra_info={"prev_pct": prev_pct, "delta": delta, "fund_type": fund_type},
                ))

    logger_chk2a.info("Previous month comparison completed: %d exceptions", len(exceptions))
    return exceptions


def check_2b_exposure_profile(
    disclosure_rows: list[DisclosureRow],
    all_funds: dict[int, MutualFund],
    mizrahi_fund_ids: set[int]
) -> list[ExceptionRow]:
    """Check 2ב: Cross-check disclosure exposure codes against fund's exposure profile."""
    logger_chk2b.info("Starting exposure profile check")

    exceptions: list[ExceptionRow] = []

    # Group rows by fund
    fund_rows: dict[int, list[DisclosureRow]] = defaultdict(list)
    for row in disclosure_rows:
        if row.fund_no in mizrahi_fund_ids:
            fund_rows[row.fund_no].append(row)

    for fund_no, rows in fund_rows.items():
        fund = all_funds.get(fund_no)
        if not fund or not fund.exposure_profile:
            continue

        profile = fund.exposure_profile.strip()
        if len(profile) < 2:
            continue

        equity_code = profile[0]
        fx_code = profile[1].upper()

        max_equity = EQUITY_EXPOSURE_PROFILES.get(equity_code, None)
        max_fx = FX_EXPOSURE_PROFILES.get(fx_code, None)

        # Aggregate exposure by code prefix for this fund
        equity_total = 0.0  # code 01 - מניות
        fx_total = 0.0      # code 06 - מט"ח

        for row in rows:
            code = row.effective_code
            if not code:
                continue
            pct = row.percent_from_fund or 0
            if code.startswith("01"):
                equity_total += pct
            elif code.startswith("06") and not code.startswith("0602"):
                fx_total += pct

        # Check equity exposure (code 01 - מניות) against profile limit
        if max_equity is not None and equity_total > max_equity:
            logger_chk2b.warning(
                "Fund %d: Profile %s allows max %d%% equity but total is %.2f%%",
                fund_no, profile, max_equity, equity_total
            )
            exceptions.append(ExceptionRow(
                check_id="2ב",
                reason=f"פרופיל {profile} מתיר עד {max_equity}% מניות אך סה\"כ חשיפה = {equity_total:.2f}%",
                fund_no=fund_no,
                fund_name=rows[0].fund_name if rows else "",
                effective_code="01",
                percent_from_fund=equity_total,
                row_num=None,
            ))

        # Check FX exposure (code 06 - מט"ח) against profile limit
        if max_fx is not None and fx_total > max_fx:
            logger_chk2b.warning(
                "Fund %d: Profile %s allows max %d%% FX but total is %.2f%%",
                fund_no, profile, max_fx, fx_total
            )
            exceptions.append(ExceptionRow(
                check_id="2ב",
                reason=f'פרופיל {profile} מתיר עד {max_fx}% מט"ח אך סה\"כ חשיפה = {fx_total:.2f}%',
                fund_no=fund_no,
                fund_name=rows[0].fund_name if rows else "",
                effective_code="06",
                percent_from_fund=fx_total,
                row_num=None,
            ))

    logger_chk2b.info("Exposure profile check completed: %d exceptions", len(exceptions))
    return exceptions


def check_3_combinations(
    disclosure_rows: list[DisclosureRow],
    mizrahi_fund_ids: set[int]
) -> dict[str, list[ExceptionRow]]:
    """Check 3א-3ח: Within-month code combinations and cross-checks."""
    logger_chk3.info("Starting within-month combinations check")

    # Group rows by fund
    fund_rows: dict[int, list[DisclosureRow]] = defaultdict(list)
    fund_names: dict[int, str] = {}
    for row in disclosure_rows:
        if row.fund_no in mizrahi_fund_ids:
            fund_rows[row.fund_no].append(row)
            if row.fund_name:
                fund_names[row.fund_no] = row.fund_name

    def has_code_prefix(codes: set[str], prefix: str) -> bool:
        """Check if any code starts with the given prefix."""
        return any(c.startswith(prefix) for c in codes)

    def get_codes_with_prefix(codes: set[str], prefix: str) -> set[str]:
        """Get all codes that start with the given prefix."""
        return {c for c in codes if c.startswith(prefix)}

    results: dict[str, list[ExceptionRow]] = {
        "3א": [],  # FX exposure
        "3ב": [],  # Bond exposure
        "3ג": [],  # Government bonds (shekel)
        "3ד": [],  # Government bonds (linked)
        "3ה": [],  # Government bonds (linked FX)
        "3ו": [],  # Corporate bonds (shekel)
        "3ז": [],  # Corporate bonds (linked)
        "3ח": [],  # Corporate bonds (linked FX)
    }

    for fund_no, rows in fund_rows.items():
        # Get all effective codes for this fund (most granular level)
        codes = set()
        # Get TIER 1 codes specifically (column C / level_1) for check 3ב
        tier1_codes = set()
        # Get TIER 2 codes specifically (column D / level_2) for check 3א
        tier2_codes = set()
        for row in rows:
            code = row.effective_code
            if code:
                codes.add(code)
            # Collect TIER 1 codes for check 3ב
            if row.level_1 and str(row.level_1).strip():
                tier1_codes.add(str(row.level_1).strip())
            # Collect TIER 2 codes for check 3א
            if row.level_2 and row.level_2.strip():
                tier2_codes.add(row.level_2.strip())

        fund_name = fund_names.get(fund_no, "")

        # Check 3א - FX Exposure
        # If 0102, 0302, or 0502 exists in TIER 2 -> 06 must exist in TIER 2
        # If 06 exists in TIER 2 -> at least one of 0102, 0302, 0502 must exist in TIER 2
        # NOTE: We check TIER 2 specifically because these codes have different meanings at deeper levels
        fx_codes_found = []
        if "0102" in tier2_codes:
            fx_codes_found.append("0102")
        if "0302" in tier2_codes:
            fx_codes_found.append("0302")
        if "0502" in tier2_codes:
            fx_codes_found.append("0502")
        fx_related = len(fx_codes_found) > 0
        has_06 = any(c.startswith("06") for c in tier2_codes)

        if fx_related and not has_06:
            found_codes_with_desc = ", ".join(code_desc(c) for c in fx_codes_found)
            logger_chk3.warning("Fund %d: Has FX-related codes (%s) but missing code 06", fund_no, found_codes_with_desc)
            results["3א"].append(ExceptionRow(
                check_id="3א",
                reason=f'נמצאו: {found_codes_with_desc}\nאך חסרים: {code_desc("06")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_06 and not fx_related:
            logger_chk3.warning("Fund %d: Has code 06 but missing FX-related codes", fund_no)
            results["3א"].append(ExceptionRow(
                check_id="3א",
                reason=f'נמצאו: {code_desc("06")}\nאך חסרים: {codes_desc(["0102", "0302", "0502"])}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ב - Bond Exposure
        # If 03 exists -> 07 and 08 must exist
        # If 07 or 08 exists -> 03 must exist
        # NOTE: We check TIER 1 (level_1) because 03, 07, 08 are TIER 1 codes per ISA spec
        has_03 = "3" in tier1_codes or "03" in tier1_codes
        has_07 = "7" in tier1_codes or "07" in tier1_codes
        has_08 = "8" in tier1_codes or "08" in tier1_codes

        if has_03 and not (has_07 and has_08):
            missing = []
            if not has_07:
                missing.append(code_desc("07"))
            if not has_08:
                missing.append(code_desc("08"))
            logger_chk3.warning("Fund %d: Has bonds (03) but missing %s", fund_no, ", ".join(missing))
            results["3ב"].append(ExceptionRow(
                check_id="3ב",
                reason=f'נמצאו: {code_desc("03")}\nאך חסרים: {", ".join(missing)}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif (has_07 or has_08) and not has_03:
            logger_chk3.warning("Fund %d: Has ratings/duration but missing bonds (03)", fund_no)
            results["3ב"].append(ExceptionRow(
                check_id="3ב",
                reason=f'נמצאו: {codes_desc(["07", "08"])}\nאך חסרים: {code_desc("03")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ג - Government Bonds (Index-Linked): 03010101 <-> 080202
        has_03010101 = "03010101" in codes
        has_080202 = "080202" in codes
        if has_03010101 and not has_080202:
            results["3ג"].append(ExceptionRow(
                check_id="3ג",
                reason=f'נמצאו: {code_desc("03010101")}\nאך חסרים: {code_desc("080202")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080202 and not has_03010101:
            results["3ג"].append(ExceptionRow(
                check_id="3ג",
                reason=f'נמצאו: {code_desc("080202")}\nאך חסרים: {code_desc("03010101")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ד - Government Bonds (Shekel/Non-Linked): 03010102 <-> 080201
        has_03010102 = "03010102" in codes
        has_080201 = "080201" in codes
        if has_03010102 and not has_080201:
            results["3ד"].append(ExceptionRow(
                check_id="3ד",
                reason=f'נמצאו: {code_desc("03010102")}\nאך חסרים: {code_desc("080201")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080201 and not has_03010102:
            results["3ד"].append(ExceptionRow(
                check_id="3ד",
                reason=f'נמצאו: {code_desc("080201")}\nאך חסרים: {code_desc("03010102")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ה - Government Bonds (Linked FX): 03010103 <-> 080203
        has_03010103 = "03010103" in codes
        has_080203 = "080203" in codes
        if has_03010103 and not has_080203:
            results["3ה"].append(ExceptionRow(
                check_id="3ה",
                reason=f'נמצאו: {code_desc("03010103")}\nאך חסרים: {code_desc("080203")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080203 and not has_03010103:
            results["3ה"].append(ExceptionRow(
                check_id="3ה",
                reason=f'נמצאו: {code_desc("080203")}\nאך חסרים: {code_desc("03010103")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ו - Corporate Bonds (Shekel): 03010202 or 03010203 <-> 080204
        corp_shekel_found = []
        if "03010202" in codes:
            corp_shekel_found.append("03010202")
        if "03010203" in codes:
            corp_shekel_found.append("03010203")
        has_corp_shekel = len(corp_shekel_found) > 0
        has_080204 = "080204" in codes
        if has_corp_shekel and not has_080204:
            found_codes_with_desc = ", ".join(code_desc(c) for c in corp_shekel_found)
            results["3ו"].append(ExceptionRow(
                check_id="3ו",
                reason=f'נמצאו: {found_codes_with_desc}\nאך חסרים: {code_desc("080204")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080204 and not has_corp_shekel:
            results["3ו"].append(ExceptionRow(
                check_id="3ו",
                reason=f'נמצאו: {code_desc("080204")}\nאך חסרים: {codes_desc(["03010202", "03010203"])}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ז - Corporate Bonds (Linked): 03010201 <-> 080205
        has_03010201 = "03010201" in codes
        has_080205 = "080205" in codes
        if has_03010201 and not has_080205:
            results["3ז"].append(ExceptionRow(
                check_id="3ז",
                reason=f'נמצאו: {code_desc("03010201")}\nאך חסרים: {code_desc("080205")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080205 and not has_03010201:
            results["3ז"].append(ExceptionRow(
                check_id="3ז",
                reason=f'נמצאו: {code_desc("080205")}\nאך חסרים: {code_desc("03010201")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

        # Check 3ח - Corporate Bonds (Linked FX): 03010204 <-> 080206
        has_03010204 = "03010204" in codes
        has_080206 = "080206" in codes
        if has_03010204 and not has_080206:
            results["3ח"].append(ExceptionRow(
                check_id="3ח",
                reason=f'נמצאו: {code_desc("03010204")}\nאך חסרים: {code_desc("080206")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))
        elif has_080206 and not has_03010204:
            results["3ח"].append(ExceptionRow(
                check_id="3ח",
                reason=f'נמצאו: {code_desc("080206")}\nאך חסרים: {code_desc("03010204")}',
                fund_no=fund_no,
                fund_name=fund_name,
            ))

    total_exceptions = sum(len(v) for v in results.values())
    logger_chk3.info("Combinations check completed: %d total exceptions", total_exceptions)
    return results


# -----------------------------
# Excel output helpers
# -----------------------------

def _rtl(ws) -> None:
    ws.sheet_view.rightToLeft = True


def _style_header(ws, row: int = 1) -> None:
    """Apply header styling to specified row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_cells(ws, start_row: int = 2) -> None:
    """Apply styling to all data cells."""
    wrap_cols: set[int] = set()
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(1, col_idx)
        header_text = str(header_cell.value).strip() if header_cell.value is not None else ""
        if header_text in WRAP_HEADERS:
            wrap_cols.add(col_idx)

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            wrap = cell.col_idx in wrap_cols
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=wrap)
            cell.border = THIN_BORDER


def _calculate_text_width(text: str, is_bold: bool = False) -> float:
    """Calculate approximate display width of text in Excel units."""
    if not text:
        return 0.0

    width = 0.0
    for char in text:
        if char == " ":
            width += 0.5
        elif '\u0590' <= char <= '\u05FF':
            width += 1.05
        elif char in '０１２３４５６７８９':
            width += 2.0
        elif ord(char) > 0x4E00:
            width += 2.0
        elif char in 'WMwm':
            width += 1.2
        elif char in 'il|!.,;:\'"':
            width += 0.6
        else:
            width += 1.0

    if is_bold:
        width *= 1.05

    return width


def _auto_fit_columns(ws, min_width: float = 8.0, max_width: float = 50.0, padding: float = 1.2) -> dict[str, float]:
    """Auto-fit column widths based on content."""
    column_widths: dict[str, float] = {}

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_width_found = 0.0

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)

            if cell.value is None:
                continue

            cell_text = str(cell.value)
            is_bold = cell.font.bold if cell.font else False
            is_header = (row_idx == 1)

            lines = cell_text.split('\n')
            for line in lines:
                line_width = _calculate_text_width(line, is_bold or is_header)
                if line_width > max_width_found:
                    max_width_found = line_width

        final_width = max_width_found + padding
        final_width = max(min_width, min(final_width, max_width))

        ws.column_dimensions[column_letter].width = final_width
        column_widths[column_letter] = final_width

    return column_widths


def _auto_fit_rows(ws, column_widths: dict[str, float] = None, line_height: float = 15.0, header_line_height: float = 18.0, min_height: float = 15.0, max_height: float = 120.0) -> None:
    """Auto-fit row heights based on content."""
    if column_widths is None:
        column_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            column_widths[col_letter] = width if width else 10.0

    for row_idx in range(1, ws.max_row + 1):
        max_lines_needed = 1
        is_header = (row_idx == 1)

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)

            if cell.value is None:
                continue

            cell_text = str(cell.value)
            col_letter = get_column_letter(col_idx)
            col_width = column_widths.get(col_letter, 10.0)
            has_wrap = cell.alignment.wrap_text if cell.alignment else False

            lines_in_cell = 0

            for line in cell_text.split('\n'):
                if not line:
                    lines_in_cell += 1
                    continue

                is_bold = cell.font.bold if cell.font else False
                text_width = _calculate_text_width(line, is_bold or is_header)
                available_width = max(col_width - 1.0, 6.0)

                if has_wrap and text_width > available_width:
                    wrapped_lines = int((text_width / available_width) + 0.99)
                    lines_in_cell += max(1, wrapped_lines)
                else:
                    lines_in_cell += 1

            if lines_in_cell > max_lines_needed:
                max_lines_needed = lines_in_cell

        if is_header:
            row_height = header_line_height * max_lines_needed
        else:
            row_height = line_height * max_lines_needed

        row_height = max(min_height, min(row_height, max_height))
        ws.row_dimensions[row_idx].height = row_height


def _set_font_calibri(ws) -> None:
    """Set Calibri font for all cells in worksheet."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name='Calibri',
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                    size=cell.font.size or 11
                )
            else:
                cell.font = Font(name='Calibri', size=11)


def _header(ws, headers: list[str], row: int = 1) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i)
        cell.value = h
    _style_header(ws, row)
    if row == 1:
        ws.freeze_panes = "A2"


# Global counter for unique table names
_table_counter = 0


def _add_table(ws, table_name_prefix: str = "Table") -> None:
    """Convert worksheet data range to a sortable/filterable Excel Table."""
    global _table_counter

    if ws.max_row < 2 or ws.max_column < 1:
        return  # No data to make into a table

    # Skip sheets with merged cells (tables cannot overlap merged cells)
    if ws.merged_cells.ranges:
        logger.debug("Skipping table for sheet %s: has merged cells", ws.title)
        return

    # Create unique table name (Excel requires unique names)
    _table_counter += 1
    table_name = f"{table_name_prefix}_{_table_counter}"
    # Sanitize table name - remove invalid characters, ensure starts with letter
    table_name = "".join(c if c.isalnum() or c == "_" else "_" for c in table_name)
    if table_name and not table_name[0].isalpha():
        table_name = "T" + table_name

    # Define the range for the table
    start_cell = "A1"
    end_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    table_range = f"{start_cell}:{end_cell}"

    # Create table with style
    table = Table(displayName=table_name, ref=table_range)

    # Use a built-in table style (medium blue to match header color)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style

    ws.add_table(table)


def _copy_code_index_sheet(wb, code_index_path: Path) -> Optional[Any]:
    """Copy the 'כל הקודים' sheet from code index file to output workbook.

    The source file should be table-formatted (headers in row 1, data from row 2).
    """
    if not code_index_path.exists():
        logger.warning("Code index file not found: %s", code_index_path)
        return None

    try:
        src_wb = openpyxl.load_workbook(code_index_path)
        src_ws = src_wb["כל הקודים"]

        # Create new sheet in destination workbook
        ws = wb.create_sheet("אינדקס קודים")
        _rtl(ws)

        # Copy all cell values and formatting (source is table-formatted, no merged cells)
        for row_idx in range(1, src_ws.max_row + 1):
            for col_idx in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(row_idx, col_idx)
                dst_cell = ws.cell(row_idx, col_idx)
                dst_cell.value = src_cell.value

                # Copy font, fill, alignment, border if they exist
                if src_cell.font:
                    dst_cell.font = Font(
                        name=src_cell.font.name or 'Calibri',
                        bold=src_cell.font.bold,
                        italic=src_cell.font.italic,
                        color=src_cell.font.color,
                        size=src_cell.font.size or 11
                    )
                if src_cell.fill and src_cell.fill.fill_type:
                    dst_cell.fill = PatternFill(
                        start_color=src_cell.fill.start_color.rgb if src_cell.fill.start_color else None,
                        end_color=src_cell.fill.end_color.rgb if src_cell.fill.end_color else None,
                        fill_type=src_cell.fill.fill_type
                    )
                if src_cell.alignment:
                    dst_cell.alignment = Alignment(
                        horizontal=src_cell.alignment.horizontal,
                        vertical=src_cell.alignment.vertical,
                        wrap_text=src_cell.alignment.wrap_text
                    )
                if src_cell.border:
                    dst_cell.border = Border(
                        left=src_cell.border.left,
                        right=src_cell.border.right,
                        top=src_cell.border.top,
                        bottom=src_cell.border.bottom
                    )

        # Copy column widths
        for col_letter, col_dim in src_ws.column_dimensions.items():
            if col_dim.width:
                ws.column_dimensions[col_letter].width = col_dim.width

        # Apply header styling to row 1
        _style_header(ws, 1)
        ws.freeze_panes = "A2"

        src_wb.close()
        logger.info("Added code index sheet from: %s", code_index_path)
        return ws

    except Exception as e:
        logger.warning("Failed to copy code index sheet: %s", e)
        return None


# -----------------------------
# Excel output
# -----------------------------

def write_output_xlsx(
    output_path: Path,
    *,
    report_month: str,
    manager_name: str,
    trustee_name: str,
    summary: dict[str, Any],
    exceptions_1a: list[ExceptionRow],
    exceptions_1b: list[ExceptionRow],
    exceptions_2a: list[ExceptionRow],
    exceptions_2b: list[ExceptionRow],
    exceptions_3: dict[str, list[ExceptionRow]],
    spec_file_path: Optional[Path] = None,
) -> None:
    """Write validation results to Excel workbook."""
    wb = openpyxl.Workbook()

    VALIDATION_COLS = ["טופל?", "שם הבודק"]

    # Sheet 1: Summary (סיכום)
    ws_sum = wb.active
    ws_sum.title = "סיכום"
    _rtl(ws_sum)
    _header(ws_sum, ["שדה", "ערך"])

    hebrew_month = _format_report_month_hebrew(report_month)
    summary_rows = [
        ("מנהל קרן", manager_name),
        ("נאמן", trustee_name),
        ("חודש נבדק", f"דוח גילוי נאות - {hebrew_month}"),
        ("מספר קרנות בדוח", summary.get("total_funds_in_report", "")),
        ("מספר קרנות בתחום (מזרחי)", summary.get("in_scope_funds", "")),
        ("מספר קרנות מחוץ לתחום", summary.get("out_of_scope_funds", "")),
        ("סה\"כ שורות בדוח", summary.get("total_rows", "")),
    ]

    rr = 2
    for k, v in summary_rows:
        ws_sum.cell(rr, 1).value = k
        ws_sum.cell(rr, 2).value = v
        rr += 1

    _style_cells(ws_sum)

    # Sheet 2: פירוט בדיקות (copy from spec file if available - supports CSV or XLSX)
    ws_spec = None
    if spec_file_path and spec_file_path.exists():
        ws_spec = wb.create_sheet("פירוט בדיקות")
        _rtl(ws_spec)

        suffix = spec_file_path.suffix.lower()
        if suffix in ('.xlsx', '.xls'):
            # Load from Excel file
            spec_wb = openpyxl.load_workbook(spec_file_path)
            spec_ws = spec_wb.active
            for row_idx in range(1, spec_ws.max_row + 1):
                row_data = []
                for col_idx in range(1, spec_ws.max_column + 1):
                    cell_value = spec_ws.cell(row_idx, col_idx).value
                    row_data.append(_clean_excel_string(cell_value) if cell_value else cell_value)
                ws_spec.append(row_data)
            spec_wb.close()
        else:
            # Load from CSV file
            with open(spec_file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for row in reader:
                    cleaned_row = [_clean_excel_string(cell) if cell else cell for cell in row]
                    ws_spec.append(cleaned_row)

        _style_header(ws_spec, 1)
        _style_cells(ws_spec)

    # Sheet 3: סטטוס בדיקות (Check Status)
    ws_checks = wb.create_sheet("סטטוס בדיקות")
    _rtl(ws_checks)
    ws_checks.append(["בדיקה", "תיאור", "סטטוס", "חריגות", "טופל?", "שם הבודק"])
    _style_header(ws_checks, 1)

    # Calculate counts for check 3
    count_3a = len(exceptions_3.get("3א", []))
    count_3b = len(exceptions_3.get("3ב", []))
    count_3c = len(exceptions_3.get("3ג", []))
    count_3d = len(exceptions_3.get("3ד", []))
    count_3e = len(exceptions_3.get("3ה", []))
    count_3f = len(exceptions_3.get("3ו", []))
    count_3g = len(exceptions_3.get("3ז", []))
    count_3h = len(exceptions_3.get("3ח", []))

    check_statuses = [
        ("בדיקה #1.א - שלמות קרנות", "הצלבה בין רשימת קרנות לדוח", len(exceptions_1a) == 0, len(exceptions_1a)),
        ("בדיקה #1.ב - תקינות תאריכים", "התאמת תאריך לחודש הדיווח", len(exceptions_1b) == 0, len(exceptions_1b)),
        ("בדיקה #2.א - סבירות מול דוח קודם", "השוואה לחודש קודם", len(exceptions_2a) == 0, len(exceptions_2a)),
        ("בדיקה #2.ב - סבירות מול מאפייני הקרן", "הצלבה מול פרופיל חשיפה", len(exceptions_2b) == 0, len(exceptions_2b)),
        ('בדיקה #3.א - חשיפה למט"ח', 'הצלבת קודי חשיפה למט"ח', count_3a == 0, count_3a),
        ('בדיקה #3.ב - חשיפה לאג"ח', 'הצלבת קודי אג"ח/דירוגים/מח"מ', count_3b == 0, count_3b),
        ('בדיקה #3.ג - אג"ח ממשלתי צמוד מדד', 'הצלבת 03010101 מול 080202', count_3c == 0, count_3c),
        ('בדיקה #3.ד - אג"ח ממשלתי שקלי', 'הצלבת 03010102 מול 080201', count_3d == 0, count_3d),
        ('בדיקה #3.ה - אג"ח ממשלתי צמוד מט"ח', 'הצלבת 03010103 מול 080203', count_3e == 0, count_3e),
        ('בדיקה #3.ו - אג"ח קונצרני שקלי', 'הצלבת 03010202/03010203 מול 080204', count_3f == 0, count_3f),
        ('בדיקה #3.ז - אג"ח קונצרני צמוד', 'הצלבת 03010201 מול 080205', count_3g == 0, count_3g),
        ('בדיקה #3.ח - אג"ח קונצרני צמוד מט"ח', 'הצלבת 03010204 מול 080206', count_3h == 0, count_3h),
    ]

    for row_idx, (name, description, passed, count) in enumerate(check_statuses, start=2):
        ws_checks.append([name, description, "✓ תקין" if passed else "✗ חריגה", count, "", ""])
        fill = PASS_FILL if passed else FAIL_FILL
        for col in range(1, 5):
            ws_checks.cell(row=row_idx, column=col).fill = fill

    _style_cells(ws_checks)

    # Sheet 4: אינדקס קודים (Code Index) - copy from k303_code_index.xlsx
    ws_codes = _copy_code_index_sheet(wb, CODE_INDEX_PATH)

    # Track optional sheets
    optional_sheets = []

    # Exception sheet helper
    def create_exception_sheet(sheet_name: str, exceptions: list[ExceptionRow], extra_columns: list[str] = None):
        if not exceptions:
            return None

        ws = wb.create_sheet(sheet_name)
        _rtl(ws)

        headers = ["בדיקה", "סיבה", "מספר קרן", "שם קרן"]
        if extra_columns:
            headers.extend(extra_columns)
        headers.extend(VALIDATION_COLS)

        _header(ws, headers)

        for ex in exceptions:
            # Clean string values to remove illegal XML characters
            row_data = [
                _clean_excel_string(ex.check_id),
                _clean_excel_string(ex.reason),
                ex.fund_no,
                _clean_excel_string(ex.fund_name)
            ]
            if extra_columns:
                for col in extra_columns:
                    if col == "קוד חשיפה":
                        row_data.append(_clean_excel_string(ex.effective_code))
                    elif col == "%מקרן":
                        row_data.append(ex.percent_from_fund)
                    elif col == "תאריך דוח":
                        row_data.append(_fmt_date(ex.report_date))
                    elif col == "שורה בקובץ":
                        row_data.append(ex.row_num)
                    elif col == "% קודם":
                        row_data.append(ex.extra_info.get("prev_pct"))
                    elif col == "הפרש":
                        row_data.append(ex.extra_info.get("delta"))
                    elif col == "סוג קרן":
                        row_data.append(_clean_excel_string(ex.extra_info.get("fund_type")))
                    elif col == "פירוט קוד":
                        # Get full hierarchical code description
                        code = ex.effective_code
                        code_def = get_full_code_description(code) if code else ""
                        row_data.append(_clean_excel_string(code_def))
            row_data.extend(["", ""])  # Validation columns
            ws.append(row_data)

        optional_sheets.append(ws)
        return ws

    # Create exception sheets
    # Note: Excel sheet names must be <= 31 characters
    create_exception_sheet("בדיקה 1.א - שלמות קרנות", exceptions_1a)
    create_exception_sheet("בדיקה 1.ב - תקינות תאריכים", exceptions_1b,
                          ["קוד חשיפה", "תאריך דוח", "שורה בקובץ"])
    create_exception_sheet("בדיקה 2.א - סבירות דוח קודם", exceptions_2a,
                          ["סוג קרן", "קוד חשיפה", "פירוט קוד", "%מקרן", "% קודם", "הפרש"])
    create_exception_sheet("בדיקה 2.ב - מאפייני קרן", exceptions_2b,
                          ["קוד חשיפה", "%מקרן", "שורה בקובץ"])

    # Create check 3 sheets
    # Note: Sheet names must be <= 31 chars and no invalid chars: " / \ [ ] : * ?
    check_3_names = [
        ("3א", 'בדיקה 3.א - חשיפה למט"ח'),
        ("3ב", 'בדיקה 3.ב - חשיפה לאג"ח'),
        ("3ג", 'בדיקה 3.ג - אג"ח ממשלתי צמוד'),
        ("3ד", 'בדיקה 3.ד - אג"ח ממשלתי שקלי'),
        ("3ה", 'בדיקה 3.ה - אג"ח ממשלתי מט"ח'),
        ("3ו", 'בדיקה 3.ו - אג"ח קונצרני שקלי'),
        ("3ז", 'בדיקה 3.ז - אג"ח קונצרני צמוד'),
        ("3ח", 'בדיקה 3.ח - אג"ח קונצרני מט"ח'),
    ]

    for check_key, sheet_name in check_3_names:
        create_exception_sheet(sheet_name, exceptions_3.get(check_key, []))

    # Apply styling to all optional sheets
    for ws in optional_sheets:
        _style_cells(ws)

    # Auto-fit all sheets
    all_sheets = [ws_sum, ws_checks] + optional_sheets
    if ws_spec:
        all_sheets.append(ws_spec)
    if ws_codes:
        all_sheets.append(ws_codes)

    for ws in all_sheets:
        _set_font_calibri(ws)
        column_widths = _auto_fit_columns(ws)
        _auto_fit_rows(ws, column_widths=column_widths)

    # Convert all sheets to sortable Excel Tables
    table_sheets = [ws_sum, ws_checks] + optional_sheets
    if ws_spec:
        table_sheets.append(ws_spec)
    if ws_codes:
        table_sheets.append(ws_codes)

    for ws in table_sheets:
        _add_table(ws, ws.title.replace(" ", "_").replace('"', ''))

    wb.save(output_path)
    wb.close()
    logger.info("Output saved to: %s", output_path)


# -----------------------------
# CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Validate K.303 Disclosure Report (Migdal Variant)"
    )
    parser.add_argument(
        "--mutual-funds-list",
        type=Path,
        required=True,
        help="Path to Mutual Funds List CSV"
    )
    parser.add_argument(
        "--current-report",
        type=Path,
        required=True,
        help="Path to current month disclosure report CSV"
    )
    parser.add_argument(
        "--previous-report",
        type=Path,
        required=True,
        help="Path to previous month disclosure report CSV"
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        required=True,
        help="Path for output Excel file"
    )
    parser.add_argument(
        "--email-json",
        type=Path,
        default=None,
        help="Path to write email JSON payload for n8n workflow (not yet implemented)"
    )
    parser.add_argument(
        "--report-month",
        type=str,
        required=True,
        help="Report month in YYYY-MM format (e.g., 2025-12)"
    )
    parser.add_argument(
        "--manager-name",
        type=str,
        required=True,
        help=f"Fund manager name. Options: {FUND_MANAGER_NAMES}"
    )
    parser.add_argument(
        "--trustee-name",
        type=str,
        default=MIZRAHI_TRUSTEE_NAME,
        help="Trustee name filter (default: Mizrahi)"
    )
    parser.add_argument(
        "--spec-file",
        type=Path,
        default=DEFAULT_SPEC_PATH,
        help="Path to K.303 checklist CSV or XLSX (for פירוט בדיקות sheet)"
    )

    args = parser.parse_args()

    # Setup logging
    setup_logging()

    logger.info("=" * 70)
    logger.info("Disclosure Report K.303 Validator")
    logger.info("=" * 70)

    # Load data
    logger.info("Loading mutual funds list...")
    all_funds = load_mutual_funds_csv(args.mutual_funds_list)
    in_scope_fund_ids = get_in_scope_fund_ids(all_funds, args.trustee_name, args.manager_name)

    logger.info("Loading current month disclosure report...")
    current_rows = load_disclosure_report(args.current_report)

    logger.info("Loading previous month disclosure report...")
    prev_rows = load_disclosure_report(args.previous_report)

    # Report month from argument
    report_month = args.report_month
    logger.info("Report month: %s", report_month)

    # Calculate summary stats
    funds_in_report = set(r.fund_no for r in current_rows if r.fund_no is not None)
    in_scope_funds = funds_in_report & in_scope_fund_ids
    out_of_scope_funds = funds_in_report - in_scope_fund_ids

    summary = {
        "total_funds_in_report": len(funds_in_report),
        "in_scope_funds": len(in_scope_funds),
        "out_of_scope_funds": len(out_of_scope_funds),
        "total_rows": len(current_rows),
    }

    # Run checks
    logger.info("Running validation checks...")

    exceptions_1a = check_1a_fund_completeness(current_rows, in_scope_fund_ids, all_funds)
    exceptions_1b = check_1b_report_month_validity(current_rows, report_month, in_scope_fund_ids)
    exceptions_2a = check_2a_prev_month_comparison(current_rows, prev_rows, in_scope_fund_ids, all_funds)
    exceptions_2b = check_2b_exposure_profile(current_rows, all_funds, in_scope_fund_ids)
    exceptions_3 = check_3_combinations(current_rows, in_scope_fund_ids)

    # Write output
    logger.info("Writing output Excel file...")
    write_output_xlsx(
        args.output_xlsx,
        report_month=report_month,
        manager_name=args.manager_name or "",
        trustee_name=args.trustee_name,
        summary=summary,
        exceptions_1a=exceptions_1a,
        exceptions_1b=exceptions_1b,
        exceptions_2a=exceptions_2a,
        exceptions_2b=exceptions_2b,
        exceptions_3=exceptions_3,
        spec_file_path=args.spec_file,
    )

    # Print summary
    logger.info("=" * 70)
    logger.info("Validation Summary")
    logger.info("=" * 70)
    logger.info("Report Month: %s", report_month)
    logger.info("Total Funds in Report: %d", summary["total_funds_in_report"])
    logger.info("In-Scope Funds (Mizrahi): %d", summary["in_scope_funds"])
    logger.info("Out-of-Scope Funds: %d", summary["out_of_scope_funds"])
    logger.info("")
    logger.info("Check Results:")
    logger.info("  1א Fund Completeness: %d exceptions", len(exceptions_1a))
    logger.info("  1ב Date Validity: %d exceptions", len(exceptions_1b))
    logger.info("  2א Previous Month Comparison: %d exceptions", len(exceptions_2a))
    logger.info("  2ב Exposure Profile: %d exceptions", len(exceptions_2b))
    for check_key in ["3א", "3ב", "3ג", "3ד", "3ה", "3ו", "3ז", "3ח"]:
        count = len(exceptions_3.get(check_key, []))
        logger.info("  %s: %d exceptions", check_key, count)
    logger.info("")
    logger.info("Output saved to: %s", args.output_xlsx)
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
